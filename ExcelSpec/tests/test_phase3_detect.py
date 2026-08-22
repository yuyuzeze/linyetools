"""Phase 3 tests: RegionDetector, RegionRouter, Semantic Profiles, modes."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl.utils import get_column_letter

from excelspec.detect.detector import RegionDetector, detect_sheet
from excelspec.detect.models import CandidateRegionType, CellBounds
from excelspec.detect.router import RegionRouter
from excelspec.ingest.sparse_model import SparseCell, SparseSheet, SparseWorkbookIR
from excelspec.models.document_ir import AssetIR, AssetType, SourceRef, StyleIR
from excelspec.profile.loader import ProfileValidationError, parse_profile
from excelspec.profile.model import SemanticProfile
from excelspec.profile.normalize import normalize_header

FIXTURES = Path(__file__).resolve().parent / "fixtures"


def _cell(row: int, col: int, value, *, style_id=None, formula=None) -> SparseCell:
    coord = f"{get_column_letter(col)}{row}"
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        data_type = "s"
    else:
        data_type = "n"
    display = None if value is None else str(value)
    return SparseCell(
        row=row,
        column=col,
        coordinate=coord,
        raw_value=value,
        display_value=display,
        data_type="f" if formula else data_type,
        formula=formula,
        cached_value=value,
        style_id=style_id,
    )


def _sheet(name: str, values: dict, *, merges=None, assets=None) -> SparseSheet:
    cells: dict[tuple[int, int], SparseCell] = {}
    for (row, col), value in values.items():
        cells[(row, col)] = value if isinstance(value, SparseCell) else _cell(row, col, value)
    rows = [r for r, _ in cells] or [1]
    cols = [c for _, c in cells] or [1]
    sheet = SparseSheet(
        name=name,
        sheet_id="sheet-1",
        index=0,
        cells=cells,
        content_bounds=(min(rows), min(cols), max(rows), max(cols)),
        assets=assets or [],
    )
    for master, span in (merges or {}).items():
        mr, mc = master
        rows_span, cols_span = span
        sheet.merge_spans[(mr, mc)] = (rows_span, cols_span)
        for r in range(mr, mr + rows_span):
            for c in range(mc, mc + cols_span):
                sheet.merge_members[(r, c)] = (mr, mc)
    return sheet


def _workbook(sheet: SparseSheet, styles=None) -> SparseWorkbookIR:
    return SparseWorkbookIR(
        path=str(FIXTURES / "workbooks" / "api-spec.xlsx"),
        sheets=[sheet],
        styles=styles or {},
    )


HEADER_STYLE = {1: StyleIR(font={"bold": True})}


def _table_values(start_row: int, header, rows) -> dict:
    values = {}
    for col, label in enumerate(header, start=1):
        values[(start_row, col)] = _cell(start_row, col, label, style_id=1)
    for r, row in enumerate(rows, start=start_row + 1):
        for col, value in enumerate(row, start=1):
            if value is not None:
                values[(r, col)] = value
    return values


# --------------------------------------------------------------------------- #
# RegionDetector
# --------------------------------------------------------------------------- #

class RegionDetectorTests(unittest.TestCase):
    def test_two_tables_separated_by_blank_rows(self) -> None:
        values = {}
        values.update(_table_values(1, ["A", "B"], [["1", "2"], ["3", "4"]]))
        # blank rows 4,5 then a second table at row 6
        values.update(_table_values(6, ["C", "D"], [["5", "6"], ["7", "8"]]))
        sheet = _sheet("S", values)
        cands = [c for c in detect_sheet(sheet, HEADER_STYLE) if c.region_type == CandidateRegionType.TABLE]
        self.assertEqual(2, len(cands))

    def test_two_tables_separated_by_blank_columns(self) -> None:
        values = {}
        values.update(_table_values(1, ["A", "B"], [["1", "2"], ["3", "4"]]))
        # second table starting at column 5 (blank columns 3,4)
        for col, label in enumerate(["C", "D"], start=5):
            values[(1, col)] = _cell(1, col, label, style_id=1)
        values[(2, 5)] = "9"; values[(2, 6)] = "10"
        values[(3, 5)] = "11"; values[(3, 6)] = "12"
        sheet = _sheet("S", values)
        tables = [c for c in detect_sheet(sheet, HEADER_STYLE) if c.region_type == CandidateRegionType.TABLE]
        self.assertEqual(2, len(tables))

    def test_internal_blank_row_does_not_split_table(self) -> None:
        values = {}
        values.update(_table_values(1, ["A", "B"], [["1", "2"]]))
        # row 3 blank, row 4 continues the same table
        values[(4, 1)] = "3"; values[(4, 2)] = "4"
        sheet = _sheet("S", values)
        tables = [c for c in detect_sheet(sheet, HEADER_STYLE) if c.region_type == CandidateRegionType.TABLE]
        self.assertEqual(1, len(tables))
        self.assertGreaterEqual(tables[0].bounds.max_row, 4)

    def test_title_row_peeled_from_table(self) -> None:
        values = {(1, 1): _cell(1, 1, "画面項目一覧", style_id=1)}
        values.update(_table_values(2, ["項目名", "型"], [["a", "b"], ["c", "d"]]))
        sheet = _sheet("S", values)
        tables = [c for c in detect_sheet(sheet, HEADER_STYLE) if c.region_type == CandidateRegionType.TABLE]
        self.assertEqual(1, len(tables))
        self.assertEqual("画面項目一覧", tables[0].title)
        self.assertEqual(2, tables[0].bounds.min_row)  # title row excluded

    def test_key_value_region_detected(self) -> None:
        values = {
            (1, 1): "画面ID", (1, 2): "SCR-001",
            (2, 1): "画面名", (2, 2): "一覧",
            (3, 1): "作成者", (3, 2): "山田",
        }
        sheet = _sheet("S", values)
        cands = detect_sheet(sheet, HEADER_STYLE)
        kinds = {c.region_type for c in cands}
        self.assertIn(CandidateRegionType.KEY_VALUE, kinds)

    def test_image_and_shape_anchors_become_regions(self) -> None:
        assets = [
            AssetIR(asset_id="img1", asset_type=AssetType.IMAGE, uri="x.png", anchor="B10"),
            AssetIR(asset_id="shp1", asset_type=AssetType.SHAPE, uri="", anchor="B20", description="ボタン"),
        ]
        sheet = _sheet("S", {(1, 1): "title"}, assets=assets)
        cands = detect_sheet(sheet, HEADER_STYLE)
        kinds = {c.region_type for c in cands}
        self.assertIn(CandidateRegionType.IMAGE, kinds)
        self.assertIn(CandidateRegionType.SHAPE, kinds)

    def test_low_confidence_region_emits_diagnostic(self) -> None:
        sheet = _sheet("S", {(1, 1): "loneword"})
        cands = detect_sheet(sheet, HEADER_STYLE)
        codes = {d.code for c in cands for d in c.diagnostics}
        self.assertIn("detect.low_confidence_region", codes)

    def test_every_value_cell_is_covered(self) -> None:
        values = {}
        values.update(_table_values(1, ["A", "B"], [["1", "2"]]))
        values[(10, 5)] = "orphan"
        sheet = _sheet("S", values)
        cands = detect_sheet(sheet, HEADER_STYLE)
        covered = {c for cand in cands for c in cand.source_cells}
        all_value = {cell.coordinate for cell in sheet.cells.values() if cell.raw_value is not None}
        self.assertTrue(all_value.issubset(covered))

    def test_detection_does_not_materialize_grid(self) -> None:
        sheet = _sheet("S", _table_values(1, ["A", "B"], [["1", "2"]]))
        with mock.patch("excelspec.ingest.adapter.materialize_region") as spy:
            detect_sheet(sheet, HEADER_STYLE)
        spy.assert_not_called()

    def test_features_are_explainable(self) -> None:
        sheet = _sheet("S", _table_values(1, ["A", "B"], [["1", "2"], ["3", "4"]]))
        table = next(c for c in detect_sheet(sheet, HEADER_STYLE) if c.region_type == CandidateRegionType.TABLE)
        self.assertTrue(table.detection_method)
        for key in ("nonempty_cell_count", "density", "repeated_row_score", "header_score"):
            self.assertIn(key, table.features)


# --------------------------------------------------------------------------- #
# RegionRouter
# --------------------------------------------------------------------------- #

class RegionRouterTests(unittest.TestCase):
    def _route(self, sheet, styles=None):
        workbook = _workbook(sheet, styles)
        router = RegionRouter(workbook)
        return router, [router.route(sheet, c) for c in detect_sheet(sheet, styles or HEADER_STYLE)]

    def test_table_region_keeps_columns_and_header(self) -> None:
        sheet = _sheet("S", _table_values(1, ["A", "B", "C"], [["1", "2", "3"]]))
        _, regions = self._route(sheet)
        table_region = next(r for r in regions if r.region_type.value == "table")
        table = table_region.tables[0]
        self.assertGreaterEqual(table.header_rows, 1)
        self.assertEqual({1, 2, 3}, {c.column for c in table.cells})

    def test_missing_cell_not_shifted(self) -> None:
        values = _table_values(1, ["A", "B", "C"], [["1", None, "3"]])
        sheet = _sheet("S", values)
        _, regions = self._route(sheet)
        table = next(r for r in regions if r.region_type.value == "table").tables[0]
        by_coord = {c.coordinate: c for c in table.cells}
        self.assertIn("B2", by_coord)
        self.assertIsNone(by_coord["B2"].raw_value)
        self.assertEqual("3", by_coord["C2"].raw_value)

    def test_key_value_region_values(self) -> None:
        values = {
            (1, 1): "画面ID", (1, 2): "SCR-001",
            (2, 1): "画面名", (2, 2): "一覧",
        }
        sheet = _sheet("S", values)
        _, regions = self._route(sheet)
        kv = next((r for r in regions if r.region_type.value == "key_value"), None)
        self.assertIsNotNone(kv)
        self.assertEqual("SCR-001", kv.values.get("画面ID"))

    def test_image_region_has_asset_no_cells(self) -> None:
        assets = [AssetIR(asset_id="img1", asset_type=AssetType.IMAGE, uri="x.png", anchor="B10")]
        sheet = _sheet("S", {(1, 1): "t"}, assets=assets)
        _, regions = self._route(sheet)
        image = next(r for r in regions if r.region_type.value == "image")
        self.assertEqual(["img1"], image.asset_ids)
        self.assertEqual([], image.tables)

    def test_layout_fast_mode_no_com_but_marked_visual(self) -> None:
        # A short text block with a drawing anchored over it -> layout (visual),
        # and fast routing keeps structure only (no Excel COM launched).
        from excelspec.render import excel_capture

        values = {(1, 1): "画面レイアウト", (2, 1): "説明文"}
        assets = [AssetIR(asset_id="img1", asset_type=AssetType.IMAGE, uri="x.png", anchor="A2")]
        sheet = _sheet("S", values, assets=assets)
        before = excel_capture.capture_launch_count()
        _, regions = self._route(sheet)
        visual = [r for r in regions if r.metadata.get("visual")]
        self.assertTrue(visual)
        self.assertEqual(before, excel_capture.capture_launch_count())


# --------------------------------------------------------------------------- #
# Semantic Profile
# --------------------------------------------------------------------------- #

class ProfileTests(unittest.TestCase):
    def _profile(self) -> SemanticProfile:
        return parse_profile(
            {
                "schema_version": "1",
                "profile_id": "screen-design",
                "document_type": "screen_design",
                "match": {
                    "filename_patterns": ["^SCR-"],
                    "sheet_aliases": {"fields": ["画面項目", "項目一覧"]},
                },
                "fields": {
                    "field_name": {"aliases": ["項目名", "画面項目名"]},
                    "required": {"aliases": ["必須"]},
                },
                "validation": [{"concept": "field_name", "required": True}],
                "overrides": [{"sheet_alias": "fields", "ignore": ["A1:Z1"]}],
            }
        )

    def test_rejects_legacy_coordinate_fields(self) -> None:
        with self.assertRaises(ProfileValidationError) as ctx:
            parse_profile(
                {
                    "profile_id": "x",
                    "document_type": "y",
                    "fields": {"f": {"aliases": ["a"], "range": "A1:B2"}},
                }
            )
        self.assertTrue(any("range" in e for e in ctx.exception.errors))

    def test_rejects_locator_and_offsets(self) -> None:
        for bad in ("locator", "width", "height", "row_offset", "anchor_text", "repeat_anchor"):
            with self.assertRaises(ProfileValidationError):
                parse_profile(
                    {"profile_id": "x", "document_type": "y", "match": {bad: 1}}
                )

    def test_overrides_may_contain_ignore(self) -> None:
        profile = self._profile()  # ignore under overrides must NOT be rejected
        self.assertEqual(["A1:Z1"], profile.overrides[0].ignore)

    def test_overrides_may_contain_visual_range(self) -> None:
        profile = parse_profile(
            {
                "profile_id": "p",
                "document_type": "d",
                "overrides": [{"sheet": "Layout", "visual_range": "A5:Z40"}],
            }
        )
        self.assertEqual("A5:Z40", profile.overrides[0].visual_range)

    def test_invalid_visual_range_is_rejected(self) -> None:
        with self.assertRaises(ProfileValidationError):
            parse_profile(
                {
                    "profile_id": "p",
                    "document_type": "d",
                    "overrides": [{"sheet": "Layout", "visual_range": "not-a-range"}],
                }
            )

    def test_normalize_header_nfkc_and_case(self) -> None:
        # full-width spaces trimmed/normalized; casing folded
        self.assertEqual(normalize_header("　項目名　"), "項目名")
        self.assertEqual(normalize_header("ＦＩＥＬＤ"), "field")
        self.assertEqual(normalize_header("Field　Name"), "field name")

    def test_sheet_role_via_alias(self) -> None:
        self.assertEqual("fields", self._profile().sheet_role("画面項目"))
        self.assertIsNone(self._profile().sheet_role("無関係"))

    def test_field_concept_exact_alias(self) -> None:
        concepts, method = self._profile().match_field("項目名")
        self.assertEqual(["field_name"], concepts)
        self.assertEqual("exact_alias", method)

    def test_unmatched_header_preserved(self) -> None:
        concepts, method = self._profile().match_field("謎の列")
        self.assertEqual([], concepts)
        self.assertEqual("unmatched", method)

    def test_ambiguous_header_reported_in_enrichment(self) -> None:
        from excelspec.profile.enrich import enrich_regions
        from excelspec.models.document_ir import RegionIR, RegionType, TableIR

        profile = parse_profile(
            {
                "profile_id": "p", "document_type": "d",
                "fields": {
                    "a": {"aliases": ["共通"]},
                    "b": {"aliases": ["共通"]},
                },
            }
        )
        region = RegionIR(
            region_id="t", region_type=RegionType.TABLE,
            tables=[TableIR(table_id="t", cells=[], metadata={"header_labels": {"A": "共通"}})],
        )
        diagnostics = enrich_regions(profile, "Sheet", [region])
        self.assertIn("profile.ambiguous_header", {d.code for d in diagnostics})


# --------------------------------------------------------------------------- #
# Modes (integration)
# --------------------------------------------------------------------------- #

class ModeTests(unittest.TestCase):
    WORKBOOK = FIXTURES / "workbooks" / "screen-design.xlsx"

    def test_fast_mode_zero_config_no_excel(self) -> None:
        from excelspec.pipeline import run_pipeline
        from excelspec.render import excel_capture

        before = excel_capture.capture_launch_count()
        with tempfile.TemporaryDirectory() as directory:
            result = run_pipeline(self.WORKBOOK, mode="fast", asset_dir=Path(directory))
        self.assertEqual("zero-config:fast", result.match.mode)
        self.assertFalse(result.validation.failed())
        self.assertEqual(before, excel_capture.capture_launch_count())
        # every sheet produced at least one region
        self.assertTrue(all(sheet.regions for sheet in result.document.sheets))

    def test_profile_mode_applies_semantics(self) -> None:
        from excelspec.pipeline import run_pipeline

        with tempfile.TemporaryDirectory() as directory:
            result = run_pipeline(
                self.WORKBOOK,
                mode="fast",
                profile=Path("profiles/screen-design.yaml"),
                asset_dir=Path(directory),
            )
        self.assertTrue(result.match.mode.startswith("profile:"))
        semantics = {
            semantic
            for sheet in result.document.sheets
            for region in sheet.regions
            for table in region.tables
            for semantic in table.column_semantics.values()
        }
        self.assertIn("field_name", semantics)

    def test_template_takes_precedence_over_mode(self) -> None:
        from excelspec.pipeline import run_pipeline

        with tempfile.TemporaryDirectory() as directory:
            result = run_pipeline(
                self.WORKBOOK,
                template=FIXTURES / "templates" / "screen-design.yaml",
                mode="fast",  # ignored because a legacy template is present
                asset_dir=Path(directory),
            )
        self.assertEqual("fixture-screen-design", result.document.template_id)

    def _layout_workbook(self, directory: str) -> SparseWorkbookIR:
        assets = [
            AssetIR(
                asset_id="shape1",
                asset_type=AssetType.SHAPE,
                uri="ooxml://drawing#shape-1",
                anchor="A2:D8",
                description="flow",
            )
        ]
        sheet = _sheet("Layout", {(1, 1): "画面レイアウト", (2, 1): "説明"}, assets=assets)
        workbook = _workbook(sheet)
        workbook.metadata["asset_directory"] = directory
        return workbook

    def test_visual_mode_reuses_single_session_and_captures(self) -> None:
        from excelspec.detect.assemble import assemble_document
        import excelspec.render as render

        launches = {"n": 0}

        class FakeSession:
            def __init__(self, path):
                launches["n"] += 1

            def __enter__(self):
                return self

            def __exit__(self, *exc):
                return False

            def capture(self, sheet_name, a1, destination):
                path = Path(destination)
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(b"png")
                return path

        with tempfile.TemporaryDirectory() as directory:
            workbook = self._layout_workbook(directory)
            with mock.patch.object(render, "ExcelCaptureSession", FakeSession):
                document, _ = assemble_document(workbook, mode="visual")
        self.assertEqual(1, launches["n"])
        screenshots = [
            a for s in document.sheets for a in s.assets if a.asset_type == AssetType.SCREENSHOT
        ]
        self.assertTrue(screenshots)
        self.assertEqual("A1:D8", screenshots[0].anchor)
        self.assertEqual("shape_anchor_union", screenshots[0].metadata["capture_strategy"])

    def test_visual_mode_reuses_embedded_image_without_excel(self) -> None:
        from excelspec.detect.assemble import assemble_document
        import excelspec.render as render

        class MustNotStart:
            def __init__(self, path):
                raise AssertionError("embedded image must not start Excel")

        with tempfile.TemporaryDirectory() as directory:
            assets = [
                AssetIR(
                    asset_id="img1",
                    asset_type=AssetType.IMAGE,
                    uri=str(Path(directory) / "layout.png"),
                    anchor="A2",
                    description="layout",
                )
            ]
            sheet = _sheet(
                "Layout",
                {(1, 1): "画面レイアウト", (2, 1): "説明"},
                assets=assets,
            )
            workbook = _workbook(sheet)
            workbook.metadata["asset_directory"] = directory
            with mock.patch.object(render, "ExcelCaptureSession", MustNotStart):
                document, _ = assemble_document(workbook, mode="visual")
        visual = next(
            region
            for sheet in document.sheets
            for region in sheet.regions
            if region.metadata.get("visual")
        )
        self.assertEqual(
            "reuse_embedded_asset", visual.metadata["screenshot_strategy"]
        )
        self.assertEqual(["img1"], visual.metadata["visual_source_asset_ids"])

    def test_visual_range_profile_override_wins(self) -> None:
        from excelspec.detect.assemble import assemble_document
        import excelspec.render as render

        captures: list[str] = []

        class FakeSession:
            def __init__(self, path):
                pass

            def __enter__(self):
                return self

            def __exit__(self, *exc):
                return False

            def capture(self, sheet_name, a1, destination):
                captures.append(a1)
                path = Path(destination)
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(b"png")
                return path

        profile = parse_profile(
            {
                "profile_id": "p",
                "document_type": "d",
                "overrides": [{"sheet": "Layout", "visual_range": "B5:Z40"}],
            }
        )
        with tempfile.TemporaryDirectory() as directory:
            workbook = self._layout_workbook(directory)
            with mock.patch.object(render, "ExcelCaptureSession", FakeSession):
                document, _ = assemble_document(workbook, mode="visual", profile=profile)
        self.assertEqual(["B5:Z40"], captures)
        screenshot = next(
            asset
            for sheet in document.sheets
            for asset in sheet.assets
            if asset.asset_type == AssetType.SCREENSHOT
        )
        self.assertEqual("profile_override", screenshot.metadata["capture_strategy"])

    def test_border_drawn_layout_uses_detected_box(self) -> None:
        from excelspec.detect.assemble import assemble_document
        import excelspec.render as render

        captures: list[str] = []

        class FakeSession:
            def __init__(self, path):
                pass

            def __enter__(self):
                return self

            def __exit__(self, *exc):
                return False

            def capture(self, sheet_name, a1, destination):
                captures.append(a1)
                path = Path(destination)
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_bytes(b"png")
                return path

        sheet = _sheet("Layout", {})
        sheet.style_only = {(row, col): 1 for row in range(5, 10) for col in range(1, 5)}
        workbook = _workbook(sheet, {1: StyleIR(border={"left": {"style": "thin"}})})
        with tempfile.TemporaryDirectory() as directory:
            workbook.metadata["asset_directory"] = directory
            with mock.patch.object(render, "ExcelCaptureSession", FakeSession):
                document, _ = assemble_document(workbook, mode="visual")
        self.assertEqual(["A5:D9"], captures)
        visual = next(
            region
            for output_sheet in document.sheets
            for region in output_sheet.regions
            if region.metadata.get("visual")
        )
        self.assertEqual("detected_visual_bounds", visual.metadata["screenshot_strategy"])

    def test_tiny_visual_range_is_skipped_without_excel(self) -> None:
        from excelspec.detect.assemble import assemble_document
        import excelspec.render as render

        class MustNotStart:
            def __init__(self, path):
                raise AssertionError("tiny layout must not start Excel")

        sheet = _sheet(
            "Layout",
            {(1, 1): _cell(1, 1, "标题", style_id=1), (2, 1): "说明"},
            merges={(1, 1): (1, 2)},
        )
        workbook = _workbook(sheet, {1: StyleIR(font={"bold": True})})
        with mock.patch.object(render, "ExcelCaptureSession", MustNotStart):
            document, _ = assemble_document(workbook, mode="visual")
        visual_regions = [
            region
            for output_sheet in document.sheets
            for region in output_sheet.regions
            if region.metadata.get("visual")
        ]
        # The detector may conservatively keep this as text; if it marks it as
        # visual, the capture policy must still refuse the tiny strip.
        for region in visual_regions:
            self.assertEqual("skip_tiny_visual_range", region.metadata["screenshot_strategy"])

    def test_visual_mode_screenshot_failure_keeps_structure(self) -> None:
        from excelspec.detect.assemble import assemble_document
        import excelspec.render as render

        class BoomSession:
            def __init__(self, path):
                pass

            def __enter__(self):
                return self

            def __exit__(self, *exc):
                return False

            def capture(self, *a, **k):
                raise RuntimeError("no excel here")

        with tempfile.TemporaryDirectory() as directory:
            workbook = self._layout_workbook(directory)
            with mock.patch.object(render, "ExcelCaptureSession", BoomSession):
                document, _ = assemble_document(workbook, mode="visual")
        # structured content survives a screenshot failure
        self.assertTrue(all(sheet.regions for sheet in document.sheets))
        layout = next(
            r for s in document.sheets for r in s.regions if r.metadata.get("visual")
        )
        diag_codes = {d["code"] for d in layout.metadata.get("diagnostics", [])}
        self.assertIn("route.screenshot_failed", diag_codes)


class DefaultBehaviorTests(unittest.TestCase):
    WORKBOOK = FIXTURES / "workbooks" / "screen-design.xlsx"
    TEMPLATE = FIXTURES / "templates" / "screen-design.yaml"

    def test_default_is_zero_config_fast(self) -> None:
        from excelspec.pipeline import run_pipeline

        with tempfile.TemporaryDirectory() as directory:
            result = run_pipeline(self.WORKBOOK, asset_dir=Path(directory))
        self.assertEqual("zero-config:fast", result.match.mode)
        self.assertEqual("zero-config", result.processing["processing_mode"])
        self.assertEqual("fast", result.processing["detection_mode"])

    def test_default_equivalent_to_mode_fast(self) -> None:
        from excelspec.pipeline import run_pipeline

        def _structure(result):
            return [
                (sheet.name, [(r.region_id, r.region_type.value, r.source.range) for r in sheet.regions])
                for sheet in result.document.sheets
            ]

        with tempfile.TemporaryDirectory() as d1, tempfile.TemporaryDirectory() as d2:
            default = run_pipeline(self.WORKBOOK, asset_dir=Path(d1))
            explicit = run_pipeline(self.WORKBOOK, mode="fast", asset_dir=Path(d2))
        self.assertEqual(_structure(default), _structure(explicit))

    def test_default_does_not_load_bundled_templates(self) -> None:
        from excelspec import pipeline

        with tempfile.TemporaryDirectory() as directory:
            with mock.patch.object(pipeline, "load_template_candidates") as spy:
                pipeline.run_pipeline(self.WORKBOOK, asset_dir=Path(directory))
        spy.assert_not_called()

    def test_default_does_not_launch_excel(self) -> None:
        from excelspec.pipeline import run_pipeline
        from excelspec.render import excel_capture

        before = excel_capture.capture_launch_count()
        with tempfile.TemporaryDirectory() as directory:
            run_pipeline(self.WORKBOOK, asset_dir=Path(directory))
        self.assertEqual(before, excel_capture.capture_launch_count())

    def test_explicit_template_uses_legacy(self) -> None:
        from excelspec.pipeline import run_pipeline

        with tempfile.TemporaryDirectory() as directory:
            result = run_pipeline(
                self.WORKBOOK, template=self.TEMPLATE, asset_dir=Path(directory)
            )
        self.assertEqual("legacy-template", result.processing["processing_mode"])
        self.assertEqual("fixture-screen-design", result.processing["legacy_template_id"])

    def test_auto_legacy_template_restores_automatch(self) -> None:
        from excelspec import pipeline

        with tempfile.TemporaryDirectory() as directory:
            with mock.patch.object(
                pipeline, "load_template_candidates", return_value=([], False)
            ) as spy:
                result = pipeline.run_pipeline(
                    self.WORKBOOK,
                    auto_legacy_template=True,
                    asset_dir=Path(directory),
                )
        spy.assert_called_once()
        self.assertEqual("legacy-template", result.processing["processing_mode"])

    def test_processing_metadata_fields(self) -> None:
        from excelspec.pipeline import run_pipeline

        with tempfile.TemporaryDirectory() as directory:
            result = run_pipeline(
                self.WORKBOOK,
                mode="fast",
                profile=Path("profiles/screen-design.yaml"),
                asset_dir=Path(directory),
            )
        for key in (
            "processing_mode",
            "detection_mode",
            "profile_id",
            "legacy_template_id",
            "ingest_engine",
        ):
            self.assertIn(key, result.processing)
        self.assertEqual("screen-design", result.processing["profile_id"])
        self.assertEqual("sparse", result.processing["ingest_engine"])


if __name__ == "__main__":
    unittest.main()
