"""Phase 1 optimization tests: schema mode, SheetIndex, COM session reuse."""

from __future__ import annotations

import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from excelspec.index import SheetIndex, normalize_text
from excelspec.models.document_ir import (
    CellIR,
    DocumentIR,
    RegionIR,
    RegionType,
    SheetIR,
    StyleIR,
    TableIR,
)
from excelspec.pipeline import run_pipeline
from excelspec.render.excel_capture import ExcelCaptureSession, capture_launch_count
from excelspec.schemas import get_validator
from excelspec.validate import validate_document


def _tiny_document() -> DocumentIR:
    cells = [
        CellIR("A1", 1, 1, "ID", "ID", style=StyleIR(font={"bold": True})),
        CellIR("B1", 1, 2, "名称", "名称", style=StyleIR(font={"bold": True})),
        CellIR("A2", 2, 1, 1, "1"),
        CellIR("B2", 2, 2, "  項目  A ", "  項目  A "),
    ]
    return DocumentIR(
        document_id="doc",
        title="Doc",
        sheets=[
            SheetIR(
                sheet_id="sheet-1",
                name="Sheet1",
                index=0,
                regions=[
                    RegionIR(
                        region_id="raw-grid",
                        region_type=RegionType.FREEFORM,
                        tables=[TableIR(table_id="raw-grid", cells=cells)],
                    )
                ],
            )
        ],
    )


class SchemaModeTests(unittest.TestCase):
    def test_validator_is_cached(self) -> None:
        self.assertIs(get_validator("document-ir"), get_validator("document-ir"))

    def test_fast_mode_skips_full_schema(self) -> None:
        document = _tiny_document()
        with mock.patch(
            "excelspec.validate.core.validate_ir_schema", return_value=[]
        ) as spy:
            validate_document(document, strict_schema=False)
        spy.assert_not_called()

    def test_strict_mode_runs_full_schema(self) -> None:
        document = _tiny_document()
        with mock.patch(
            "excelspec.validate.core.validate_ir_schema", return_value=[]
        ) as spy:
            validate_document(document, strict_schema=True)
        spy.assert_called_once()

    def test_run_pipeline_xlsx_default_skips_full_schema(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook = Path(directory) / "w.xlsx"
            book = Workbook()
            book.active.title = "S"
            book.active.append(["A", "B"])
            book.save(workbook)

            with mock.patch(
                "excelspec.validate.core.validate_ir_schema", return_value=[]
            ) as spy:
                run_pipeline(workbook, asset_dir=Path(directory) / "assets")
            spy.assert_not_called()

            with mock.patch(
                "excelspec.validate.core.validate_ir_schema", return_value=[]
            ) as spy:
                run_pipeline(
                    workbook,
                    asset_dir=Path(directory) / "assets2",
                    strict_schema=True,
                )
            spy.assert_called_once()


class SheetIndexTests(unittest.TestCase):
    def test_index_builds_all_lookup_tables(self) -> None:
        index = SheetIndex.from_sheet(_tiny_document().sheets[0])
        self.assertEqual(4, index.cell_count)
        self.assertEqual("ID", index.cell("A1").raw_value)
        self.assertEqual("名称", index.at(1, 2).raw_value)
        self.assertEqual((1, 1, 2, 2), index.bounds())
        # Two bold header cells share one style id.
        bold_id = index.style_ids["A1"]
        self.assertEqual(bold_id, index.style_ids["B1"])
        self.assertEqual(2, len(index.by_style_id[bold_id]))
        self.assertNotEqual(bold_id, index.style_ids["A2"])

    def test_normalized_text_lookup_is_whitespace_tolerant(self) -> None:
        index = SheetIndex.from_sheet(_tiny_document().sheets[0])
        self.assertEqual(normalize_text("  項目   A "), "項目 a")
        found = index.find_text("項目 A")
        self.assertEqual(["B2"], [cell.coordinate for cell in found])


class ExcelCaptureSessionTests(unittest.TestCase):
    def test_session_launches_excel_and_opens_workbook_once(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook = Path(directory) / "w.xlsx"
            Workbook().save(workbook)

            dispatch = mock.MagicMock()
            client_mod = types.ModuleType("win32com.client")
            client_mod.DispatchEx = dispatch
            win32_mod = types.ModuleType("win32com")
            win32_mod.client = client_mod

            fake_image = mock.MagicMock()
            fake_image.mode = "RGB"

            with mock.patch.dict(
                sys.modules, {"win32com": win32_mod, "win32com.client": client_mod}
            ), mock.patch("PIL.ImageGrab.grabclipboard", return_value=fake_image):
                before = capture_launch_count()
                with ExcelCaptureSession(workbook) as session:
                    session.capture("Sheet1", "A1:B2", Path(directory) / "a.png")
                    session.capture("Sheet1", "A3:B4", Path(directory) / "b.png")
                    session.capture("Sheet1", "A5:B6", Path(directory) / "c.png")
                after = capture_launch_count()

            self.assertEqual(1, after - before)
            self.assertEqual(1, dispatch.call_count)
            workbooks_open = dispatch.return_value.Workbooks.Open
            self.assertEqual(1, workbooks_open.call_count)
            # Excel is quit exactly once on context exit.
            self.assertEqual(1, dispatch.return_value.Quit.call_count)


if __name__ == "__main__":
    unittest.main()
