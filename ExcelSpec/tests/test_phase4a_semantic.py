"""Phase 4A tests: SemanticDocumentIR, references, KnowledgeChunkIR, JSONL."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl.utils import get_column_letter

from excelspec.chunking import chunk_document
from excelspec.models.document_ir import (
    AssetIR,
    AssetType,
    CellIR,
    DocumentIR,
    RegionIR,
    RegionType,
    SheetIR,
    SourceRef,
    TableIR,
)
from excelspec.models.semantic import ReferenceType
from excelspec.semantic import assemble_semantic
from excelspec.semantic.references import extract_targets

FIXTURES = Path(__file__).resolve().parent / "fixtures"


def _col(index: int) -> str:
    return get_column_letter(index)


def _table_region(
    sheet_name: str,
    region_id: str,
    header: list[str],
    rows: list[list],
    *,
    header_rows: int = 1,
    column_semantics: dict | None = None,
    field_mapping: list | None = None,
) -> RegionIR:
    cells: list[CellIR] = []
    for ci, label in enumerate(header, start=1):
        cells.append(CellIR(coordinate=f"{_col(ci)}1", row=1, column=ci, raw_value=label, display_value=label))
    for ri, row in enumerate(rows, start=2):
        for ci, value in enumerate(row, start=1):
            if value is None:
                continue
            coord = f"{_col(ci)}{ri}"
            if isinstance(value, tuple):  # (display, formula)
                display, formula = value
                cells.append(
                    CellIR(coord, ri, ci, raw_value=formula, display_value=display, formula=formula, data_type="f")
                )
            else:
                cells.append(CellIR(coord, ri, ci, raw_value=value, display_value=str(value)))
    max_row = 1 + len(rows)
    max_col = len(header)
    rng = f"A1:{_col(max_col)}{max_row}"
    source = SourceRef(sheet=sheet_name, range=rng, workbook_path="wb.xlsx")
    meta = {"header_labels": {_col(i + 1): h for i, h in enumerate(header)}}
    if field_mapping is not None:
        meta["field_mapping"] = field_mapping
    table = TableIR(
        table_id=region_id, cells=cells, source=source, header_rows=header_rows,
        column_semantics=column_semantics or {}, metadata=meta,
    )
    return RegionIR(
        region_id=region_id, region_type=RegionType.TABLE, source=source,
        tables=[table], confidence=0.9,
        metadata={"candidate_type": "table", "detection_method": "test"},
    )


def _document(regions: list[RegionIR], sheet_name: str = "Sheet1", assets=None) -> DocumentIR:
    return DocumentIR(
        document_id="doc",
        title="Doc",
        source_path="wb.xlsx",
        sheets=[
            SheetIR(sheet_id="sheet-1", name=sheet_name, index=0, regions=regions, assets=assets or [])
        ],
        metadata={"extraction_mode": "fast"},
    )


# --------------------------------------------------------------------------- #
# SemanticDocumentIR
# --------------------------------------------------------------------------- #

class SemanticDocumentTests(unittest.TestCase):
    def test_basic_structure(self) -> None:
        region = _table_region("Sheet1", "t1", ["ID", "Name"], [["1", "a"], ["2", "b"]])
        sem = assemble_semantic(_document([region]), processing_mode="fast")
        self.assertEqual("1", sem.schema_version)
        self.assertEqual("doc", sem.document_id)
        self.assertEqual(1, len(sem.sheets))
        self.assertEqual(1, len(sem.regions))
        self.assertEqual("fast", sem.processing_mode)

    def test_every_region_has_source(self) -> None:
        region = _table_region("Sheet1", "t1", ["ID"], [["1"]])
        sem = assemble_semantic(_document([region]))
        for sem_region in sem.regions:
            self.assertIsNotNone(sem_region.source_range)

    def test_table_columns_and_rows_structured(self) -> None:
        region = _table_region("Sheet1", "t1", ["ID", "Name"], [["1", "a"], ["2", "b"]])
        sem = assemble_semantic(_document([region]))
        table = sem.regions[0].table
        self.assertEqual(["A", "B"], [c.column_id for c in table.columns])
        self.assertEqual(2, len(table.rows))

    def test_unmatched_header_preserved(self) -> None:
        region = _table_region("Sheet1", "t1", ["謎の列"], [["x"]])
        sem = assemble_semantic(_document([region]))
        column = sem.regions[0].table.columns[0]
        self.assertEqual("謎の列", column.source_header)
        self.assertIsNone(column.semantic_name)

    def test_missing_cell_not_shifted(self) -> None:
        region = _table_region("Sheet1", "t1", ["A", "B", "C"], [["1", None, "3"]])
        row = assemble_semantic(_document([region])).regions[0].table.rows[0]
        self.assertIsNone(row.values["B"])
        self.assertEqual("3", row.values["C"])

    def test_row_source_range(self) -> None:
        region = _table_region("Sheet1", "t1", ["A", "B"], [["1", "2"]])
        row = assemble_semantic(_document([region])).regions[0].table.rows[0]
        self.assertEqual("A2:B2", row.source_range)

    def test_formula_text_and_cached_value_coexist(self) -> None:
        region = _table_region("Sheet1", "t1", ["A", "B"], [["x", ("42", "=C1+1")]])
        row = assemble_semantic(_document([region])).regions[0].table.rows[0]
        self.assertEqual("42", row.values["B"])       # cached display
        self.assertEqual("=C1+1", row.formulas["B"])  # formula text


# --------------------------------------------------------------------------- #
# Formula references
# --------------------------------------------------------------------------- #

class ReferenceTests(unittest.TestCase):
    def test_same_sheet(self) -> None:
        targets, kind, resolved = extract_targets("=A1+B2")
        self.assertEqual(ReferenceType.SAME_SHEET, kind)
        self.assertTrue(resolved)
        self.assertEqual({"A1", "B2"}, {t.range for t in targets})

    def test_cross_sheet(self) -> None:
        targets, kind, _ = extract_targets("=Sheet2!A1:B5")
        self.assertEqual(ReferenceType.CROSS_SHEET, kind)
        self.assertEqual("Sheet2", targets[0].sheet)
        self.assertEqual("A1:B5", targets[0].range)

    def test_sheet_name_with_spaces(self) -> None:
        targets, kind, _ = extract_targets("='基本 情報'!B3")
        self.assertEqual(ReferenceType.CROSS_SHEET, kind)
        self.assertEqual("基本 情報", targets[0].sheet)
        self.assertEqual("B3", targets[0].range)

    def test_external_workbook(self) -> None:
        targets, kind, _ = extract_targets("=[Book.xlsx]Sheet1!A1")
        self.assertEqual(ReferenceType.EXTERNAL, kind)
        self.assertEqual("Book.xlsx", targets[0].workbook)

    def test_unparseable_named_range_keeps_original(self) -> None:
        targets, kind, resolved = extract_targets("=MyNamedRange")
        self.assertEqual(ReferenceType.NAMED_RANGE, kind)
        self.assertFalse(resolved)
        self.assertEqual("MyNamedRange", targets[0].name)

    def test_references_attached_to_document(self) -> None:
        region = _table_region("Sheet1", "t1", ["A", "B"], [["x", ("42", "=Sheet2!B3")]])
        sem = assemble_semantic(_document([region]))
        self.assertEqual(1, len(sem.references))
        self.assertEqual("=Sheet2!B3", sem.references[0].formula)
        self.assertEqual("42", sem.references[0].display_value)
        self.assertIn(sem.references[0].reference_id, sem.regions[0].formula_refs)


# --------------------------------------------------------------------------- #
# Chunking
# --------------------------------------------------------------------------- #

class ChunkingTests(unittest.TestCase):
    def _table_doc(self, rows):
        region = _table_region("Sheet1", "t1", ["ID", "Name"], rows)
        return assemble_semantic(_document([region]))

    def test_text_chunk(self) -> None:
        region = RegionIR(
            region_id="txt", region_type=RegionType.FREEFORM,
            source=SourceRef(sheet="Sheet1", range="A1:A2"),
            tables=[TableIR(table_id="txt", cells=[
                CellIR("A1", 1, 1, raw_value="行1", display_value="行1"),
                CellIR("A2", 2, 1, raw_value="行2", display_value="行2"),
            ])],
            confidence=0.5, metadata={"candidate_type": "text"},
        )
        chunks = chunk_document(assemble_semantic(_document([region])))
        self.assertEqual(1, len(chunks))
        self.assertEqual("text", chunks[0].chunk_type)
        self.assertIn("行1", chunks[0].text)

    def test_key_value_chunk(self) -> None:
        region = RegionIR(
            region_id="kv", region_type=RegionType.KEY_VALUE,
            source=SourceRef(sheet="Sheet1", range="A1:B2"),
            values={"画面ID": "SCR-001", "画面名": "一覧"},
            confidence=0.9, metadata={"candidate_type": "key_value"},
        )
        chunks = chunk_document(assemble_semantic(_document([region])))
        self.assertEqual("key_value", chunks[0].chunk_type)
        self.assertEqual(2, len(chunks[0].structured_data["key_values"]))

    def test_table_multi_row_chunk_and_no_row_split(self) -> None:
        rows = [[str(i), f"n{i}"] for i in range(1, 6)]  # 5 data rows
        from excelspec.chunking import ChunkingOptions, KnowledgeChunker
        sem = self._table_doc(rows)
        chunks = KnowledgeChunker(ChunkingOptions(max_rows=2)).chunk(sem)
        table_chunks = [c for c in chunks if c.chunk_type == "table"]
        self.assertEqual(3, len(table_chunks))  # 2 + 2 + 1
        total_rows = sum(len(c.structured_data["rows"]) for c in table_chunks)
        self.assertEqual(5, total_rows)  # no row dropped or split

    def test_every_table_chunk_repeats_columns(self) -> None:
        rows = [[str(i), f"n{i}"] for i in range(1, 6)]
        from excelspec.chunking import ChunkingOptions, KnowledgeChunker
        chunks = KnowledgeChunker(ChunkingOptions(max_rows=2)).chunk(self._table_doc(rows))
        for chunk in (c for c in chunks if c.chunk_type == "table"):
            self.assertEqual(2, len(chunk.structured_data["columns"]))

    def test_all_table_rows_end_up_in_chunks(self) -> None:
        rows = [[str(i), f"n{i}"] for i in range(1, 10)]
        sem = self._table_doc(rows)
        chunks = chunk_document(sem)
        emitted = sum(len(c.structured_data.get("rows", [])) for c in chunks if c.chunk_type == "table")
        self.assertEqual(9, emitted)

    def test_image_chunk_has_asset_refs_and_no_fabricated_text(self) -> None:
        assets = [AssetIR(asset_id="img1", asset_type=AssetType.IMAGE, uri="x.png", anchor="B2")]
        region = RegionIR(
            region_id="image-asset-1", region_type=RegionType.IMAGE,
            source=SourceRef(sheet="Sheet1", range="B2:B2"),
            asset_ids=["img1"], confidence=0.9,
            metadata={"candidate_type": "image"},
        )
        chunks = chunk_document(assemble_semantic(_document([region], assets=assets)))
        image = next(c for c in chunks if c.chunk_type == "image")
        self.assertEqual(["img1"], image.asset_refs)
        self.assertEqual("", image.text)  # no description fabricated

    def test_chunk_ids_stable_across_runs(self) -> None:
        sem1 = self._table_doc([["1", "a"], ["2", "b"]])
        sem2 = self._table_doc([["1", "a"], ["2", "b"]])
        self.assertEqual(
            [c.chunk_id for c in chunk_document(sem1)],
            [c.chunk_id for c in chunk_document(sem2)],
        )

    def test_chunk_order_and_index_stable(self) -> None:
        region1 = _table_region("Sheet1", "t1", ["A"], [["1"]])
        region2 = _table_region("Sheet1", "t2", ["B"], [["2"]])
        chunks = chunk_document(assemble_semantic(_document([region1, region2])))
        self.assertEqual([0, 1], [c.chunk_index for c in chunks])
        self.assertEqual(["doc:sheet-1:t1:t0", "doc:sheet-1:t2:t0"], [c.chunk_id for c in chunks])

    def test_no_duplicate_chunks(self) -> None:
        sem = self._table_doc([["1", "a"], ["2", "b"]])
        ids = [c.chunk_id for c in chunk_document(sem)]
        self.assertEqual(len(ids), len(set(ids)))

    def test_every_chunk_has_source(self) -> None:
        sem = self._table_doc([["1", "a"]])
        for chunk in chunk_document(sem):
            self.assertIsNotNone(chunk.source.sheet)
            self.assertIsNotNone(chunk.source.range)


# --------------------------------------------------------------------------- #
# JSONL exporter
# --------------------------------------------------------------------------- #

class JsonlExportTests(unittest.TestCase):
    def test_jsonl_one_object_per_line_and_japanese_preserved(self) -> None:
        from excelspec.exporters import ChunksJsonlExporter

        region = _table_region("画面項目", "t1", ["項目名", "型"], [["契約番号", "string"]])
        text = ChunksJsonlExporter().render(_document([region], sheet_name="画面項目"))
        lines = text.splitlines()
        for line in lines:
            json.loads(line)  # each line is a valid JSON object
        self.assertIn("契約番号", text)
        self.assertNotIn("\\u", text)  # not escaped to \uXXXX

    def test_semantic_json_roundtrips(self) -> None:
        from excelspec.exporters import SemanticJsonExporter

        region = _table_region("Sheet1", "t1", ["A"], [["1"]])
        payload = json.loads(SemanticJsonExporter().render(_document([region])))
        self.assertEqual("doc", payload["document_id"])
        self.assertIn("regions", payload)


# --------------------------------------------------------------------------- #
# Compatibility / stress
# --------------------------------------------------------------------------- #

class CompatibilityTests(unittest.TestCase):
    def test_legacy_and_default_exports_unaffected(self) -> None:
        # existing formats still produce output for a legacy-template run
        from excelspec.exporters import HtmlExporter, JsonExporter, MarkdownExporter
        region = _table_region("Sheet1", "t1", ["A"], [["1"]])
        document = _document([region])
        self.assertIn("Doc", MarkdownExporter().render(document))
        self.assertIn("doc", JsonExporter().render(document))
        self.assertIn("<!doctype html>", HtmlExporter().render(document))

    def test_semantic_assembly_on_stress_workbook_stays_sparse(self) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from excelspec.ingest import ingest_sparse_workbook
        from excelspec.detect.assemble import assemble_document

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "stress.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            for r in range(1, 6):
                sheet.cell(r, 1, f"A{r}")
                sheet.cell(r, 2, r)
            sheet.cell(row=1048576, column=16384).font = Font(bold=True)
            workbook.save(path)
            sparse = ingest_sparse_workbook(path, asset_dir=Path(directory) / "a")
            document, _ = assemble_document(sparse, mode="fast")
            sem = assemble_semantic(document)
        # content stays tiny despite the inflated dimension / distant style
        total_rows = sum(len(r.table.rows) for r in sem.regions if r.table)
        self.assertLessEqual(total_rows, 10)


if __name__ == "__main__":
    unittest.main()
