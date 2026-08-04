"""Rebuild the synthetic, anonymized XLSX acceptance fixtures."""

from __future__ import annotations

import json
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw


ROOT = Path(__file__).resolve().parent
WORKBOOKS = ROOT / "workbooks"
SCREENS = ROOT / "screens"
TEMPLATES = ROOT / "templates"

SCREEN_TEMPLATE = """\
schema_version: "1.0"
template_id: fixture-screen-design
version: "1.0"
name: 脱敏画面設計書
match:
  sheet_name_patterns: ["^表紙$", "^改訂履歴$", "^画面項目$"]
  fingerprints:
    - sheet_name_pattern: "^表紙$"
      required_text: ["画面設計書", "画面ID"]
    - sheet_name_pattern: "^画面項目$"
      required_text: ["画面項目一覧", "データ型"]
  minimum_score: 0.65
sheets:
  - sheet_id: cover
    name_pattern: "^表紙$"
    order: 10
    regions:
      - region_id: document-info
        region_type: key_value
        title: 表紙
        required: true
        locator: {mode: fixed, range: "A1:D10"}
        extractor:
          kind: key_value
          key_column: 1
          value_column: 2
          key_semantics:
            画面ID: screen_id
            画面名: screen_name
            作成者: author
            作成日: created_at
      - region_id: screen-layout
        region_type: layout
        title: 画面レイアウト
        locator:
          mode: anchor
          anchor_text: 画面レイアウト
          row_offset: 1
          height: 12
          width: 8
        extractor: {kind: asset}
        screenshot_bindings:
          - asset_id: fixture-screen-shot
            path: ../screens/layout.png
            asset_type: layout
            description: 脱敏検索画面
  - sheet_id: revisions
    name_pattern: "^改訂履歴$"
    order: 20
    regions:
      - region_id: revision-history
        region_type: table
        title: 改訂履歴
        required: true
        locator: {mode: fixed, range: "A1:D4"}
        extractor:
          kind: table
          header_rows: 2
          column_semantics:
            版: revision
            日付: changed_at
            変更内容: change
            担当: author
  - sheet_id: items
    name_pattern: "^画面項目$"
    order: 30
    regions:
      - region_id: screen-item-table
        region_type: table
        title: 画面項目表
        required: true
        locator:
          mode: anchor
          anchor_pattern: "^画面項目一覧$"
          end_anchor_pattern: "^備考$"
          row_offset: 1
          width: 8
        extractor:
          kind: table
          header_rows: 2
          column_semantics:
            項目ID: item_id
            項目名: item_name
            データ型: data_type
            桁数: length
            必須: required
            表示名: label
            入力チェック: validation
validation_rules:
  - rule_id: require-screen-id
    kind: required
    field: cover.document-info.screen_id
metadata: {fixture: true, document_kind: screen-design}
"""

API_TEMPLATE = {
    "schema_version": "1.0",
    "template_id": "fixture-api-spec",
    "version": "1.0",
    "name": "脱敏API仕様書",
    "match": {
        "sheet_name_patterns": ["^API概要$", "^リクエスト$", "^レスポンス$"],
        "fingerprints": [
            {
                "sheet_name_pattern": "^API概要$",
                "required_text": ["API仕様書", "エンドポイント", "HTTPメソッド"],
            }
        ],
        "minimum_score": 0.65,
    },
    "sheets": [
        {
            "sheet_id": "overview",
            "name_pattern": "^API概要$",
            "regions": [
                {
                    "region_id": "api-info",
                    "region_type": "key_value",
                    "required": True,
                    "locator": {"mode": "fixed", "range": "A1:D10"},
                    "extractor": {
                        "kind": "key_value",
                        "key_column": 1,
                        "value_column": 2,
                        "key_semantics": {
                            "API ID": "api_id",
                            "API名": "api_name",
                            "エンドポイント": "endpoint",
                            "HTTPメソッド": "http_method",
                        },
                    },
                }
            ],
        },
        {
            "sheet_id": "request",
            "name_pattern": "^リクエスト$",
            "regions": [
                {
                    "region_id": "request-fields",
                    "region_type": "table",
                    "required": True,
                    "locator": {
                        "mode": "anchor",
                        "anchor_text": "リクエスト項目",
                        "row_offset": 1,
                        "width": 5,
                    },
                    "extractor": {
                        "kind": "table",
                        "header_rows": 1,
                        "column_semantics": {
                            "パラメータ名": "parameter_name",
                            "場所": "location",
                            "データ型": "data_type",
                            "必須": "required",
                            "説明": "description",
                        },
                        "options": {"stop_after_blank_rows": 1},
                    },
                }
            ],
        },
        {
            "sheet_id": "response",
            "name_pattern": "^レスポンス$",
            "regions": [
                {
                    "region_id": "response-fields",
                    "region_type": "table",
                    "locator": {
                        "mode": "anchor",
                        "anchor_text": "レスポンス項目",
                        "row_offset": 1,
                        "width": 4,
                    },
                    "extractor": {
                        "kind": "table",
                        "header_rows": 1,
                        "column_semantics": {
                            "項目名": "field_name",
                            "データ型": "data_type",
                            "必須": "required",
                            "説明": "description",
                        },
                        "options": {"stop_after_blank_rows": 1},
                    },
                }
            ],
        },
    ],
    "validation_rules": [
        {
            "rule_id": "require-endpoint",
            "kind": "required",
            "field": "overview.api-info.endpoint",
        }
    ],
    "metadata": {"fixture": True, "document_kind": "api-specification"},
}

INVALID_TEMPLATE = """\
schema_version: "1.0"
template_id: broken-fixture
version: "1.0"
name: エラーテンプレート
sheets:
  - sheet_id: broken
    name_pattern: "^表紙$"
    regions:
      - region_id: missing-range
        region_type: table
        locator:
          mode: fixed
"""


def _style_title(sheet, range_ref: str) -> None:
    sheet.merge_cells(range_ref)
    cell = sheet[range_ref.split(":")[0]]
    cell.font = Font(size=16, bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center")


def _inject_shape_text(workbook_path: Path) -> None:
    namespace = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
    ET.register_namespace("xdr", namespace)
    ET.register_namespace("a", drawing_ns)
    replacement = workbook_path.with_suffix(".tmp")
    with zipfile.ZipFile(workbook_path) as source, zipfile.ZipFile(
        replacement, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "xl/drawings/drawing1.xml":
                root = ET.fromstring(data)
                anchor = next(iter(root))
                shape = ET.fromstring(
                    f"""<xdr:sp xmlns:xdr="{namespace}" xmlns:a="{drawing_ns}">
<xdr:nvSpPr><xdr:cNvPr id="99" name="検索注記"/><xdr:cNvSpPr/></xdr:nvSpPr>
<xdr:spPr/><xdr:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r>
<a:t>検索ボタンで一覧を更新</a:t></a:r></a:p></xdr:txBody></xdr:sp>"""
                )
                anchor.insert(max(0, len(anchor) - 1), shape)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target.writestr(info, data)
    replacement.replace(workbook_path)


def build_screen_workbook(path: Path, image_path: Path) -> None:
    workbook = Workbook()
    workbook.properties.title = "脱敏利用者検索画面"
    workbook.properties.creator = "Fixture Team"
    cover = workbook.active
    cover.title = "表紙"
    _style_title(cover, "A1:D1")
    cover["A1"] = "画面設計書"
    rows = [
        ("画面ID", "SCR-DEMO-001"),
        ("画面名", "利用者検索"),
        ("作成者", "担当A"),
        ("作成日", "2026-01-15"),
    ]
    for row, values in enumerate(rows, start=2):
        cover.cell(row, 1, values[0])
        cover.cell(row, 2, values[1])
    cover["A12"] = "画面レイアウト"
    cover["A12"].font = Font(bold=True)
    picture = XlsxImage(image_path)
    picture.width = 256
    picture.height = 160
    picture.anchor = "B13"
    cover.add_image(picture)

    history = workbook.create_sheet("改訂履歴")
    _style_title(history, "A1:D1")
    history["A1"] = "改訂履歴"
    history.append(["版", "日付", "変更内容", "担当"])
    history.append(["1.0", "2026-01-15", "新規作成", "担当A"])
    history.append(["1.1", "2026-02-01", "検索条件を追加", "担当B"])

    items = workbook.create_sheet("画面項目")
    _style_title(items, "A1:H1")
    items["A1"] = "画面項目一覧"
    for range_ref in ("A2:A3", "B2:B3", "C2:D2", "E2:E3", "F2:F3", "G2:G3", "H2:H3"):
        items.merge_cells(range_ref)
    headers = {
        "A2": "項目ID",
        "B2": "項目名",
        "C2": "属性",
        "C3": "データ型",
        "D3": "桁数",
        "E2": "必須",
        "F2": "表示名",
        "G2": "入力チェック",
        "H2": "備考",
    }
    for coordinate, value in headers.items():
        items[coordinate] = value
        items[coordinate].font = Font(bold=True)
        items[coordinate].fill = PatternFill("solid", fgColor="E2F0D9")
    for values in (
        ("USR-ID", "利用者ID", "文字列", 12, "○", "利用者ID", "英数字", ""),
        ("USR-NAME", "利用者名", "文字列", 40, "", "利用者名", "", ""),
        ("SEARCH", "検索", "ボタン", None, "", "検索", "", "押下で検索"),
    ):
        items.append(values)
    items["A7"] = "備考"
    items["A8"] = "個人名・実データを含まない合成 fixture"
    workbook.save(path)
    workbook.close()
    _inject_shape_text(path)


def build_api_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.properties.title = "脱敏利用者検索API"
    workbook.properties.creator = "Fixture Team"
    overview = workbook.active
    overview.title = "API概要"
    _style_title(overview, "A1:D1")
    overview["A1"] = "API仕様書"
    for row, values in enumerate(
        (
            ("API ID", "API-DEMO-001"),
            ("API名", "利用者検索API"),
            ("エンドポイント", "/v1/demo-users"),
            ("HTTPメソッド", "GET"),
        ),
        start=2,
    ):
        overview.cell(row, 1, values[0])
        overview.cell(row, 2, values[1])

    request = workbook.create_sheet("リクエスト")
    request["A1"] = "リクエスト項目"
    request.append(["パラメータ名", "場所", "データ型", "必須", "説明"])
    request.append(["query", "query", "string", "false", "検索語"])
    request.append(["limit", "query", "integer", "false", "最大件数"])

    response = workbook.create_sheet("レスポンス")
    response["A1"] = "レスポンス項目"
    response.append(["項目名", "データ型", "必須", "説明"])
    response.append(["userId", "string", "true", "合成利用者ID"])
    response.append(["displayName", "string", "true", "表示名"])
    workbook.save(path)
    workbook.close()


def main() -> None:
    for directory in (WORKBOOKS, SCREENS, TEMPLATES):
        directory.mkdir(parents=True, exist_ok=True)
    image_path = SCREENS / "layout.png"
    image = Image.new("RGB", (256, 160), "#f4f7fb")
    draw = ImageDraw.Draw(image)
    draw.rectangle((12, 12, 244, 42), fill="#d9eaf7", outline="#5b7894")
    draw.rectangle((20, 60, 170, 84), fill="white", outline="#708090")
    draw.rectangle((182, 60, 236, 84), fill="#4f81bd", outline="#315f91")
    draw.rectangle((20, 100, 236, 145), fill="white", outline="#708090")
    image.save(image_path, format="PNG")
    build_screen_workbook(WORKBOOKS / "screen-design.xlsx", image_path)
    build_api_workbook(WORKBOOKS / "api-spec.xlsx")
    (TEMPLATES / "screen-design.yaml").write_text(SCREEN_TEMPLATE, encoding="utf-8")
    (TEMPLATES / "api-spec.json").write_text(
        json.dumps(API_TEMPLATE, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (TEMPLATES / "invalid-template.yaml").write_text(
        INVALID_TEMPLATE, encoding="utf-8"
    )
    (ROOT / "screenshots.json").write_text(
        json.dumps(
            {
                "version": "1",
                "assets": [
                    {
                        "asset_id": "fixture-screen-shot",
                        "path": "screens/layout.png",
                        "sheet": "表紙",
                        "anchor": "B13:H24",
                        "asset_type": "layout",
                        "description": "脱敏検索画面",
                        "ocr": {"status": "pending"},
                        "vlm": {"status": "not_requested"},
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
