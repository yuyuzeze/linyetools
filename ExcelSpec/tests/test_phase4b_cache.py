"""Phase 4B tests: content-hash cache, OCR/VLM providers, semantic coverage."""

from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excelspec.detect.assemble import assemble_document
from excelspec.exporters import ChunksJsonlExporter
from excelspec.ingest.sparse_model import SparseCell, SparseSheet, SparseWorkbookIR
from excelspec.models.document_ir import AssetIR, AssetType
from excelspec.pipeline import run_pipeline
from excelspec.providers import (
    NullOcrProvider,
    NullVlmProvider,
    ProviderContext,
    ProviderResult,
)

FIXTURES = Path(__file__).resolve().parent / "fixtures"
WORKBOOK = FIXTURES / "workbooks" / "screen-design.xlsx"


# --------------------------------------------------------------------------- #
# Cache
# --------------------------------------------------------------------------- #

class CacheTests(unittest.TestCase):
    def test_second_run_hits_cache(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            assets = Path(directory) / "assets"
            cache_dir = Path(directory) / "c"
            miss = run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
            hit = run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
        self.assertEqual("miss", miss.processing["cache"])
        self.assertEqual("hit", hit.processing["cache"])

    def test_cache_hit_is_byte_identical(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            assets = Path(directory) / "assets"
            cache_dir = Path(directory) / "c"
            miss = run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
            hit = run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
            self.assertEqual(
                ChunksJsonlExporter().render(miss.document),
                ChunksJsonlExporter().render(hit.document),
            )

    def test_workbook_change_invalidates(self) -> None:
        from openpyxl import Workbook

        directory = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, directory, ignore_errors=True)
        workbook = directory / "w.xlsx"
        assets = directory / "assets"
        cache_dir = directory / "c"

        book = Workbook()
        sheet = book.active
        sheet.title = "S"
        sheet.append(["ID", "Name"])
        sheet.append(["1", "a"])
        book.save(workbook)
        first = run_pipeline(workbook, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
        hit = run_pipeline(workbook, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)

        sheet["A3"] = "changed"
        book.save(workbook)
        third = run_pipeline(workbook, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)

        self.assertEqual("miss", first.processing["cache"])
        self.assertEqual("hit", hit.processing["cache"])
        self.assertEqual("miss", third.processing["cache"])  # hash changed

    def test_profile_change_invalidates(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            profile = Path(directory) / "p.yaml"
            profile.write_text(
                "profile_id: p\ndocument_type: d\nfields: {}\n", encoding="utf-8"
            )
            assets = Path(directory) / "assets"
            cache_dir = Path(directory) / "c"
            run_pipeline(WORKBOOK, mode="fast", profile=profile, asset_dir=assets, cache=True, cache_dir=cache_dir)
            profile.write_text(
                "profile_id: p\ndocument_type: d\nfields:\n  f:\n    aliases: [X]\n",
                encoding="utf-8",
            )
            second = run_pipeline(WORKBOOK, mode="fast", profile=profile, asset_dir=assets, cache=True, cache_dir=cache_dir)
        self.assertEqual("miss", second.processing["cache"])  # profile hash changed

    def test_chunk_params_do_not_invalidate_document_cache(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            assets = Path(directory) / "assets"
            cache_dir = Path(directory) / "c"
            run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
            hit = run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
        self.assertEqual("hit", hit.processing["cache"])
        # different chunk params only change the chunk output, not the doc cache
        few = ChunksJsonlExporter(max_rows=1).render(hit.document)
        many = ChunksJsonlExporter(max_rows=100).render(hit.document)
        self.assertNotEqual(few, many)

    def test_corrupt_cache_recovers(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            assets = Path(directory) / "assets"
            cache_dir = Path(directory) / "c"
            run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
            entries = list((cache_dir / ".excelspec-cache" / "document").glob("*.json"))
            entries[0].write_text("{ not valid json", encoding="utf-8")
            recovered = run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
        self.assertEqual("miss", recovered.processing["cache"])
        self.assertTrue(recovered.processing.get("cache_warnings"))

    def test_cache_not_created_next_to_source(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            assets = Path(directory) / "assets"
            cache_dir = Path(directory) / "out"
            run_pipeline(WORKBOOK, mode="fast", asset_dir=assets, cache=True, cache_dir=cache_dir)
        # cache lives under the given cache_dir, never beside the workbook
        self.assertFalse((WORKBOOK.parent / ".excelspec-cache").exists())


# --------------------------------------------------------------------------- #
# Providers
# --------------------------------------------------------------------------- #

class _SpyVlm:
    available = True

    def __init__(self) -> None:
        self.calls: list = []

    def describe(self, asset, context: ProviderContext) -> ProviderResult:
        self.calls.append((asset, context))
        return ProviderResult(text="a diagram", provider="spy", source="vlm", confidence=0.8)


class _BoomVlm:
    available = True

    def describe(self, asset, context: ProviderContext) -> ProviderResult:
        raise RuntimeError("provider exploded")


def _layout_workbook(directory: str) -> SparseWorkbookIR:
    cells = {
        (1, 1): SparseCell(1, 1, "A1", "画面レイアウト", "画面レイアウト", "s", None, None, None),
        (2, 1): SparseCell(2, 1, "A2", "説明", "説明", "s", None, None, None),
    }
    sheet = SparseSheet(
        name="Layout", sheet_id="sheet-1", index=0, cells=cells,
        content_bounds=(1, 1, 2, 1),
        assets=[AssetIR(asset_id="img1", asset_type=AssetType.IMAGE, uri="x.png", anchor="A2")],
    )
    workbook = SparseWorkbookIR(path=str(WORKBOOK), sheets=[sheet], styles={})
    workbook.metadata["asset_directory"] = directory
    return workbook


class ProviderTests(unittest.TestCase):
    def test_null_providers_report_unavailable(self) -> None:
        self.assertFalse(NullOcrProvider().available)
        self.assertFalse(NullVlmProvider().available)

    def test_fast_mode_never_calls_provider(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            spy = _SpyVlm()
            assemble_document(_layout_workbook(directory), mode="fast", vlm=spy)
        self.assertEqual([], spy.calls)

    def test_null_provider_does_not_change_conversion(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            plain, _ = assemble_document(_layout_workbook(directory), mode="fast")
            withnull, _ = assemble_document(
                _layout_workbook(directory), mode="fast", ocr=NullOcrProvider(), vlm=NullVlmProvider()
            )
        self.assertEqual(
            [r.region_id for s in plain.sheets for r in s.regions],
            [r.region_id for s in withnull.sheets for r in s.regions],
        )

    def test_auto_mode_calls_available_provider_and_tags_result(self) -> None:
        import excelspec.render as render

        class _NoExcel:
            def __init__(self, path): pass
            def __enter__(self): return self
            def __exit__(self, *exc): return False
            def capture(self, *a, **k): raise RuntimeError("no excel")

        with tempfile.TemporaryDirectory() as directory:
            spy = _SpyVlm()
            with unittest_mock_patch(render, "ExcelCaptureSession", _NoExcel):
                document, _ = assemble_document(_layout_workbook(directory), mode="auto", vlm=spy)
        self.assertTrue(spy.calls)
        visual = next(r for s in document.sheets for r in s.regions if r.metadata.get("visual"))
        self.assertEqual("spy", visual.metadata["vlm_result"]["provider"])
        self.assertEqual("vlm", visual.metadata["vlm_result"]["source"])

    def test_provider_failure_keeps_assets_and_records_diagnostic(self) -> None:
        import excelspec.render as render

        class _NoExcel:
            def __init__(self, path): pass
            def __enter__(self): return self
            def __exit__(self, *exc): return False
            def capture(self, *a, **k): raise RuntimeError("no excel")

        with tempfile.TemporaryDirectory() as directory:
            with unittest_mock_patch(render, "ExcelCaptureSession", _NoExcel):
                document, _ = assemble_document(_layout_workbook(directory), mode="auto", vlm=_BoomVlm())
        visual = next(r for s in document.sheets for r in s.regions if r.metadata.get("visual"))
        self.assertIn("img1", visual.asset_ids)  # asset preserved
        codes = {d["code"] for d in visual.metadata.get("diagnostics", [])}
        self.assertIn("provider.vlm_failed", codes)


# --------------------------------------------------------------------------- #
# Coverage
# --------------------------------------------------------------------------- #

class CoverageTests(unittest.TestCase):
    def test_coverage_stats_and_full_row_coverage(self) -> None:
        from excelspec.semantic import assemble_semantic
        from excelspec.semantic.coverage import analyze_coverage
        from excelspec.chunking import chunk_document
        from excelspec.ingest import ingest_sparse_workbook

        with tempfile.TemporaryDirectory() as directory:
            sparse = ingest_sparse_workbook(WORKBOOK, asset_dir=Path(directory))
            document, _ = assemble_document(sparse, mode="fast")
        semantic = assemble_semantic(document)
        chunks = chunk_document(semantic)
        report = analyze_coverage(semantic, chunks)
        self.assertEqual(len(semantic.regions), report.stats["semantic_region_count"])
        self.assertEqual(len(chunks), report.stats["chunk_count"])
        # every table row is chunked
        self.assertEqual(
            report.stats["table_row_count"], report.stats["chunked_table_row_count"]
        )
        self.assertEqual(1.0, report.stats["source_coverage"])


def unittest_mock_patch(target, name, value):
    from unittest import mock

    return mock.patch.object(target, name, value)


if __name__ == "__main__":
    unittest.main()
