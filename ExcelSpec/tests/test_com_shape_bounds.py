"""Excel COM Shape-bounds (mock COM) tests for dynamic-bottom capture."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest import mock

from excelspec.models.document_ir import (
    AssetIR,
    AssetType,
    CellIR,
    DocumentIR,
    RegionIR,
    RegionType,
    SheetIR,
    SourceRef,
    TableIR,
)
from excelspec.models.template import (
    ExtractionSpec,
    LocatorMode,
    RegionLocator,
    RegionTemplate,
    SheetTemplate,
    TemplateMatch,
    TemplateSpec,
)
from excelspec.render.excel_capture import ExcelCaptureSession
from excelspec.templates import MatchResult, extract_with_template


# --------------------------------------------------------------------------- #
# fake COM object model
# --------------------------------------------------------------------------- #

class _FakeCell:
    def __init__(self, row, col):
        self.Row = row
        self.Column = col


class _FakeShape:
    def __init__(self, name, stype, top_left=None, bottom_right=None, group_items=None):
        self.Name = name
        self.Type = stype
        self._tl = _FakeCell(*top_left) if top_left else None
        self._br = _FakeCell(*bottom_right) if bottom_right else None
        self.GroupItems = group_items

    @property
    def TopLeftCell(self):
        if self._tl is None:
            raise RuntimeError("no anchor cell")
        return self._tl

    @property
    def BottomRightCell(self):
        if self._br is None:
            raise RuntimeError("no anchor cell")
        return self._br


class _FakeGroupItems:
    def __init__(self, items):
        self._items = items

    @property
    def Count(self):
        return len(self._items)

    def Item(self, index):
        return self._items[index - 1]


class _FakeWorksheet:
    def __init__(self, shapes):
        self.Shapes = shapes


class _FakeWorkbook:
    def __init__(self, sheet_name, shapes):
        self._name = sheet_name
        self._ws = _FakeWorksheet(shapes)

    def Worksheets(self, name):
        assert name == self._name
        return self._ws


def _session(sheet_name, shapes):
    session = ExcelCaptureSession("x.xlsx")
    session._workbook = _FakeWorkbook(sheet_name, shapes)  # bypass real open
    return session


class ResolveShapeBoundsTests(unittest.TestCase):
    def test_connector_below_nodes_sets_bottom(self) -> None:
        shapes = [
            _FakeShape("node", 1, (10, 3), (30, 5)),
            _FakeShape("connector", 9, (30, 4), (55, 4)),  # deeper than nodes
        ]
        result = _session("S", shapes).resolve_shape_bounds(
            "S", top_row=6, left_column="B", right_column="BZ", section_limit=200
        )
        self.assertEqual(55, result["bottom"])
        self.assertEqual({"node", "connector"}, {s["name"] for s in result["included_shapes"]})

    def test_group_outer_bounds_used_when_valid(self) -> None:
        group = _FakeShape("grp", 6, (10, 3), (60, 6))  # outer bounds valid
        result = _session("S", [group]).resolve_shape_bounds(
            "S", top_row=6, left_column="B", right_column="BZ", section_limit=200
        )
        self.assertEqual(60, result["bottom"])

    def test_group_iterates_items_when_outer_invalid(self) -> None:
        items = _FakeGroupItems([
            _FakeShape("i1", 1, (10, 3), (40, 5)),
            _FakeShape("i2", 1, (45, 3), (70, 5)),
        ])
        group = _FakeShape("grp", 6, None, None, group_items=items)  # outer invalid
        result = _session("S", [group]).resolve_shape_bounds(
            "S", top_row=6, left_column="B", right_column="BZ", section_limit=200
        )
        self.assertEqual(70, result["bottom"])

    def test_shape_beyond_section_excluded(self) -> None:
        shapes = [
            _FakeShape("keep", 1, (10, 3), (40, 5)),
            _FakeShape("next_section", 1, (60, 3), (90, 5)),
        ]
        result = _session("S", shapes).resolve_shape_bounds(
            "S", top_row=6, left_column="B", right_column="BZ", section_limit=50
        )
        self.assertEqual(40, result["bottom"])
        self.assertIn("next_section", {s["name"] for s in result["excluded_shapes"]})

    def test_shape_outside_band_excluded(self) -> None:
        shapes = [
            _FakeShape("in", 1, (10, 3), (40, 5)),
            _FakeShape("far_right", 1, (10, 100), (40, 105)),  # beyond BZ(78)
        ]
        result = _session("S", shapes).resolve_shape_bounds(
            "S", top_row=6, left_column="B", right_column="BZ", section_limit=200
        )
        self.assertEqual(40, result["bottom"])
        self.assertIn("far_right", {s["name"] for s in result["excluded_shapes"]})

    def test_repeated_queries_launch_excel_once(self) -> None:
        import sys
        import types

        dispatch = mock.MagicMock()
        fake_ws = _FakeWorksheet([_FakeShape("n", 1, (10, 3), (30, 5))])
        dispatch.return_value.Workbooks.Open.return_value.Worksheets.return_value = fake_ws
        client_mod = types.ModuleType("win32com.client")
        client_mod.DispatchEx = dispatch
        win32_mod = types.ModuleType("win32com")
        win32_mod.client = client_mod

        with tempfile.TemporaryDirectory() as directory:
            workbook = Path(directory) / "w.xlsx"
            from openpyxl import Workbook
            Workbook().save(workbook)
            with mock.patch.dict(sys.modules, {"win32com": win32_mod, "win32com.client": client_mod}):
                session = ExcelCaptureSession(workbook)
                session.resolve_shape_bounds("S", top_row=6, left_column="B", right_column="BZ")
                session.resolve_shape_bounds("S", top_row=6, left_column="B", right_column="BZ")
        self.assertEqual(1, dispatch.call_count)
        self.assertEqual(1, dispatch.return_value.Workbooks.Open.call_count)


# --------------------------------------------------------------------------- #
# engine integration: COM bottom > OOXML bottom, and COM failure fallback
# --------------------------------------------------------------------------- #

def _transition_doc(directory, assets):
    cells = [CellIR("A5", 5, 1, "■画面遷移図", display_value="■画面遷移図")]
    table = TableIR(table_id="raw-grid", cells=cells, source=SourceRef(sheet="画面遷移図", range="A5:BZ200"))
    region = RegionIR(region_id="raw-grid", region_type=RegionType.FREEFORM,
                      source=SourceRef(sheet="画面遷移図", range="A5:BZ200"), tables=[table])
    sheet = SheetIR(sheet_id="sheet-1", name="画面遷移図", index=0, regions=[region], assets=assets)
    workbook = Path(directory) / "画面遷移図.xlsx"
    workbook.write_bytes(b"x")
    return DocumentIR(document_id="d", title="d", source_path=str(workbook),
                      sheets=[sheet], metadata={"asset_directory": str(Path(directory) / "assets")})


def _transition_template():
    return TemplateSpec(
        template_id="t", version="1", name="t", schema_version="1.0",
        match=TemplateMatch(minimum_score=0.1),
        sheets=[SheetTemplate(sheet_id="st", name_pattern="^画面遷移図$", regions=[
            RegionTemplate(region_id="transition-diagram", region_type="layout", title="図",
                locator=RegionLocator(mode=LocatorMode.ANCHOR, anchor_pattern="^\\s*■?\\s*画面遷移図\\s*$", row_offset=1),
                extractor=ExtractionSpec(kind="asset", options={
                    "screenshot": True, "screenshot_bounds": "dynamic_bottom",
                    "left_column": "B", "right_column": "BZ",
                    "padding_bottom_rows": 2, "max_bottom_row": 2000, "text_fallback": True,
                }))
        ])],
    )


class _FakeSession:
    def __init__(self, path, bottom, raise_error=False):
        self.workbook_path = Path(path).resolve()
        self._bottom = bottom
        self._raise = raise_error

    def resolve_shape_bounds(self, sheet_name, *, top_row, left_column, right_column, section_limit=None):
        if self._raise:
            raise RuntimeError("com boom")
        return {"bottom": self._bottom, "included_shapes": [{"name": "conn", "type": 9}], "excluded_shapes": []}


class ComIntegrationTests(unittest.TestCase):
    def _run(self, assets, fake_session):
        captured = {}

        def capture(*, destination, workbook_path, sheet_name, a1_range):
            path = Path(destination); path.parent.mkdir(parents=True, exist_ok=True); path.write_bytes(b"png")
            captured["range"] = a1_range
            return path, "excel_com"

        with tempfile.TemporaryDirectory() as directory:
            document = _transition_doc(directory, assets)
            with mock.patch("excelspec.templates.engine.render_region_screenshot", side_effect=capture), \
                 mock.patch("excelspec.render.excel_capture.active_capture_session",
                            return_value=_FakeSession(document.source_path, *fake_session)):
                result = extract_with_template(document, MatchResult(mode="template", template=_transition_template(), candidates=[]))
        region = next(r for r in result.document.sheets[0].regions if r.region_id == "transition-diagram")
        asset = next(a for a in result.document.sheets[0].assets if a.asset_id in region.asset_ids)
        return captured, region, asset, result

    def test_com_bottom_larger_than_ooxml_is_used(self) -> None:
        # OOXML drawing bottom 50, COM shape bottom 80 -> final 80 + padding 2
        assets = [AssetIR(asset_id="d1", asset_type=AssetType.SHAPE, uri="", anchor="C10:C50")]
        captured, region, asset, _ = self._run(assets, (80, False))
        self.assertEqual("B6:BZ82", captured["range"])
        self.assertEqual(80, asset.metadata["com_shape_bottom"])
        self.assertEqual(50, asset.metadata["ooxml_bottom"])
        self.assertEqual("com_shape_bottom", asset.metadata["dominant_source"])
        self.assertTrue(asset.metadata["included_shapes"])

    def test_com_failure_falls_back_to_ooxml_with_diagnostic(self) -> None:
        assets = [AssetIR(asset_id="d1", asset_type=AssetType.SHAPE, uri="", anchor="C10:C50")]
        captured, region, asset, result = self._run(assets, (None, True))  # raises
        self.assertEqual("B6:BZ52", captured["range"])  # OOXML 50 + padding 2
        codes = {d.code for sheet in result.document.sheets for d in sheet.diagnostics}
        self.assertIn("template.com_shape_bounds_failed", codes)


class NoComTests(unittest.TestCase):
    def test_default_fast_does_not_call_com_shape_bounds(self) -> None:
        from openpyxl import Workbook
        from excelspec.pipeline import run_pipeline

        with tempfile.TemporaryDirectory() as directory:
            workbook = Path(directory) / "画面遷移図.xlsx"
            book = Workbook(); book.active.title = "画面遷移図"; book.active.append(["■画面遷移図"])
            book.save(workbook)
            with mock.patch.object(ExcelCaptureSession, "resolve_shape_bounds") as spy:
                run_pipeline(workbook, asset_dir=Path(directory) / "a")  # default fast
        spy.assert_not_called()

    def test_no_screenshot_config_does_not_call_com_shape_bounds(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            cells = [CellIR("A1", 1, 1, "x", display_value="x")]
            table = TableIR(table_id="raw-grid", cells=cells, source=SourceRef(sheet="S", range="A1:A1"))
            region = RegionIR(region_id="raw-grid", region_type=RegionType.FREEFORM,
                              source=SourceRef(sheet="S", range="A1:A1"), tables=[table])
            document = DocumentIR(document_id="d", title="d", source_path=str(Path(directory) / "w.xlsx"),
                                  sheets=[SheetIR(sheet_id="s1", name="S", index=0, regions=[region])], metadata={})
            template = TemplateSpec(
                template_id="t", version="1", name="t", schema_version="1.0",
                match=TemplateMatch(minimum_score=0.1),
                sheets=[SheetTemplate(sheet_id="s", name_pattern="^S$", regions=[
                    RegionTemplate(region_id="r", region_type="freeform",
                                   locator=RegionLocator(mode=LocatorMode.FIXED, range="A1:A1"),
                                   extractor=ExtractionSpec(kind="freeform"))  # no screenshot
                ])],
            )
            with mock.patch.object(ExcelCaptureSession, "resolve_shape_bounds") as spy:
                extract_with_template(document, MatchResult(mode="template", template=template, candidates=[]))
        spy.assert_not_called()


if __name__ == "__main__":
    unittest.main()
