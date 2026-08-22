"""Phase 2 tests: sparse OOXML ingestion, adapter, engine dispatch, SheetIndex."""

from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from excelspec.ingest import ingest_xlsx
from excelspec.ingest.base import UnsupportedWorkbookError, ingest_with_engine
from excelspec.ingest.sparse import SparseOoxmlIngestor
from excelspec.ingest.workbook import XlsxIngestOptions, XlsxIngestor
from excelspec.models.document_ir import DocumentIR

FIXTURES = Path(__file__).resolve().parent / "fixtures"


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

_CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>"""

_ROOT_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""

_WORKBOOK = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="Data" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""

_WORKBOOK_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""

_SHARED_STRINGS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>共有文字列</t></si>
</sst>"""

# s=1 is a bold font so openpyxl reports has_style=True (a real style-only cell).
_STYLES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><sz val="11"/><name val="Calibri"/></font></fonts>
<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>
<borders count="1"><border/></borders>
<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
<cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/></cellXfs>
<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>"""

# A1 shared string, A2 number, A3 bool, A4 error, A5 inline string,
# A6 formula with cached value, B1:C1 merged (B1 value), and a distant
# style-only cell at XFD1048576 with an inflated <dimension>.
_SHEET = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<dimension ref="A1:XFD1048576"/>
<sheetData>
<row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1" t="inlineStr"><is><t>merged</t></is></c></row>
<row r="2"><c r="A2"><v>12</v></c></row>
<row r="3"><c r="A3" t="b"><v>1</v></c></row>
<row r="4"><c r="A4" t="e"><v>#DIV/0!</v></c></row>
<row r="5"><c r="A5" t="inlineStr"><is><t>inline文字</t></is></c></row>
<row r="6"><c r="A6"><f>A2+10</f><v>22</v></c></row>
<row r="1048576"><c r="XFD1048576" s="1"/></row>
</sheetData>
<mergeCells count="1"><mergeCell ref="B1:C1"/></mergeCells>
</worksheet>"""


def _write_crafted_xlsx(path: Path) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _CONTENT_TYPES)
        archive.writestr("_rels/.rels", _ROOT_RELS)
        archive.writestr("xl/workbook.xml", _WORKBOOK)
        archive.writestr("xl/_rels/workbook.xml.rels", _WORKBOOK_RELS)
        archive.writestr("xl/worksheets/sheet1.xml", _SHEET)
        archive.writestr("xl/sharedStrings.xml", _SHARED_STRINGS)
        archive.writestr("xl/styles.xml", _STYLES)


def _cells_by_coord(document: DocumentIR, sheet_index: int = 0) -> dict:
    table = document.sheets[sheet_index].regions[0].tables[0]
    return {cell.coordinate: cell for cell in table.cells}


def _normalize(document: DocumentIR) -> str:
    data = document.to_dict()
    meta = data.get("metadata", {})
    for key in (
        "ingestor",
        "legacy_fallback",
        "fallback_reason",
        "sparse_stats",
        "asset_directory",
    ):
        meta.pop(key, None)
    # asset uris embed the temp asset dir; drop them for content comparison
    for sheet in data["sheets"]:
        for asset in sheet.get("assets", []):
            asset["uri"] = "<uri>"
    return json.dumps(data, ensure_ascii=False, sort_keys=True)


# --------------------------------------------------------------------------- #
# Equivalence + engine dispatch
# --------------------------------------------------------------------------- #

class SparseLegacyEquivalenceTests(unittest.TestCase):
    FIXTURE_WORKBOOKS = [
        FIXTURES / "workbooks" / "screen-design.xlsx",
        FIXTURES / "workbooks" / "api-spec.xlsx",
    ]

    def test_sparse_matches_legacy_business_content(self) -> None:
        for workbook in self.FIXTURE_WORKBOOKS:
            with self.subTest(workbook=workbook.name):
                with tempfile.TemporaryDirectory() as d1, tempfile.TemporaryDirectory() as d2:
                    legacy = XlsxIngestor(XlsxIngestOptions(asset_dir=Path(d1))).ingest(workbook)
                    sparse = SparseOoxmlIngestor(XlsxIngestOptions(asset_dir=Path(d2))).ingest(workbook)
                self.assertEqual(_normalize(legacy), _normalize(sparse))

    def test_images_and_shapes_preserved(self) -> None:
        workbook = FIXTURES / "workbooks" / "screen-design.xlsx"
        with tempfile.TemporaryDirectory() as d1, tempfile.TemporaryDirectory() as d2:
            legacy = XlsxIngestor(XlsxIngestOptions(asset_dir=Path(d1))).ingest(workbook)
            sparse = SparseOoxmlIngestor(XlsxIngestOptions(asset_dir=Path(d2))).ingest(workbook)
        legacy_assets = [
            (a.asset_type, a.anchor, a.description)
            for sheet in legacy.sheets
            for a in sheet.assets
        ]
        sparse_assets = [
            (a.asset_type, a.anchor, a.description)
            for sheet in sparse.sheets
            for a in sheet.assets
        ]
        self.assertTrue(legacy_assets)
        self.assertEqual(legacy_assets, sparse_assets)

    def test_auto_engine_uses_sparse(self) -> None:
        workbook = FIXTURES / "workbooks" / "api-spec.xlsx"
        with tempfile.TemporaryDirectory() as directory:
            document = ingest_xlsx(workbook, asset_dir=Path(directory), engine="auto")
        self.assertEqual("sparse-ooxml", document.metadata["ingestor"])
        self.assertFalse(document.metadata["legacy_fallback"])

    def test_multi_sheet_relationships(self) -> None:
        workbook = FIXTURES / "workbooks" / "screen-design.xlsx"
        with tempfile.TemporaryDirectory() as directory:
            document = ingest_xlsx(workbook, asset_dir=Path(directory), engine="sparse")
        self.assertEqual(
            [s.name for s in document.sheets],
            ["表紙", "改訂履歴", "画面項目"],
        )
        self.assertEqual([0, 1, 2], [s.index for s in document.sheets])


# --------------------------------------------------------------------------- #
# OOXML feature coverage (crafted workbook)
# --------------------------------------------------------------------------- #

class CraftedWorkbookTests(unittest.TestCase):
    def setUp(self) -> None:
        self._dir = tempfile.TemporaryDirectory()
        self.addCleanup(self._dir.cleanup)
        self.workbook = Path(self._dir.name) / "variety.xlsx"
        _write_crafted_xlsx(self.workbook)
        self.document = ingest_xlsx(
            self.workbook, asset_dir=Path(self._dir.name) / "assets", engine="sparse"
        )
        self.cells = _cells_by_coord(self.document)

    def test_shared_string(self) -> None:
        self.assertEqual("共有文字列", self.cells["A1"].raw_value)

    def test_inline_string(self) -> None:
        self.assertEqual("inline文字", self.cells["A5"].raw_value)

    def test_number(self) -> None:
        self.assertEqual(12, self.cells["A2"].raw_value)

    def test_boolean(self) -> None:
        self.assertTrue(self.cells["A3"].raw_value)

    def test_error_value(self) -> None:
        self.assertEqual("#DIV/0!", self.cells["A4"].raw_value)

    def test_formula_and_cached_from_same_cell(self) -> None:
        cell = self.cells["A6"]
        self.assertEqual("=A2+10", cell.formula)
        # cached <v>22</v> read from the same <c> node -> display "22"
        self.assertEqual("22", cell.display_value)

    def test_merge_records_range_without_materializing_members(self) -> None:
        sparse = SparseOoxmlIngestor(
            XlsxIngestOptions(asset_dir=Path(self._dir.name) / "a2")
        )
        workbook_ir = sparse._build_sparse(
            __import__("openpyxl").load_workbook(self.workbook), self.workbook
        )
        sheet = workbook_ir.sheets[0]
        self.assertIn("B1:C1", sheet.merges)
        # Member C1 is not a value cell in the sparse store.
        self.assertNotIn((1, 3), {(c.row, c.column) for c in sheet.cells.values() if c.raw_value is not None})
        # But C1 is present as an empty member after materialization (no shift).
        self.assertIn("C1", self.cells)
        self.assertEqual("B1", self.cells["C1"].merged_master)
        self.assertIsNone(self.cells["C1"].raw_value)

    def test_missing_cell_is_empty_not_shifted(self) -> None:
        # Row 2 only has A2; B2 must materialize as empty in place.
        self.assertIn("B2", self.cells)
        self.assertIsNone(self.cells["B2"].raw_value)
        self.assertEqual(2, self.cells["B2"].column)

    def test_distant_style_does_not_inflate_content(self) -> None:
        stats = self.document.metadata["sparse_stats"]
        # 7 value cells: A1,B1,A2,A3,A4,A5,A6 ; 1 style-only (XFD1048576)
        self.assertEqual(7, stats["value_cell_count"])
        self.assertEqual(1, stats["style_only_cell_count"])
        # The distant cell is never materialized into the document.
        self.assertNotIn("XFD1048576", self.cells)
        # Content stays a tight box, never the inflated dimension.
        self.assertLess(len(self.cells), 50)
        self.assertTrue(all(cell.row <= 6 for cell in self.cells.values()))


# --------------------------------------------------------------------------- #
# Stress fixture (generated, not committed)
# --------------------------------------------------------------------------- #

class StressFixtureTests(unittest.TestCase):
    def _build_stress(self, path: Path) -> None:
        from openpyxl.styles import Font

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Stress"
        for row in range(1, 11):
            sheet.cell(row, 1, f"A{row}")
            sheet.cell(row, 2, row)
        sheet.merge_cells("A1:B1")
        sheet["A11"] = "=SUM(B1:B10)"
        # A lone styled cell at the far corner + whole-column format.
        far = sheet.cell(row=1048576, column=16384)
        far.font = Font(bold=True)
        sheet.column_dimensions["H"].font = Font(bold=True)
        workbook.save(path)

    def test_distant_style_and_inflated_dimension_stay_sparse(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook = Path(directory) / "stress.xlsx"
            self._build_stress(workbook)
            document = ingest_xlsx(
                workbook, asset_dir=Path(directory) / "assets", engine="sparse"
            )
        stats = document.metadata["sparse_stats"]
        # 20 real value cells: A1..A10 (10) + B2..B10 (9, B1 cleared by the
        # A1:B1 merge) + A11 formula (1). Never the millions the inflated
        # dimension / distant styled cell would imply.
        self.assertEqual(20, stats["value_cell_count"])
        self.assertEqual(1, stats["style_only_cell_count"])
        cells = _cells_by_coord(document)
        # Materialized grid ~ content box (rows 1-11 x cols 1-2), never millions.
        self.assertLessEqual(len(cells), 40)
        self.assertNotIn("XFD1048576", cells)
        self.assertTrue(all(cell.row <= 11 for cell in cells.values()))


# --------------------------------------------------------------------------- #
# Fallback boundary
# --------------------------------------------------------------------------- #

class FallbackTests(unittest.TestCase):
    def _workbook(self, directory: str) -> Path:
        path = Path(directory) / "w.xlsx"
        book = Workbook()
        book.active.title = "S"
        book.active.append(["A", "B"])
        book.save(path)
        return path

    def test_auto_falls_back_to_legacy_on_unsupported(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook = self._workbook(directory)
            options = XlsxIngestOptions(asset_dir=Path(directory) / "assets")
            with mock.patch.object(
                SparseOoxmlIngestor,
                "ingest",
                side_effect=UnsupportedWorkbookError("simulated unsupported feature"),
            ):
                document = ingest_with_engine(workbook, options, engine="auto")
        self.assertTrue(document.metadata["legacy_fallback"])
        self.assertIn("UnsupportedWorkbookError", document.metadata["fallback_reason"])
        codes = {d.code for d in document.diagnostics}
        self.assertIn("INGEST_LEGACY_FALLBACK", codes)

    def test_auto_does_not_swallow_bugs(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workbook = self._workbook(directory)
            options = XlsxIngestOptions(asset_dir=Path(directory) / "assets")
            with mock.patch.object(
                SparseOoxmlIngestor,
                "ingest",
                side_effect=RuntimeError("a real bug, not an unsupported workbook"),
            ):
                with self.assertRaises(RuntimeError):
                    ingest_with_engine(workbook, options, engine="auto")


# --------------------------------------------------------------------------- #
# SheetIndex integration
# --------------------------------------------------------------------------- #

class SheetIndexIntegrationTests(unittest.TestCase):
    def test_scan_built_once_per_sheet(self) -> None:
        from excelspec.pipeline import run_pipeline

        workbook = FIXTURES / "workbooks" / "screen-design.xlsx"
        template = FIXTURES / "templates" / "screen-design.yaml"
        import excelspec.templates.engine as engine

        original = engine._scan_sheet
        calls: list[str] = []

        def _counting_scan(sheet):
            calls.append(sheet.name)
            return original(sheet)

        with tempfile.TemporaryDirectory() as directory:
            with mock.patch.object(engine, "_scan_sheet", _counting_scan):
                run_pipeline(workbook, template=template, asset_dir=Path(directory))
        # Every sheet is scanned exactly once (no per-region rebuild).
        self.assertEqual(sorted(calls), sorted(set(calls)))
        self.assertEqual(3, len(calls))

    def test_locate_regions_scan_matches_unscanned(self) -> None:
        from excelspec.templates.engine import _scan_sheet, locate_regions
        from excelspec.models.template import LocatorMode, RegionLocator

        workbook = FIXTURES / "workbooks" / "screen-design.xlsx"
        with tempfile.TemporaryDirectory() as directory:
            document = ingest_xlsx(workbook, asset_dir=Path(directory), engine="sparse")
        sheet = document.sheets[2]
        locator = RegionLocator(mode=LocatorMode.ANCHOR, anchor_pattern="項目ID")
        scan = _scan_sheet(sheet)
        self.assertEqual(
            locate_regions(sheet, locator),
            locate_regions(sheet, locator, scan=scan),
        )


if __name__ == "__main__":
    unittest.main()
