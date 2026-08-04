from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from pathlib import Path

import jsonschema
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from excelspec.ingest import ingest_xlsx
from excelspec.ingest.ooxml import extract_sheet_drawings
from excelspec.schemas import load_schema


class XlsxIngestTests(unittest.TestCase):
    def test_cells_merges_formula_style_and_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "spec.xlsx"
            screenshot_path = root / "screen.png"
            screenshot_path.write_bytes(b"png")

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "画面"
            sheet["A1"] = "标题"
            sheet["A1"].font = Font(bold=True)
            sheet["A1"].fill = PatternFill("solid", fgColor="FFFF00")
            sheet.merge_cells("A1:B1")
            sheet["A2"] = "=1+1"
            workbook.save(workbook_path)
            workbook.close()

            manifest_path = root / "screenshots.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "version": "1",
                        "assets": [
                            {
                                "asset_id": "screen-main",
                                "path": "screen.png",
                                "sheet": "画面",
                                "region_id": "raw-grid",
                                "description": "主画面",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            document = ingest_xlsx(
                workbook_path,
                asset_dir=root / "assets",
                screenshot_manifest=manifest_path,
            )
            sheet_ir = document.sheets[0]
            cells = {
                cell.coordinate: cell
                for cell in sheet_ir.regions[0].tables[0].cells
            }

            self.assertEqual(cells["A1"].row_span, 1)
            self.assertEqual(cells["A1"].col_span, 2)
            self.assertEqual(cells["B1"].merged_master, "A1")
            self.assertTrue(cells["A1"].style.font["bold"])
            self.assertEqual(cells["A2"].formula, "=1+1")
            self.assertIsNone(cells["A2"].display_value)
            self.assertIn(
                "FORMULA_CACHE_MISSING",
                {diagnostic.code for diagnostic in sheet_ir.diagnostics},
            )
            self.assertEqual(sheet_ir.assets[0].asset_id, "screen-main")
            self.assertEqual(sheet_ir.regions[0].asset_ids, ["screen-main"])
            self.assertEqual(sheet_ir.assets[0].metadata["ocr"]["status"], "pending")
            self.assertEqual(sheet_ir.assets[0].metadata["vlm"]["status"], "pending")
            jsonschema.validate(document.to_dict(), load_schema("document-ir"))

    def test_drawing_image_and_shape_text_extraction(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            package_path = root / "drawing.zip"
            sheet_xml = """\
<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <drawing r:id="rId1"/>
</worksheet>"""
            sheet_rels = """\
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Target="../drawings/drawing1.xml"/>
</Relationships>"""
            drawing_xml = """\
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
 xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
 <xdr:twoCellAnchor>
  <xdr:from><xdr:col>1</xdr:col><xdr:row>2</xdr:row></xdr:from>
  <xdr:to><xdr:col>3</xdr:col><xdr:row>5</xdr:row></xdr:to>
  <xdr:pic><xdr:nvPicPr><xdr:cNvPr name="画面图" descr="截图"/></xdr:nvPicPr>
   <xdr:blipFill><a:blip r:embed="rIdImg"/></xdr:blipFill></xdr:pic>
  <xdr:sp><xdr:txBody><a:p><a:r><a:t>形状文字</a:t></a:r></a:p></xdr:txBody></xdr:sp>
 </xdr:twoCellAnchor>
</xdr:wsDr>"""
            drawing_rels = """\
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rIdImg" Target="../media/image1.png"/>
</Relationships>"""
            with zipfile.ZipFile(package_path, "w") as archive:
                archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
                archive.writestr(
                    "xl/worksheets/_rels/sheet1.xml.rels", sheet_rels
                )
                archive.writestr("xl/drawings/drawing1.xml", drawing_xml)
                archive.writestr(
                    "xl/drawings/_rels/drawing1.xml.rels", drawing_rels
                )
                archive.writestr("xl/media/image1.png", b"image")

            with zipfile.ZipFile(package_path) as archive:
                assets, diagnostics = extract_sheet_drawings(
                    archive,
                    sheet_part="xl/worksheets/sheet1.xml",
                    sheet_name="画面",
                    output_dir=root / "assets",
                )

            self.assertFalse(diagnostics)
            self.assertEqual([asset.kind for asset in assets], ["image", "shape"])
            self.assertEqual(assets[0].anchor, "B3:D6")
            self.assertEqual(Path(assets[0].uri).read_bytes(), b"image")
            self.assertEqual(assets[1].description, "形状文字")


if __name__ == "__main__":
    unittest.main()
