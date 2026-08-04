"""Create two multi-sheet Japanese design-doc demos matching the provided screenshots."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw

ROOT = Path(__file__).resolve().parent
WORKBOOKS = ROOT / "workbooks"
ASSETS = ROOT / "assets"
TEMPLATES = ROOT / "templates"

GREEN = PatternFill("solid", fgColor="92D050")
LIGHT_GREEN = PatternFill("solid", fgColor="C6EFCE")
HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
TITLE_FILL = PatternFill("solid", fgColor="D9EAD3")
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set(cell, value, *, fill=None, bold=False, align=CENTER, size=11):
    cell.value = value
    cell.font = Font(name="Yu Gothic", size=size, bold=bold)
    cell.alignment = align
    cell.border = THIN
    if fill is not None:
        cell.fill = fill


def _merge_fill(ws, range_ref: str, value, *, fill=None, bold=False, align=CENTER, size=11):
    ws.merge_cells(range_ref)
    top_left = range_ref.split(":")[0]
    _set(ws[top_left], value, fill=fill, bold=bold, align=align, size=size)


def _meta_block(ws, start_row: int, pairs: list[tuple[str, str]]) -> None:
    row = start_row
    for label, value in pairs:
        _merge_fill(ws, f"A{row}:B{row}", label, fill=GREEN, bold=True)
        _merge_fill(ws, f"C{row}:F{row}", value, align=LEFT)
        row += 1


def _make_mock_image(path: Path, title: str, subtitle: str, kind: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    image = Image.new("RGB", (720, 420), "#f7f9fc")
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, 719, 48), fill="#92D050")
    draw.text((18, 14), title, fill="#103010")
    draw.rectangle((20, 70, 700, 110), fill="#ffffff", outline="#708090")
    draw.text((30, 82), subtitle, fill="#203040")
    if kind == "screen":
        draw.rectangle((20, 130, 120, 160), fill="#4f81bd", outline="#315f91")
        draw.text((42, 138), "新規", fill="white")
        draw.rectangle((140, 130, 240, 160), fill="#4f81bd", outline="#315f91")
        draw.text((162, 138), "検索", fill="white")
        headers = ["選択", "保証番号", "被保証者名", "保証金額", "受付日", "状態"]
        x = 20
        for header in headers:
            draw.rectangle((x, 180, x + 110, 210), fill="#d9ead3", outline="#708090")
            draw.text((x + 8, 188), header, fill="#203040")
            x += 110
        for i in range(4):
            y = 210 + i * 32
            x = 20
            values = ["□", f"G-00{i+1}", f"顧客{i+1}", f"{(i+1)*100000}", f"2025/0{i+1}/10", "審査中"]
            for value in values:
                draw.rectangle((x, y, x + 110, y + 32), fill="white", outline="#708090")
                draw.text((x + 8, y + 8), value, fill="#203040")
                x += 110
    else:
        draw.ellipse((40, 140, 70, 170), fill="#2b9e9e")
        draw.text((50, 147), "1", fill="white")
        draw.rectangle((80, 140, 280, 170), fill="white", outline="#708090")
        draw.text((90, 148), "受付日", fill="#203040")
        draw.ellipse((40, 190, 70, 220), fill="#2b9e9e")
        draw.text((50, 197), "2", fill="white")
        draw.rectangle((80, 190, 280, 220), fill="white", outline="#708090")
        draw.text((90, 198), "タイトル", fill="#203040")
        draw.line((300, 155, 420, 205), fill="#708090", width=2)
        draw.rectangle((40, 260, 680, 380), fill="white", outline="#708090")
        draw.text((50, 275), "証券番号 / 残高 / 備考", fill="#203040")
        draw.text((50, 310), "RPT-B0020 帳票レイアウト（モック）", fill="#406080")
    image.save(path, format="PNG")


def build_screen_design(path: Path, layout_image: Path) -> None:
    wb = Workbook()

    # 表紙
    cover = wb.active
    cover.title = "表紙"
    for col in range(1, 20):
        cover.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 4.5
    for col in range(1, 12):
        cover.column_dimensions[chr(64 + col)].width = 8
    _merge_fill(
        cover,
        "B2:K3",
        "林業業務システム更改に係る設計・開発",
        fill=TITLE_FILL,
        bold=True,
        size=14,
    )
    _merge_fill(cover, "B5:K5", "債務保証_平時", bold=True, size=16)
    _merge_fill(cover, "B6:K6", "基本設計", size=14)
    _merge_fill(cover, "B7:K7", "SCR-A0010", bold=True, size=14)
    _merge_fill(cover, "B8:K8", "保証一覧", bold=True, size=16)
    _merge_fill(cover, "B9:K9", "画面設計書", bold=True, size=16)
    meta = [
        ("文書番号", "文書番号を取得出来ませんでした。"),
        ("バージョン", "0.01"),
        ("作成者", "OKI"),
        ("作成日", "2026/3/24"),
    ]
    for i, (label, value) in enumerate(meta):
        row = 11 + i
        _merge_fill(cover, f"D{row}:E{row}", label, fill=GREEN, bold=True)
        _merge_fill(cover, f"F{row}:I{row}", value, align=LEFT)

    # 修正履歴
    hist = wb.create_sheet("修正履歴")
    _merge_fill(hist, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(hist, "A3:B4", "修正履歴", fill=GREEN, bold=True, size=14)
    header_meta = [
        ("プロダクト", "林業業務システム更改に係る設計・開発"),
        ("サブシステム", "債務保証_平時"),
        ("機能名", "一覧"),
        ("画面ID", "SCR-A0010"),
        ("作成日", "2025/3/24"),
        ("更新日", "2025/7/30"),
        ("作成者", "OKI"),
        ("更新者", "OKI"),
    ]
    for i, (label, value) in enumerate(header_meta):
        col = 4 + (i % 4) * 2
        row = 2 + i // 4
        _set(hist.cell(row, col), label, fill=LIGHT_GREEN, bold=True)
        _set(hist.cell(row, col + 1), value, align=LEFT)
    headers = ["No.", "版数", "シート", "修正箇所", "修正内容", "修正者", "修正日"]
    for col, h in enumerate(headers, start=1):
        _set(hist.cell(6, col), h, fill=GREEN, bold=True)
    rows = [
        (1, "0.80", "画面入出力項目一覧", "-", "初期作成", "OKI", "2025/3/24"),
        (2, "0.90", "画面アクション一覧", "-", "イベント概要を追加", "OKI", "2025/5/10"),
        (3, "0.91", "画面入出力項目一覧", "No.8", "必須判定を見直し", "OKI", "2025/7/30"),
        (4, "0.95", "画面レイアウト", "-", "検索条件エリアを更新", "OKI", "2025/8/15"),
        (5, "1.00", "入力チェック", "-", "桁数チェック追加", "OKI", "2025/9/01"),
    ]
    for r, values in enumerate(rows, start=7):
        for c, value in enumerate(values, start=1):
            _set(hist.cell(r, c), value, align=LEFT if c in (3, 5) else CENTER)

    # 画面レイアウト
    layout = wb.create_sheet("画面レイアウト")
    _merge_fill(layout, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(layout, "A3:B4", "画面レイアウト", fill=GREEN, bold=True, size=14)
    for i, (label, value) in enumerate(
        [
            ("プロダクト", "林業業務システム更改に係る設計・開発"),
            ("サブシステム", "債務保証_平時"),
            ("画面名", "保証審査一覧"),
            ("画面ID", "SCR-A0010"),
            ("作成日", "2025/3/24"),
            ("更新日", "2025/7/30"),
        ]
    ):
        col = 4 + (i % 3) * 2
        row = 2 + i // 3
        _set(layout.cell(row, col), label, fill=LIGHT_GREEN, bold=True)
        _set(layout.cell(row, col + 1), value, align=LEFT)
    _merge_fill(layout, "A6:H6", "■画面イメージ", bold=True, align=LEFT)
    picture = XlsxImage(layout_image)
    picture.width = 480
    picture.height = 280
    picture.anchor = "A8"
    layout.add_image(picture)
    _merge_fill(layout, "A28:H28", "※画面モックは埋め込み画像として保持", align=LEFT)

    # 画面入出力項目一覧
    io = wb.create_sheet("画面入出力項目一覧")
    _merge_fill(io, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(io, "A3:B4", "画面入出力項目一覧", fill=GREEN, bold=True, size=13)
    for i, (label, value) in enumerate(
        [
            ("プロダクト", "林業業務システム更改に係る設計・開発"),
            ("サブシステム", "債務保証_平時"),
            ("画面名", "保証一覧"),
            ("画面ID", "SCR-A0010"),
        ]
    ):
        _set(io.cell(2 + i, 4), label, fill=LIGHT_GREEN, bold=True)
        _set(io.cell(2 + i, 5), value, align=LEFT)
    _merge_fill(io, "A6:M6", "凡例: ●=表示/必須  ○=任意  ▲=条件表示  x=非活性  -=対象外", align=LEFT)
    headers = [
        "No.",
        "項目名称",
        "種別",
        "全半角",
        "桁数",
        "I/O",
        "表示",
        "活性",
        "必須",
        "データ元",
        "初期値",
        "イベントID",
        "補足説明",
    ]
    for col, h in enumerate(headers, start=1):
        _set(io.cell(8, col), h, fill=GREEN, bold=True)
        io.column_dimensions[chr(64 + col)].width = 12
    io.column_dimensions["B"].width = 18
    io.column_dimensions["J"].width = 28
    io.column_dimensions["M"].width = 22
    items = [
        (1, "画面タイトル", "ラベル", "-", "-", "O", "●", "x", "-", "固定文言", "保証一覧", "-", "ヘッダ表示"),
        (2, "新規", "ボタン", "-", "-", "I", "●", "●", "-", "-", "-", "EV01", "基本情報入力へ遷移"),
        (3, "検索", "ボタン", "-", "-", "I", "●", "●", "-", "-", "-", "EV02", "一覧再検索"),
        (4, "保証番号", "テキスト", "半角", 20, "I/O", "●", "●", "○", "保証番号入力値", "", "-", "検索条件"),
        (5, "被保証者名", "テキスト", "全角", 40, "I/O", "●", "●", "○", "被保証者名入力値", "", "-", "部分一致"),
        (6, "保証金額", "テキスト", "半角", 12, "O", "●", "x", "-", "保証テーブル.金額", "-", "-", "一覧表示"),
        (7, "受付日", "日付", "-", "-", "O", "●", "x", "-", "保証テーブル.受付日", "-", "-", "YYYY/MM/DD"),
        (
            8,
            "保証割合",
            "ラベル",
            "-",
            "-",
            "O",
            "▲",
            "x",
            "-",
            "IF 保証更新入力の場合:\n  更新割合を表示\nELSE:\n  初期割合を表示",
            "-",
            "-",
            "条件表示",
        ),
        (9, "選択", "チェック", "-", "-", "I", "●", "●", "○", "-", "OFF", "EV03", "複数選択可"),
        (10, "状態", "ラベル", "-", "-", "O", "●", "x", "-", "保証テーブル.状態", "-", "-", "審査中/承認済"),
    ]
    for r, values in enumerate(items, start=9):
        for c, value in enumerate(values, start=1):
            _set(io.cell(r, c), value, align=LEFT if c in (2, 10, 13) else CENTER)
    # Emulate multi-row merged item without overwriting following records.
    io.insert_rows(17, 2)
    io.merge_cells("A16:A18")
    io.merge_cells("B16:B18")
    io["J16"].alignment = LEFT
    io.row_dimensions[16].height = 18
    io.row_dimensions[17].height = 18
    io.row_dimensions[18].height = 18

    # 画面アクション一覧
    actions = wb.create_sheet("画面アクション一覧")
    _merge_fill(actions, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(actions, "A3:B4", "画面アクション一覧", fill=GREEN, bold=True, size=13)
    for i, (label, value) in enumerate(
        [
            ("プロダクト", "林業業務システム更改に係る設計・開発"),
            ("サブシステム", "債務保証_平時"),
            ("画面名", "保証一覧"),
            ("画面ID", "SCR-A0010"),
        ]
    ):
        _set(actions.cell(2 + i, 4), label, fill=LIGHT_GREEN, bold=True)
        _set(actions.cell(2 + i, 5), value, align=LEFT)
    headers = ["No.", "イベントID", "イベント名", "画面項目名", "種別", "トリガー", "イベント概要"]
    for col, h in enumerate(headers, start=1):
        _set(actions.cell(7, col), h, fill=GREEN, bold=True)
    actions.column_dimensions["C"].width = 22
    actions.column_dimensions["G"].width = 42
    action_rows = [
        (1, "EV01", "初期処理", "-", "-", "初期表示時", "初期表示内容を設定し、検索条件をクリアする。"),
        (2, "EV02", "新規ボタン押下処理", "新規", "ボタン", "押下時", "基本情報入力画面へ遷移する。"),
        (3, "EV03", "検索ボタン押下処理", "検索", "ボタン", "押下時", "入力条件で保証一覧を再検索し、結果を表示する。"),
        (4, "EV04", "行選択処理", "選択", "チェック", "変更時", "選択行を保持し、一括操作対象とする。"),
        (5, "EV05", "状態プルダウン変更", "状態", "プルダウンリスト", "変更時", "表示対象状態を絞り込み、一覧を更新する。"),
    ]
    for r, values in enumerate(action_rows, start=8):
        for c, value in enumerate(values, start=1):
            _set(actions.cell(r, c), value, align=LEFT if c in (3, 7) else CENTER)

    # 入力チェック
    checks = wb.create_sheet("入力チェック")
    _merge_fill(checks, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(checks, "A3:B4", "入力チェック", fill=GREEN, bold=True, size=13)
    headers = ["No.", "項目名称", "チェック内容", "エラーメッセージ", "イベントID"]
    for col, h in enumerate(headers, start=1):
        _set(checks.cell(6, col), h, fill=GREEN, bold=True)
    checks.column_dimensions["C"].width = 36
    checks.column_dimensions["D"].width = 36
    check_rows = [
        (1, "保証番号", "半角英数字20桁以内", "保証番号の形式が不正です。", "EV03"),
        (2, "被保証者名", "全角40文字以内", "被保証者名は全角40文字以内で入力してください。", "EV03"),
        (3, "受付日", "日付形式 YYYY/MM/DD", "受付日の形式が不正です。", "EV03"),
        (4, "選択", "1件以上選択", "対象を1件以上選択してください。", "EV04"),
    ]
    for r, values in enumerate(check_rows, start=7):
        for c, value in enumerate(values, start=1):
            _set(checks.cell(r, c), value, align=LEFT if c in (2, 3, 4) else CENTER)

    wb.save(path)
    wb.close()


def build_report_design(path: Path, layout_image: Path) -> None:
    wb = Workbook()

    cover = wb.active
    cover.title = "表紙"
    for col in range(1, 12):
        cover.column_dimensions[chr(64 + col)].width = 8
    _merge_fill(
        cover,
        "B2:K3",
        "林業業務システム更改に係る設計・開発",
        fill=TITLE_FILL,
        bold=True,
        size=14,
    )
    _merge_fill(cover, "B5:K5", "債務保証_有事", bold=True, size=16)
    _merge_fill(cover, "B6:K6", "基本設計", size=14)
    _merge_fill(cover, "B7:K7", "RPT-B0020", bold=True, size=14)
    _merge_fill(cover, "B8:K8", "予見・事故記録簿", bold=True, size=16)
    _merge_fill(cover, "B9:K9", "帳票設計書", bold=True, size=16)
    for i, (label, value) in enumerate(
        [
            ("文書番号", "文書番号を取得出来ませんでした。"),
            ("バージョン", "0.90"),
            ("作成者", "OKI"),
            ("作成日", "2024/4/9"),
        ]
    ):
        row = 11 + i
        _merge_fill(cover, f"D{row}:E{row}", label, fill=GREEN, bold=True)
        _merge_fill(cover, f"F{row}:I{row}", value, align=LEFT)

    hist = wb.create_sheet("修正履歴")
    _merge_fill(hist, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(hist, "A3:B4", "修正履歴", fill=GREEN, bold=True, size=14)
    for i, (label, value) in enumerate(
        [
            ("プロジェクト", "林業業務システム更改に係る設計・開発"),
            ("サブシステム", "債務保証_有事"),
            ("帳票名", "予見・事故記録簿"),
            ("帳票ID", "RPT-B0020"),
            ("作成日", "2023/4/9"),
            ("作成者", "OKI"),
        ]
    ):
        col = 4 + (i % 3) * 2
        row = 2 + i // 3
        _set(hist.cell(row, col), label, fill=LIGHT_GREEN, bold=True)
        _set(hist.cell(row, col + 1), value, align=LEFT)
    headers = ["No.", "版数", "シート", "修正箇所", "修正内容", "修正者", "修正日"]
    for col, h in enumerate(headers, start=1):
        _set(hist.cell(6, col), h, fill=GREEN, bold=True)
    for c, value in enumerate(
        (1, "0.80", "帳票概要", "-", "初期作成", "OKI", "2023/4/9"), start=1
    ):
        _set(hist.cell(7, c), value, align=LEFT if c in (3, 5) else CENTER)

    overview = wb.create_sheet("帳票概要")
    _merge_fill(overview, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(overview, "A3:B4", "帳票概要", fill=GREEN, bold=True, size=14)
    for i, (label, value) in enumerate(
        [
            ("プロジェクト", "林業業務システム更改に係る設計・開発"),
            ("サブシステム", "債務保証_有事"),
            ("帳票名", "予見・事故記録簿"),
            ("帳票ID", "RPT-B0020"),
            ("作成日", "2023/4/9"),
            ("作成者", "OKI"),
        ]
    ):
        col = 4 + (i % 3) * 2
        row = 2 + i // 3
        _set(overview.cell(row, col), label, fill=LIGHT_GREEN, bold=True)
        _set(overview.cell(row, col + 1), value, align=LEFT)

    specs = [
        ("機能ID", "FNC-B0040"),
        ("機能名", "予見・事故記録出力"),
        ("画面ID・バッチID", "SCR-B0040"),
        ("出力様式", "EXCEL, PDF"),
        ("帳票サイズ", "A4"),
        ("帳票方向", "横"),
        ("出力枚数", "2P / 回"),
    ]
    for i, (label, value) in enumerate(specs):
        row = 6 + i
        _merge_fill(overview, f"A{row}:B{row}", label, fill=GREEN, bold=True)
        _merge_fill(overview, f"C{row}:F{row}", value, align=LEFT)

    _merge_fill(overview, "A14:F14", "■帳票概要", bold=True, align=LEFT)
    _merge_fill(
        overview,
        "A15:F17",
        "1. 帳票内容：予見または事故情報を参照し、記録簿を出力する。\n"
        "2. 並び順：なし\n"
        "3. 出力条件：保証審査画面から出力指示された場合に生成する。",
        align=LEFT,
    )

    layout = wb.create_sheet("帳票レイアウト")
    _merge_fill(layout, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(layout, "A3:B4", "帳票レイアウト", fill=GREEN, bold=True, size=14)
    for i, (label, value) in enumerate(
        [
            ("プロジェクト", "林業業務システム更改に係る設計・開発"),
            ("サブシステム", "債務保証_有事"),
            ("帳票名", "予見・事故記録簿"),
            ("帳票ID", "RPT-B0020"),
        ]
    ):
        col = 4 + (i % 2) * 2
        row = 2 + i // 2
        _set(layout.cell(row, col), label, fill=LIGHT_GREEN, bold=True)
        _set(layout.cell(row, col + 1), value, align=LEFT)
    _merge_fill(layout, "A6:H6", "■帳票イメージ（番号付き項目は定義表と対応）", bold=True, align=LEFT)
    picture = XlsxImage(layout_image)
    picture.width = 480
    picture.height = 280
    picture.anchor = "A8"
    layout.add_image(picture)
    _merge_fill(layout, "A28:H28", "※円番号と線はモック画像内に表現", align=LEFT)

    define = wb.create_sheet("帳票編集定義")
    _merge_fill(define, "A2:B2", "基本設計", fill=GREEN, bold=True)
    _merge_fill(define, "A3:B4", "帳票編集定義", fill=GREEN, bold=True, size=14)
    for i, (label, value) in enumerate(
        [
            ("プロジェクト", "林業業務システム更改に係る設計・開発"),
            ("サブシステム", "債務保証_有事"),
            ("帳票名", "予見・事故記録簿"),
            ("帳票ID", "RPT-B0020"),
        ]
    ):
        _set(define.cell(2 + i, 4), label, fill=LIGHT_GREEN, bold=True)
        _set(define.cell(2 + i, 5), value, align=LEFT)

    _merge_fill(define, "A7:J7", "1. 改ページ条件：明細がページを超えた場合\n2. ヘッダ条件：各ページ先頭に帳票タイトルを出力\n3. フッタ条件：ページ番号を出力\n4. 項目概要", align=LEFT)
    headers = [
        "No.",
        "項目名",
        "全/半角",
        "桁数",
        "書体",
        "フォントサイズ",
        "水平配置",
        "垂直配置",
        "データ元",
        "備考",
    ]
    for col, h in enumerate(headers, start=1):
        _set(define.cell(12, col), h, fill=GREEN, bold=True)
    define.column_dimensions["B"].width = 14
    define.column_dimensions["I"].width = 28
    define.column_dimensions["J"].width = 28
    define_rows = [
        (1, "受付日", "半角", 10, "MS P明朝", 10, "中央揃え", "中央揃え", "事故情報.受付日", "【フォーマット】和暦 ggge年m月d日"),
        (2, "起票日", "半角", 10, "MS P明朝", 10, "中央揃え", "中央揃え", "事故情報.起票日", "【フォーマット】和暦"),
        (3, "タイトル", "全角", 40, "MS P明朝", 18, "中央揃え", "中央揃え", "固定文言", "事故（延滞）報告"),
        (4, "証券番号", "半角", 20, "MS P明朝", 9, "左詰め", "中央揃え", "保証テーブル.証券番号", ""),
        (5, "残高", "半角", 12, "MS P明朝", 9, "右詰め", "中央揃え", "保証テーブル.残高", "カンマ編集"),
        (6, "備考", "全角", 200, "MS P明朝", 9, "左詰め", "上詰め", "事故情報.備考", "改行可"),
    ]
    for r, values in enumerate(define_rows, start=13):
        for c, value in enumerate(values, start=1):
            _set(define.cell(r, c), value, align=LEFT if c in (2, 9, 10) else CENTER)

    wb.save(path)
    wb.close()


SCREEN_TEMPLATE = """\
schema_version: "1.0"
template_id: demo-screen-design
version: "1.0"
name: 画面設計書デモ
match:
  sheet_name_patterns: ["^表紙$", "^修正履歴$", "^画面入出力項目一覧$", "^画面アクション一覧$"]
  fingerprints:
    - sheet_name_pattern: "^表紙$"
      required_text: ["画面設計書", "SCR-A0010"]
    - sheet_name_pattern: "^画面入出力項目一覧$"
      required_text: ["項目名称", "データ元"]
  minimum_score: 0.55
sheets:
  - sheet_id: cover
    name_pattern: "^表紙$"
    order: 10
    regions:
      - region_id: document-info
        region_type: key_value
        title: 表紙
        required: true
        locator: {mode: fixed, range: "D11:I14"}
        extractor:
          kind: key_value
          key_column: 1
          value_column: 3
          key_semantics:
            文書番号: document_no
            バージョン: version
            作成者: author
            作成日: created_at
  - sheet_id: revisions
    name_pattern: "^修正履歴$"
    order: 20
    regions:
      - region_id: revision-table
        region_type: table
        title: 修正履歴
        locator:
          mode: anchor
          anchor_text: "No."
          width: 7
        extractor:
          kind: table
          header_rows: 1
          column_semantics:
            'No.': seq_no
            版数: revision
            シート: sheet_name
            修正箇所: location
            修正内容: change
            修正者: author
            修正日: changed_at
          options: {stop_after_blank_rows: 1, shrink_to_content: true, trim_empty_columns: true}
  - sheet_id: layout
    name_pattern: "^画面レイアウト$"
    order: 30
    regions:
      - region_id: screen-layout
        region_type: layout
        title: 画面レイアウト
        locator:
          mode: anchor
          anchor_pattern: "^■画面イメージ"
          row_offset: 1
          height: 20
          width: 8
        extractor: {kind: asset}
  - sheet_id: io-items
    name_pattern: "^画面入出力項目一覧$"
    order: 40
    regions:
      - region_id: io-table
        region_type: table
        title: 画面入出力項目一覧
        required: true
        locator:
          mode: anchor
          anchor_text: "No."
          width: 13
        extractor:
          kind: table
          header_rows: 1
          column_semantics:
            'No.': seq_no
            項目名称: field_name
            種別: control_type
            全半角: charset
            桁数: length
            "I/O": io
            表示: visible
            活性: enabled
            必須: required
            データ元: data_source
            初期値: default_value
            イベントID: event_id
            補足説明: remarks
          options: {stop_after_blank_rows: 1, shrink_to_content: true, trim_empty_columns: true}
  - sheet_id: actions
    name_pattern: "^画面アクション一覧$"
    order: 50
    regions:
      - region_id: action-table
        region_type: table
        title: 画面アクション一覧
        locator:
          mode: anchor
          anchor_text: "No."
          width: 7
        extractor:
          kind: table
          header_rows: 1
          column_semantics:
            'No.': seq_no
            イベントID: event_id
            イベント名: event_name
            画面項目名: field_name
            種別: control_type
            トリガー: trigger
            イベント概要: overview
          options: {stop_after_blank_rows: 1, shrink_to_content: true, trim_empty_columns: true}
  - sheet_id: checks
    name_pattern: "^入力チェック$"
    order: 60
    regions:
      - region_id: check-table
        region_type: table
        title: 入力チェック
        locator:
          mode: anchor
          anchor_text: "No."
          width: 5
        extractor:
          kind: table
          header_rows: 1
          column_semantics:
            'No.': seq_no
            項目名称: field_name
            チェック内容: validation
            エラーメッセージ: error_message
            イベントID: event_id
          options: {stop_after_blank_rows: 1, shrink_to_content: true, trim_empty_columns: true}
validation_rules: []
metadata:
  document_kind: screen-design
  demo: true
"""

REPORT_TEMPLATE = """\
schema_version: "1.0"
template_id: demo-report-design
version: "1.0"
name: 帳票設計書デモ
match:
  sheet_name_patterns: ["^表紙$", "^帳票概要$", "^帳票編集定義$"]
  fingerprints:
    - sheet_name_pattern: "^表紙$"
      required_text: ["帳票設計書", "RPT-B0020"]
    - sheet_name_pattern: "^帳票編集定義$"
      required_text: ["項目名", "データ元"]
  minimum_score: 0.55
sheets:
  - sheet_id: cover
    name_pattern: "^表紙$"
    order: 10
    regions:
      - region_id: document-info
        region_type: key_value
        title: 表紙
        required: true
        locator: {mode: fixed, range: "D11:I14"}
        extractor:
          kind: key_value
          key_column: 1
          value_column: 3
          key_semantics:
            文書番号: document_no
            バージョン: version
            作成者: author
            作成日: created_at
  - sheet_id: revisions
    name_pattern: "^修正履歴$"
    order: 20
    regions:
      - region_id: revision-table
        region_type: table
        title: 修正履歴
        locator:
          mode: anchor
          anchor_text: "No."
          width: 7
        extractor:
          kind: table
          header_rows: 1
          column_semantics:
            'No.': seq_no
            版数: revision
            シート: sheet_name
            修正箇所: location
            修正内容: change
            修正者: author
            修正日: changed_at
          options: {stop_after_blank_rows: 1, shrink_to_content: true, trim_empty_columns: true}
  - sheet_id: overview
    name_pattern: "^帳票概要$"
    order: 30
    regions:
      - region_id: report-specs
        region_type: key_value
        title: 帳票仕様
        locator: {mode: fixed, range: "A6:F12"}
        extractor:
          kind: key_value
          key_column: 1
          value_column: 3
          key_semantics:
            機能ID: function_id
            機能名: function_name
            "画面ID・バッチID": screen_or_batch_id
            出力様式: output_format
            帳票サイズ: paper_size
            帳票方向: orientation
            出力枚数: page_count
      - region_id: report-summary
        region_type: freeform
        title: 帳票概要説明
        locator:
          mode: anchor
          anchor_text: "■帳票概要"
          row_offset: 1
          height: 4
          width: 6
        extractor: {kind: freeform}
  - sheet_id: layout
    name_pattern: "^帳票レイアウト$"
    order: 40
    regions:
      - region_id: report-layout
        region_type: layout
        title: 帳票レイアウト
        locator:
          mode: anchor
          anchor_pattern: "^■帳票イメージ"
          row_offset: 1
          height: 20
          width: 8
        extractor: {kind: asset}
  - sheet_id: definitions
    name_pattern: "^帳票編集定義$"
    order: 50
    regions:
      - region_id: field-table
        region_type: table
        title: 帳票項目定義
        required: true
        locator:
          mode: anchor
          anchor_text: "No."
          width: 10
        extractor:
          kind: table
          header_rows: 1
          column_semantics:
            'No.': seq_no
            項目名: field_name
            "全/半角": charset
            桁数: length
            書体: font_family
            フォントサイズ: font_size
            水平配置: h_align
            垂直配置: v_align
            データ元: data_source
            備考: remarks
          options: {stop_after_blank_rows: 1, shrink_to_content: true, trim_empty_columns: true}
validation_rules: []
metadata:
  document_kind: report-design
  demo: true
"""


def main() -> None:
    WORKBOOKS.mkdir(parents=True, exist_ok=True)
    ASSETS.mkdir(parents=True, exist_ok=True)
    TEMPLATES.mkdir(parents=True, exist_ok=True)

    screen_img = ASSETS / "screen-layout.png"
    report_img = ASSETS / "report-layout.png"
    _make_mock_image(screen_img, "保証審査一覧", "SCR-A0010 / 検索・一覧モック", "screen")
    _make_mock_image(report_img, "予見・事故記録簿", "RPT-B0020 / 帳票レイアウトモック", "report")

    build_screen_design(WORKBOOKS / "SCR-A0010_画面設計書_保証一覧.xlsx", screen_img)
    build_report_design(WORKBOOKS / "RPT-B0020_帳票設計書_予見事故記録簿.xlsx", report_img)

    (TEMPLATES / "demo-screen-design.yaml").write_text(SCREEN_TEMPLATE, encoding="utf-8")
    (TEMPLATES / "demo-report-design.yaml").write_text(REPORT_TEMPLATE, encoding="utf-8")
    print(f"created: {WORKBOOKS}")
    print(f"templates: {TEMPLATES}")


if __name__ == "__main__":
    main()
