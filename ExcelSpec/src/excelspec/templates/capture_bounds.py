"""Resolve screenshot capture ranges from sparse content + drawing anchors.

Works on a region's already-materialised cells (bounded to content, never the
whole sheet) and the sheet's extracted drawing assets (OOXML anchors). Two
strategies:

* ``connected_region`` — shrink a range to the connected content/bordered block
  around its anchor (used for the 凡例 legend box), so it never captures the full
  table width or a far isolated styled cell.
* ``dynamic_bottom`` — fixed top/left/right + a bottom computed from drawings,
  content, borders and merges (used for 画面遷移図 transition diagrams).
"""

from __future__ import annotations

import re
from dataclasses import dataclass

from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)

from ..models.document_ir import AssetIR, CellIR, StyleIR

_EXCEL_MAX_ROW = 1048576
_A1 = re.compile(r"^\$?([A-Z]{1,3})\$?(\d+)")


def _display(cell: CellIR) -> str:
    if cell.display_value is not None:
        return str(cell.display_value).strip()
    if cell.raw_value is not None:
        return str(cell.raw_value).strip()
    return ""


def _has_border(style: StyleIR | None) -> bool:
    return bool(style and style.border)


def _has_fill(style: StyleIR | None) -> bool:
    if not style or not style.fill:
        return False
    fill_type = style.fill.get("type")
    return bool(fill_type) and fill_type not in (None, "none")


def _is_visual(cell: CellIR) -> bool:
    return bool(_display(cell)) or _has_border(cell.style) or _has_fill(cell.style)


def _bounds_from_a1(a1: str) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(a1)
    return (min_row, min_col, max_row, max_col)


def _a1(bounds: tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def _anchor_bottom(anchor: str | None) -> int | None:
    """Return the bottom row referenced by an A1 cell or ``A1:B2`` range anchor."""

    if not anchor:
        return None
    parts = anchor.split(":")
    rows = []
    for part in parts:
        match = _A1.match(part.strip().upper())
        if match:
            rows.append(int(match.group(2)))
    return max(rows) if rows else None


def _anchor_columns(anchor: str | None) -> tuple[int, int] | None:
    if not anchor:
        return None
    cols = []
    for part in anchor.split(":"):
        match = _A1.match(part.strip().upper())
        if match:
            cols.append(column_index_from_string(match.group(1)))
    return (min(cols), max(cols)) if cols else None


def _contiguous_run(values: list[int], anchor: int, gap: int = 1) -> tuple[int, int]:
    """Return the contiguous (gap-tolerant) run of ``values`` containing ``anchor``."""

    ordered = sorted(set(values))
    if anchor not in ordered:
        ordered = sorted(set(ordered) | {anchor})
    low = high = anchor
    index = ordered.index(anchor)
    for value in ordered[index + 1:]:
        if value - high - 1 <= gap:
            high = value
        else:
            break
    for value in reversed(ordered[:index]):
        if low - value - 1 <= gap:
            low = value
        else:
            break
    return low, high


@dataclass(slots=True)
class CaptureResolution:
    range_a1: str
    bounds_method: str
    requested_range: str
    diagnostics: list[tuple[str, str]]  # (code, message)
    metadata: dict = None  # component bottoms / band, for audit

    def __post_init__(self) -> None:
        if self.metadata is None:
            self.metadata = {}


@dataclass(slots=True)
class DynamicBand:
    top: int
    left: int
    right: int
    section_ceiling: int
    hard_cap: int


def dynamic_band(
    base: tuple[int, int, int, int], options: dict, *, section_max_row: int
) -> DynamicBand:
    """Fixed top/left/right band + section-exclusion ceiling + hard cap."""

    min_row, _min_col, _max_row, max_col = base
    top = int(options["top_row"]) if isinstance(options.get("top_row"), int) else min_row
    left_col = options.get("left_column")
    right_col = options.get("right_column")
    left = column_index_from_string(str(left_col)) if left_col else base[1]
    right = column_index_from_string(str(right_col)) if right_col else max_col
    if right < left:
        left, right = right, left
    section_ceiling = min(section_max_row, _EXCEL_MAX_ROW)
    hard_cap = min(
        int(options.get("max_bottom_row", section_ceiling) or section_ceiling),
        section_ceiling,
    )
    return DynamicBand(top=top, left=left, right=right, section_ceiling=section_ceiling, hard_cap=hard_cap)


def resolve_connected_region(
    cells: list[CellIR],
    base: tuple[int, int, int, int],
    *,
    padding_rows: int = 1,
    padding_columns: int = 1,
) -> CaptureResolution:
    """Shrink ``base`` to the connected content/bordered block around its anchor."""

    min_row, min_col, max_row, max_col = base
    in_bounds = [
        cell
        for cell in cells
        if min_row <= cell.row <= max_row
        and min_col <= cell.column <= max_col
        and _is_visual(cell)
    ]
    if not in_bounds:
        return CaptureResolution(_a1(base), "locator_range", _a1(base), [])

    content_cols = [cell.column for cell in in_bounds]
    content_rows = [cell.row for cell in in_bounds]
    left, right = _contiguous_run(content_cols, min_col)
    top = min(content_rows)
    bottom = max(content_rows)

    resolved = (
        max(1, top - padding_rows),
        max(1, left - padding_columns),
        min(_EXCEL_MAX_ROW, bottom + padding_rows),
        right + padding_columns,
    )
    return CaptureResolution(_a1(resolved), "connected_region", _a1(base), [])


def resolve_dynamic_bottom(
    cells: list[CellIR],
    assets: list[AssetIR],
    base: tuple[int, int, int, int],
    options: dict,
    *,
    section_max_row: int,
    com_shape_bottom: int | None = None,
) -> CaptureResolution:
    """Fixed top/left/right + a dynamic bottom.

    ``com_shape_bottom`` is the union bottom read from Excel COM Shapes (priority
    2). Priority order for the resolved bottom is: OOXML drawing anchor -> COM
    Shape bounds -> content -> bordered -> merged. The largest wins, so a
    Connector deeper than every node is never clipped.
    """

    band = dynamic_band(base, options, section_max_row=section_max_row)
    top, left, right = band.top, band.left, band.right
    ceiling = band.section_ceiling
    hard_cap = band.hard_cap

    # 1) OOXML drawing anchors within the band and section
    ooxml_bottom: int | None = None
    for asset in assets:
        columns = _anchor_columns(asset.anchor)
        bottom = _anchor_bottom(asset.anchor)
        if bottom is None or bottom < top or bottom > ceiling:
            continue
        if columns is not None and (columns[1] < left or columns[0] > right):
            continue
        ooxml_bottom = bottom if ooxml_bottom is None else max(ooxml_bottom, bottom)

    # 3/4) content / bordered / merged cells within the band and section
    content_bottom: int | None = None
    bordered_bottom: int | None = None
    merged_bottom: int | None = None
    for cell in cells:
        if cell.row < top or cell.column < left or cell.column > right:
            continue
        cell_bottom = cell.row + max(1, cell.row_span) - 1
        if cell_bottom > ceiling:
            continue
        if cell.row_span > 1:
            merged_bottom = cell_bottom if merged_bottom is None else max(merged_bottom, cell_bottom)
        if _display(cell):
            content_bottom = cell_bottom if content_bottom is None else max(content_bottom, cell_bottom)
        elif _has_border(cell.style):
            bordered_bottom = cell_bottom if bordered_bottom is None else max(bordered_bottom, cell_bottom)

    # 2) COM shape bottom is clamped to this section's ceiling
    com_bottom = com_shape_bottom
    if com_bottom is not None and com_bottom > ceiling:
        com_bottom = ceiling

    sources = {
        "ooxml_bottom": ooxml_bottom,
        "com_shape_bottom": com_bottom,
        "content_bottom": content_bottom,
        "bordered_bottom": bordered_bottom,
        "merged_bottom": merged_bottom,
    }
    present = {name: value for name, value in sources.items() if value is not None}

    diagnostics: list[tuple[str, str]] = []
    metadata = dict(sources)
    metadata["band"] = f"{get_column_letter(left)}{top}:{get_column_letter(right)}"

    if not present:
        fallback = min(top, hard_cap)
        diagnostics.append(
            (
                "screenshot.dynamic_bottom_not_found",
                f"未找到可靠的动态底部（top={top}, band={get_column_letter(left)}:{get_column_letter(right)}），"
                f"回退到 top 行并保留文本",
            )
        )
        resolved = (top, left, max(top, fallback), right)
        metadata["resolved_bottom"] = resolved[2]
        metadata["dominant_source"] = None
        return CaptureResolution(_a1(resolved), "fallback_top_only", _a1(base), diagnostics, metadata)

    dominant_source = max(present, key=present.get)
    dynamic_bottom = present[dominant_source]
    padding_bottom = int(options.get("padding_bottom_rows", 0) or 0)
    final_bottom = min(dynamic_bottom + padding_bottom, hard_cap)
    if final_bottom < top:
        final_bottom = top
    resolved = (top, left, final_bottom, right)
    metadata["resolved_bottom"] = final_bottom
    metadata["dominant_source"] = dominant_source
    return CaptureResolution(_a1(resolved), "dynamic_bottom", _a1(base), diagnostics, metadata)


__all__ = [
    "CaptureResolution",
    "DynamicBand",
    "dynamic_band",
    "resolve_connected_region",
    "resolve_dynamic_bottom",
]
