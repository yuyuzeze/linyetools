"""Template scoring, semantic extraction, and conservative freeform fallback."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Iterable

from openpyxl.utils.cell import (
    get_column_letter,
    range_boundaries,
)

from ..models.document_ir import (
    AssetIR,
    AssetType,
    CellIR,
    DiagnosticIR,
    DiagnosticSeverity,
    DocumentIR,
    RegionIR,
    RegionType,
    SheetIR,
    SourceRef,
    TableIR,
)
from ..models.template import (
    ExtractionSpec,
    FingerprintRule,
    LocatorMode,
    RegionLocator,
    RegionTemplate,
    SheetTemplate,
    TemplateSpec,
)


@dataclass(slots=True)
class TemplateCandidate:
    template_id: str
    version: str
    score: float
    sheet_score: float
    fingerprint_score: float
    accepted: bool


@dataclass(slots=True)
class MatchResult:
    mode: str
    template: TemplateSpec | None
    candidates: list[TemplateCandidate] = field(default_factory=list)


@dataclass(slots=True)
class ExtractionResult:
    document: DocumentIR
    match: MatchResult
    unrecognized_ranges: dict[str, list[str]] = field(default_factory=dict)


def _display(cell: CellIR) -> str:
    if cell.display_value is not None:
        return str(cell.display_value).strip()
    if cell.raw_value is not None:
        return str(cell.raw_value).strip()
    return ""


def _sheet_cells(sheet: SheetIR) -> dict[str, CellIR]:
    cells: dict[str, CellIR] = {}
    for region in sheet.regions:
        for table in region.tables:
            for cell in table.cells:
                cells.setdefault(cell.coordinate.upper(), cell)
    return cells


def _regex_matches(pattern: str, value: str) -> bool:
    try:
        return re.search(pattern, value, flags=re.IGNORECASE) is not None
    except re.error:
        return pattern.casefold() in value.casefold()


def _matching_sheets(document: DocumentIR, pattern: str | None) -> list[SheetIR]:
    if pattern is None:
        return document.sheets
    return [sheet for sheet in document.sheets if _regex_matches(pattern, sheet.name)]


def _fingerprint_score(document: DocumentIR, rule: FingerprintRule) -> float:
    best = 0.0
    for sheet in _matching_sheets(document, rule.sheet_name_pattern):
        cells = _sheet_cells(sheet)
        checks: list[bool] = []
        for coordinate, expected in rule.cells.items():
            cell = cells.get(coordinate.upper())
            checks.append(cell is not None and _regex_matches(expected, _display(cell)))
        haystack = "\n".join(_display(cell) for cell in cells.values())
        checks.extend(text.casefold() in haystack.casefold() for text in rule.required_text)
        score = sum(checks) / len(checks) if checks else 0.0
        best = max(best, score)
    return best


def score_template(document: DocumentIR, template: TemplateSpec) -> TemplateCandidate:
    """Return a normalized and explainable score for one candidate template."""

    patterns = template.match.sheet_name_patterns or [
        sheet.name_pattern for sheet in template.sheets
    ]
    sheet_score = (
        sum(bool(_matching_sheets(document, pattern)) for pattern in patterns) / len(patterns)
        if patterns
        else 0.0
    )
    weighted_fingerprints = [
        (_fingerprint_score(document, rule), rule.weight)
        for rule in template.match.fingerprints
    ]
    total_weight = sum(weight for _, weight in weighted_fingerprints)
    fingerprint_score = (
        sum(score * weight for score, weight in weighted_fingerprints) / total_weight
        if total_weight
        else 0.0
    )
    if patterns and weighted_fingerprints:
        score = (sheet_score * 0.35) + (fingerprint_score * 0.65)
    elif weighted_fingerprints:
        score = fingerprint_score
    elif patterns:
        score = sheet_score
    else:
        score = 0.0
    score = round(min(1.0, max(0.0, score)), 6)
    return TemplateCandidate(
        template_id=template.template_id,
        version=template.version,
        score=score,
        sheet_score=round(sheet_score, 6),
        fingerprint_score=round(fingerprint_score, 6),
        accepted=score >= template.match.minimum_score,
    )


def match_template(
    document: DocumentIR,
    templates: Iterable[TemplateSpec],
    *,
    minimum_score: float | None = None,
) -> MatchResult:
    """Rank candidates and select the highest candidate above its threshold."""

    template_list = list(templates)
    ranked = sorted(
        zip((score_template(document, item) for item in template_list), template_list),
        key=lambda pair: (-pair[0].score, pair[0].template_id, pair[0].version),
    )
    candidates = [candidate for candidate, _ in ranked]
    for candidate, template in ranked:
        threshold = template.match.minimum_score if minimum_score is None else minimum_score
        candidate.accepted = candidate.score >= threshold
    selected = next(
        (template for candidate, template in ranked if candidate.accepted),
        None,
    )
    return MatchResult(
        mode="template" if selected is not None else "freeform",
        template=selected,
        candidates=candidates,
    )


def _coordinate(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _cell_position(cell: CellIR) -> tuple[int, int]:
    return cell.row, cell.column


def _sheet_bounds(cells: dict[str, CellIR]) -> tuple[int, int, int, int] | None:
    if not cells:
        return None
    positions = [_cell_position(cell) for cell in cells.values()]
    return (
        min(column for _, column in positions),
        min(row for row, _ in positions),
        max(column for _, column in positions),
        max(row for row, _ in positions),
    )


def _find_anchor(cells: dict[str, CellIR], locator: RegionLocator) -> CellIR | None:
    for cell in sorted(cells.values(), key=_cell_position):
        value = _display(cell)
        if locator.anchor_text is not None and value == locator.anchor_text:
            return cell
        if locator.anchor_pattern is not None and _regex_matches(locator.anchor_pattern, value):
            return cell
    return None


def _find_all_anchors(cells: dict[str, CellIR], locator: RegionLocator) -> list[CellIR]:
    matches: list[CellIR] = []
    for cell in sorted(cells.values(), key=_cell_position):
        value = _display(cell)
        if locator.anchor_text is not None and value == locator.anchor_text:
            matches.append(cell)
            continue
        if locator.anchor_pattern is not None and _regex_matches(locator.anchor_pattern, value):
            matches.append(cell)
    return matches


def _find_end_anchor(
    cells: dict[str, CellIR],
    locator: RegionLocator,
    start_row: int,
    *,
    before_row: int | None = None,
) -> CellIR | None:
    for cell in sorted(cells.values(), key=_cell_position):
        if cell.row <= start_row:
            continue
        if before_row is not None and cell.row >= before_row:
            break
        value = _display(cell)
        if locator.end_anchor_text is not None and value == locator.end_anchor_text:
            return cell
        if locator.end_anchor_pattern is not None and _regex_matches(
            locator.end_anchor_pattern, value
        ):
            return cell
    return None


def _bounds_from_anchor(
    *,
    anchor: CellIR,
    locator: RegionLocator,
    sheet_min_col: int,
    sheet_min_row: int,
    sheet_max_col: int,
    sheet_max_row: int,
    cells: dict[str, CellIR],
    before_row: int | None = None,
) -> tuple[int, int, int, int]:
    min_row = max(sheet_min_row, anchor.row + locator.row_offset)
    min_col = max(sheet_min_col, anchor.column + locator.column_offset)
    end_anchor = _find_end_anchor(cells, locator, min_row, before_row=before_row)
    default_max_row = (
        before_row - 1
        if before_row is not None
        else max(min_row, sheet_max_row)
    )
    max_row = (
        min_row + locator.height - 1
        if locator.height is not None
        else end_anchor.row - 1
        if end_anchor is not None
        else default_max_row
    )
    max_col = (
        min_col + locator.width - 1
        if locator.width is not None
        else max(min_col, sheet_max_col)
    )
    if max_row < min_row:
        max_row = min_row
    if max_col < min_col:
        max_col = min_col
    return min_col, min_row, max_col, max_row


def locate_region(
    sheet: SheetIR, locator: RegionLocator
) -> tuple[int, int, int, int] | None:
    """Resolve a fixed or anchor locator to min_col, min_row, max_col, max_row."""

    located = locate_regions(sheet, locator)
    return located[0] if located else None


def locate_regions(
    sheet: SheetIR, locator: RegionLocator
) -> list[tuple[int, int, int, int]]:
    """Resolve one or more rectangular regions for a locator."""

    cells = _sheet_cells(sheet)
    bounds = _sheet_bounds(cells)
    if bounds is None:
        return []
    sheet_min_col, sheet_min_row, sheet_max_col, sheet_max_row = bounds
    if locator.mode == LocatorMode.FIXED:
        return [range_boundaries(locator.range or "")]

    anchors = (
        _find_all_anchors(cells, locator)
        if locator.repeat_anchor
        else ([anchor] if (anchor := _find_anchor(cells, locator)) is not None else [])
    )
    if not anchors:
        return []

    regions: list[tuple[int, int, int, int]] = []
    for index, anchor in enumerate(anchors):
        before_row = anchors[index + 1].row if index + 1 < len(anchors) else None
        regions.append(
            _bounds_from_anchor(
                anchor=anchor,
                locator=locator,
                sheet_min_col=sheet_min_col,
                sheet_min_row=sheet_min_row,
                sheet_max_col=sheet_max_col,
                sheet_max_row=sheet_max_row,
                cells=cells,
                before_row=before_row,
            )
        )
    return regions


def _cells_in_bounds(
    cells: dict[str, CellIR], bounds: tuple[int, int, int, int]
) -> list[CellIR]:
    min_col, min_row, max_col, max_row = bounds
    return sorted(
        (
            cell
            for cell in cells.values()
            if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col
        ),
        key=_cell_position,
    )


def _range_name(bounds: tuple[int, int, int, int]) -> str:
    min_col, min_row, max_col, max_row = bounds
    return f"{_coordinate(min_row, min_col)}:{_coordinate(max_row, max_col)}"


def _cell_has_text(cell: CellIR, by_coordinate: dict[str, CellIR]) -> bool:
    value = _display(cell)
    if value:
        return True
    if cell.merged_master:
        master = by_coordinate.get(cell.merged_master)
        return bool(master and _display(master))
    return False


def _trim_blank_rows(
    cells: list[CellIR],
    bounds: tuple[int, int, int, int],
    stop_after: int,
) -> tuple[list[CellIR], tuple[int, int, int, int]]:
    if stop_after <= 0:
        return cells, bounds
    min_col, min_row, max_col, max_row = bounds
    blank_run = 0
    effective_max = max_row
    by_coordinate = {cell.coordinate: cell for cell in cells}
    values = {
        (cell.row, cell.column): _cell_has_text(cell, by_coordinate) for cell in cells
    }
    for row in range(min_row, max_row + 1):
        is_blank = all(
            not values.get((row, column), False) for column in range(min_col, max_col + 1)
        )
        blank_run = blank_run + 1 if is_blank else 0
        if blank_run >= stop_after:
            effective_max = row - blank_run
            break
    trimmed_bounds = (min_col, min_row, max_col, max(min_row, effective_max))
    return _cells_in_bounds({cell.coordinate: cell for cell in cells}, trimmed_bounds), trimmed_bounds


def _trim_empty_columns(
    cells: list[CellIR],
    bounds: tuple[int, int, int, int],
) -> tuple[list[CellIR], tuple[int, int, int, int]]:
    min_col, min_row, max_col, max_row = bounds
    by_coordinate = {cell.coordinate: cell for cell in cells}
    kept_columns = [
        column
        for column in range(min_col, max_col + 1)
        if any(
            _cell_has_text(cell, by_coordinate)
            for cell in cells
            if cell.column == column
        )
    ]
    if not kept_columns:
        return cells, bounds
    trimmed_bounds = (kept_columns[0], min_row, kept_columns[-1], max_row)
    # Drop wholly empty columns even if they sit between content columns.
    kept = {
        cell.coordinate: cell
        for cell in cells
        if cell.column in set(kept_columns)
    }
    return _cells_in_bounds(kept, trimmed_bounds), trimmed_bounds


def _shrink_to_content(
    cells: list[CellIR],
    bounds: tuple[int, int, int, int],
) -> tuple[list[CellIR], tuple[int, int, int, int]]:
    by_coordinate = {cell.coordinate: cell for cell in cells}
    content_cells = [cell for cell in cells if _cell_has_text(cell, by_coordinate)]
    if not content_cells:
        return cells, bounds
    min_col = min(cell.column for cell in content_cells)
    max_col = max(cell.column + max(cell.col_span, 1) - 1 for cell in content_cells)
    min_row = min(cell.row for cell in content_cells)
    max_row = max(cell.row + max(cell.row_span, 1) - 1 for cell in content_cells)
    shrunk = (min_col, min_row, max_col, max_row)
    return _cells_in_bounds({cell.coordinate: cell for cell in cells}, shrunk), shrunk


def _option_flag(options: dict, name: str, default: bool = False) -> bool:
    value = options.get(name, default)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return bool(value)


def _header_labels(
    cells: list[CellIR], bounds: tuple[int, int, int, int], header_rows: int
) -> dict[int, str]:
    min_col, min_row, max_col, _ = bounds
    by_coordinate = {cell.coordinate: cell for cell in cells}
    values: dict[tuple[int, int], str] = {}
    for cell in cells:
        value = _display(cell)
        if not value and cell.merged_master:
            master = by_coordinate.get(cell.merged_master)
            value = _display(master) if master is not None else ""
        values[(cell.row, cell.column)] = value
    labels: dict[int, str] = {}
    for column in range(min_col, max_col + 1):
        parts = [
            values.get((row, column), "")
            for row in range(min_row, min_row + header_rows)
            if values.get((row, column), "")
        ]
        labels[column] = " / ".join(dict.fromkeys(parts))
    return labels


def _column_semantics(
    extractor: ExtractionSpec,
    labels: dict[int, str],
    min_col: int,
) -> dict[str, str]:
    result: dict[str, str] = {}
    for column, label in labels.items():
        relative = column - min_col + 1
        semantic = (
            extractor.column_semantics.get(str(relative))
            or extractor.column_semantics.get(get_column_letter(column))
            or extractor.column_semantics.get(label)
        )
        if semantic is None:
            for pattern, candidate in extractor.column_semantics.items():
                if _regex_matches(pattern, label):
                    semantic = candidate
                    break
        if semantic is not None:
            result[get_column_letter(column)] = semantic
    return result


def _key_semantic(extractor: ExtractionSpec, key: str) -> str:
    if key in extractor.key_semantics:
        return extractor.key_semantics[key]
    for pattern, semantic in extractor.key_semantics.items():
        if _regex_matches(pattern, key):
            return semantic
    return key


def _extract_region(
    sheet: SheetIR,
    template: RegionTemplate,
    bounds: tuple[int, int, int, int],
    *,
    region_id: str | None = None,
) -> RegionIR:
    cells = _cells_in_bounds(_sheet_cells(sheet), bounds)
    extractor = template.extractor or ExtractionSpec(kind=template.region_type)
    stop_after = int(extractor.options.get("stop_after_blank_rows", 0))
    cells, bounds = _trim_blank_rows(cells, bounds, stop_after)
    shrink = _option_flag(extractor.options, "shrink_to_content")
    if shrink:
        cells, bounds = _shrink_to_content(cells, bounds)
    trim_columns = _option_flag(
        extractor.options,
        "trim_empty_columns",
        default=extractor.kind == "table",
    )
    if trim_columns:
        cells, bounds = _trim_empty_columns(cells, bounds)
    source = SourceRef(sheet=sheet.name, range=_range_name(bounds))
    region_type = RegionType(template.region_type)
    confidence = 1.0 if template.locator.mode == LocatorMode.FIXED else 0.9
    resolved_id = region_id or template.region_id
    region = RegionIR(
        region_id=resolved_id,
        region_type=region_type,
        title=template.title,
        source=source,
        confidence=confidence,
        metadata={
            "extractor_kind": extractor.kind,
            "template_region_id": template.region_id,
            "trim_empty_columns": trim_columns,
            "shrink_to_content": shrink,
        },
    )
    min_col, min_row, _, _ = bounds
    if extractor.kind == "key_value":
        key_column = min_col + (extractor.key_column or 1) - 1
        value_column = min_col + (extractor.value_column or 2) - 1
        by_position = {(cell.row, cell.column): cell for cell in cells}
        for row in range(min_row + extractor.header_rows, bounds[3] + 1):
            key_cell = by_position.get((row, key_column))
            value_cell = by_position.get((row, value_column))
            key = _display(key_cell) if key_cell else ""
            if not key:
                continue
            semantic = _key_semantic(extractor, key)
            if value_cell is None:
                region.values[semantic] = None
            elif value_cell.formula is not None:
                # For formula cells, use Excel's cached calculated result.
                # Falling back to the formula text would expose implementation
                # details instead of the document value.
                region.values[semantic] = value_cell.display_value
            else:
                region.values[semantic] = (
                    value_cell.raw_value
                    if value_cell.raw_value is not None
                    else _display(value_cell)
                )
        region.metadata["key_labels"] = extractor.key_semantics
    elif extractor.kind == "table":
        labels = _header_labels(cells, bounds, extractor.header_rows)
        semantics = _column_semantics(extractor, labels, min_col)
        region.tables.append(
            TableIR(
                table_id=resolved_id,
                cells=cells,
                source=source,
                header_rows=extractor.header_rows,
                column_semantics=semantics,
                metadata={
                    "header_labels": {
                        get_column_letter(column): label for column, label in labels.items()
                    }
                },
            )
        )
    else:
        region.tables.append(
            TableIR(table_id=resolved_id, cells=cells, source=source)
        )
    return region


def _nonblank_bands(
    cells: dict[str, CellIR],
    excluded: set[str],
) -> list[tuple[int, int, int, int]]:
    available = [cell for cell in cells.values() if cell.coordinate not in excluded]
    if not available:
        return []
    min_row = min(cell.row for cell in available)
    max_row = max(cell.row for cell in available)
    nonblank_rows = {cell.row for cell in available if _display(cell)}
    heading_rows = {
        cell.row
        for cell in available
        if _display(cell)
        and cell.style is not None
        and (
            bool(cell.style.font.get("bold"))
            or bool(cell.style.fill.get("fill_type"))
            or bool(cell.style.fill.get("patternType"))
        )
    }
    row_bands: list[tuple[int, int]] = []
    start: int | None = None
    for row in range(min_row, max_row + 2):
        if row in nonblank_rows and start is None:
            start = row
        elif row in heading_rows and start is not None and row > start:
            row_bands.append((start, row - 1))
            start = row
        elif row not in nonblank_rows and start is not None:
            row_bands.append((start, row - 1))
            start = None
    result: list[tuple[int, int, int, int]] = []
    for row_start, row_end in row_bands:
        band = [cell for cell in available if row_start <= cell.row <= row_end]
        columns = sorted({cell.column for cell in band if _display(cell)})
        if not columns:
            continue
        col_start = previous = columns[0]
        for column in columns[1:] + [columns[-1] + 2]:
            if column > previous + 1:
                result.append((col_start, row_start, previous, row_end))
                col_start = column
            previous = column
    return result


def _freeform_regions(
    sheet: SheetIR,
    *,
    excluded: set[str] | None = None,
    prefix: str = "freeform",
) -> tuple[list[RegionIR], list[str]]:
    cells = _sheet_cells(sheet)
    excluded = excluded or set()
    regions: list[RegionIR] = []
    ranges: list[str] = []
    assigned: set[str] = set()
    for index, bounds in enumerate(_nonblank_bands(cells, excluded), start=1):
        selected = _cells_in_bounds(cells, bounds)
        selected = [cell for cell in selected if cell.coordinate not in excluded]
        if not selected:
            continue
        assigned.update(cell.coordinate for cell in selected)
        range_name = _range_name(bounds)
        ranges.append(range_name)
        source = SourceRef(sheet=sheet.name, range=range_name)
        regions.append(
            RegionIR(
                region_id=f"{prefix}-{index}",
                region_type=RegionType.FREEFORM,
                source=source,
                tables=[
                    TableIR(
                        table_id=f"{prefix}-{index}",
                        cells=selected,
                        source=source,
                    )
                ],
                confidence=0.25,
                metadata={"segmentation": "blank-and-style-boundaries"},
            )
        )
    residual = [
        cell
        for cell in sorted(cells.values(), key=_cell_position)
        if cell.coordinate not in excluded and cell.coordinate not in assigned
    ]
    if residual:
        bounds = (
            min(cell.column for cell in residual),
            min(cell.row for cell in residual),
            max(cell.column for cell in residual),
            max(cell.row for cell in residual),
        )
        range_name = _range_name(bounds)
        ranges.append(range_name)
        source = SourceRef(sheet=sheet.name, range=range_name)
        regions.append(
            RegionIR(
                region_id=f"{prefix}-grid-residual",
                region_type=RegionType.FREEFORM,
                source=source,
                tables=[
                    TableIR(
                        table_id=f"{prefix}-grid-residual",
                        cells=residual,
                        source=source,
                    )
                ],
                confidence=0.1,
                metadata={
                    "segmentation": "grid-residual",
                    "reason": "preserve-empty-or-separated-effective-cells",
                },
            )
        )
    return regions, ranges


def _copy_assets(sheet: SheetIR, regions: list[RegionIR]) -> list[AssetIR]:
    assets = list(sheet.assets)
    for asset in assets:
        reference = asset.anchor
        if reference is None and asset.source is not None:
            reference = asset.source.range or asset.source.cell
        if reference is None:
            continue
        try:
            asset_bounds = (
                range_boundaries(reference)
                if ":" in reference
                else range_boundaries(f"{reference}:{reference}")
            )
        except ValueError:
            continue
        for region in regions:
            if region.source is None or region.source.range is None:
                continue
            region_bounds = range_boundaries(region.source.range)
            if not (
                asset_bounds[2] < region_bounds[0]
                or asset_bounds[0] > region_bounds[2]
                or asset_bounds[3] < region_bounds[1]
                or asset_bounds[1] > region_bounds[3]
            ):
                if asset.asset_id not in region.asset_ids:
                    region.asset_ids.append(asset.asset_id)
                break
    return assets


def extract_with_template(document: DocumentIR, match: MatchResult) -> ExtractionResult:
    """Apply a selected template or preserve the workbook in freeform mode."""

    template = match.template
    output_sheets: list[SheetIR] = []
    unrecognized: dict[str, list[str]] = {}
    for sheet in document.sheets:
        template_sheet: SheetTemplate | None = None
        if template is not None:
            template_sheet = next(
                (item for item in template.sheets if _regex_matches(item.name_pattern, sheet.name)),
                None,
            )
        if template_sheet is None:
            regions, ranges = _freeform_regions(sheet)
            unrecognized[sheet.name] = ranges
            output_sheets.append(
                SheetIR(
                    sheet_id=sheet.sheet_id,
                    name=sheet.name,
                    index=sheet.index,
                    regions=regions,
                    assets=_copy_assets(sheet, regions),
                    diagnostics=list(sheet.diagnostics),
                    metadata={**sheet.metadata, "extraction_mode": "freeform"},
                )
            )
            continue

        cells = _sheet_cells(sheet)
        covered: set[str] = set()
        diagnostics = list(sheet.diagnostics)
        regions: list[RegionIR] = []
        for region_template in sorted(
            template_sheet.regions, key=lambda item: (item.order, item.region_id)
        ):
            located = locate_regions(sheet, region_template.locator)
            if not located:
                severity = (
                    DiagnosticSeverity.ERROR
                    if region_template.required
                    else DiagnosticSeverity.WARNING
                )
                diagnostics.append(
                    DiagnosticIR(
                        code="template.region_not_found",
                        severity=severity,
                        message=f"模板区域未找到: {region_template.region_id}",
                        source=SourceRef(sheet=sheet.name),
                        region_id=region_template.region_id,
                    )
                )
                continue
            for index, bounds in enumerate(located, start=1):
                region_id = (
                    region_template.region_id
                    if index == 1
                    else f"{region_template.region_id}-{index}"
                )
                region = _extract_region(
                    sheet, region_template, bounds, region_id=region_id
                )
                if len(located) > 1:
                    region.metadata["repeat_index"] = index
                    region.metadata["repeat_anchor"] = True
                    if region.title and index > 1:
                        region.title = f"{region.title} ({index})"
                for binding in region_template.screenshot_bindings:
                    region.asset_ids.append(binding.asset_id)
                regions.append(region)
                if region.source and region.source.range:
                    try:
                        final_bounds = range_boundaries(region.source.range)
                    except ValueError:
                        final_bounds = bounds
                else:
                    final_bounds = bounds
                covered.update(
                    cell.coordinate for cell in _cells_in_bounds(cells, final_bounds)
                )
        fallback_regions, ranges = _freeform_regions(
            sheet, excluded=covered, prefix="unrecognized"
        )
        regions.extend(fallback_regions)
        unrecognized[sheet.name] = ranges
        assets = _copy_assets(sheet, regions)
        known_assets = {asset.asset_id for asset in assets}
        for region_template in template_sheet.regions:
            for binding in region_template.screenshot_bindings:
                if binding.asset_id not in known_assets:
                    assets.append(
                        AssetIR(
                            asset_id=binding.asset_id,
                            asset_type=AssetType(binding.asset_type),
                            uri=binding.path,
                            description=binding.description,
                            source=SourceRef(sheet=sheet.name),
                            extraction_status="bound",
                            metadata={"template_binding": True},
                        )
                    )
                    known_assets.add(binding.asset_id)
        output_sheets.append(
            SheetIR(
                sheet_id=sheet.sheet_id,
                name=sheet.name,
                index=sheet.index,
                regions=regions,
                assets=assets,
                diagnostics=diagnostics,
                metadata={**sheet.metadata, "extraction_mode": "template"},
            )
        )

    output = DocumentIR(
        document_id=document.document_id,
        title=document.title,
        sheets=output_sheets,
        source_path=document.source_path,
        template_id=template.template_id if template else None,
        template_version=template.version if template else None,
        assets=list(document.assets),
        diagnostics=list(document.diagnostics),
        metadata={
            **document.metadata,
            "template_match": {
                "mode": match.mode,
                "candidates": [
                    {
                        "template_id": candidate.template_id,
                        "version": candidate.version,
                        "score": candidate.score,
                        "accepted": candidate.accepted,
                    }
                    for candidate in match.candidates
                ],
                "unrecognized_ranges": unrecognized,
            },
        },
    )
    return ExtractionResult(output, match, unrecognized)


def apply_best_template(
    document: DocumentIR,
    templates: Iterable[TemplateSpec],
    *,
    minimum_score: float | None = None,
) -> ExtractionResult:
    match = match_template(document, templates, minimum_score=minimum_score)
    return extract_with_template(document, match)


__all__ = [
    "ExtractionResult",
    "MatchResult",
    "TemplateCandidate",
    "apply_best_template",
    "extract_with_template",
    "locate_region",
    "locate_regions",
    "match_template",
    "score_template",
]
