"""Adapt a :class:`SparseWorkbookIR` into the canonical :class:`DocumentIR`.

The DocumentIR ``raw-grid`` region is materialised **only over each sheet's
content bounds** (value cells + merge extents). Distant style-only cells and an
inflated ``<dimension>`` never enter the content bounds, so the materialised
grid stays proportional to real content — a lone styled cell at ``XFD1048576``
cannot create a million-row rectangle. This reproduces the legacy dense grid
byte-for-byte on well-formed sheets (where effective bounds == content bounds)
while remaining bounded on pathological inputs.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl.utils import get_column_letter

from ..models.document_ir import (
    CellIR,
    DiagnosticIR,
    DiagnosticSeverity,
    DocumentIR,
    RegionIR,
    RegionType,
    SheetIR,
    TableIR,
)
from .sparse_model import SparseSheet, SparseWorkbookIR
from .workbook import _source


def materialize_region(
    sheet: SparseSheet,
    bounds: tuple[int, int, int, int],
    styles: dict,
    workbook_path: Path,
) -> tuple[list[CellIR], list[DiagnosticIR]]:
    """Materialise the dense grid of ``bounds`` from a sparse sheet.

    Absent cells become empty ``CellIR`` (``data_type='n'``, no value, no
    style), exactly like the legacy densification path — so a missing cell is an
    empty value in place, never a leftward column shift.
    """

    min_row, min_col, max_row, max_col = bounds
    cells: list[CellIR] = []
    diagnostics: list[DiagnosticIR] = []
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            coordinate = f"{get_column_letter(column)}{row}"
            master = sheet.merge_members.get((row, column))
            row_span, col_span = sheet.merge_spans.get((row, column), (1, 1))
            sparse_cell = sheet.cells.get((row, column))
            if sparse_cell is not None:
                raw_value = sparse_cell.raw_value
                display_value = sparse_cell.display_value
                data_type = sparse_cell.data_type
                formula = sparse_cell.formula
                style = (
                    styles.get(sparse_cell.style_id)
                    if sparse_cell.style_id is not None
                    else None
                )
                if formula is not None and sparse_cell.cached_value is None:
                    diagnostics.append(
                        DiagnosticIR(
                            code="FORMULA_CACHE_MISSING",
                            severity=DiagnosticSeverity.WARNING,
                            message=f"公式 {coordinate} 没有可用的缓存计算值",
                            source=_source(workbook_path, sheet.name, cell=coordinate),
                            details={"formula": formula},
                        )
                    )
            else:
                style_id = sheet.style_only.get((row, column))
                raw_value = None
                display_value = None
                data_type = "n"
                formula = None
                style = styles.get(style_id) if style_id is not None else None
            cells.append(
                CellIR(
                    coordinate=coordinate,
                    row=row,
                    column=column,
                    raw_value=raw_value,
                    display_value=display_value,
                    data_type=data_type,
                    formula=formula,
                    style=style,
                    merged_master=(
                        f"{get_column_letter(master[1])}{master[0]}" if master else None
                    ),
                    row_span=row_span,
                    col_span=col_span,
                    source=_source(workbook_path, sheet.name, cell=coordinate),
                )
            )
    return cells, diagnostics


def _sheet_to_ir(
    sheet: SparseSheet, styles: dict, workbook_path: Path
) -> SheetIR:
    bounds = sheet.content_bounds or (1, 1, 1, 1)
    min_row, min_col, max_row, max_col = bounds
    cells, diagnostics = materialize_region(sheet, bounds, styles, workbook_path)
    range_ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    table = TableIR(
        table_id="raw-grid",
        cells=cells,
        source=_source(workbook_path, sheet.name, range=range_ref),
        metadata={
            "min_row": min_row,
            "min_column": min_col,
            "max_row": max_row,
            "max_column": max_col,
            "merge_count": len(sheet.merges),
        },
    )
    region = RegionIR(
        region_id="raw-grid",
        region_type=RegionType.FREEFORM,
        source=_source(workbook_path, sheet.name, range=range_ref),
        tables=[table],
        confidence=1.0,
        metadata={"ingest_stage": "xlsx", "semantic_extraction": False},
    )
    return SheetIR(
        sheet_id=sheet.sheet_id,
        name=sheet.name,
        index=sheet.index,
        regions=[region],
        diagnostics=diagnostics,
        metadata={
            "state": sheet.state,
            "effective_bounds": list(bounds),
            "merged_ranges": list(sheet.merges),
        },
    )


def sparse_to_document(
    sparse: SparseWorkbookIR, *, asset_dir: Path
) -> tuple[DocumentIR, list[SheetIR]]:
    """Build a DocumentIR (without drawings/manifest) from a SparseWorkbookIR.

    Returns the document and its sheet list so the caller can still attach
    drawings and manifest assets with the shared helpers before finalising.
    """

    workbook_path = Path(sparse.path)
    sheets = [
        _sheet_to_ir(sheet, sparse.styles, workbook_path) for sheet in sparse.sheets
    ]
    document = DocumentIR(
        document_id=workbook_path.stem,
        title=sparse.properties.get("title") or workbook_path.stem,
        source_path=str(workbook_path),
        sheets=sheets,
        diagnostics=[],
        metadata={
            "ingestor": "sparse-ooxml",
            "legacy_fallback": False,
            "fallback_reason": None,
            "creator": sparse.properties.get("creator"),
            "last_modified_by": sparse.properties.get("last_modified_by"),
            "asset_directory": str(asset_dir),
            "sparse_stats": sparse.metadata.get("sparse_stats", {}),
        },
    )
    return document, sheets


__all__ = ["materialize_region", "sparse_to_document"]
