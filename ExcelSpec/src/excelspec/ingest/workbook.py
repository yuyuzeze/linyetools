"""Merge-aware XLSX ingestion into the canonical DocumentIR."""

from __future__ import annotations

import mimetypes
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models.document_ir import (
    AssetIR,
    AssetType,
    CellIR,
    DiagnosticIR,
    DiagnosticSeverity,
    DocumentIR,
    RegionIR,
    RegionType,
    SheetIR,
    SourceRef,
    StyleIR,
    TableIR,
)
from .manifest import ScreenshotManifest, load_screenshot_manifest
from .ooxml import extract_sheet_drawings, workbook_sheet_parts


@dataclass(slots=True)
class XlsxIngestOptions:
    asset_dir: Path | None = None
    screenshot_manifest: Path | None = None
    include_images: bool = True
    include_shapes: bool = True


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    return str(value)


def _color(color: Any) -> dict[str, Any]:
    if color is None:
        return {}
    result: dict[str, Any] = {}
    color_type = getattr(color, "type", None)
    if isinstance(color_type, str):
        result["type"] = color_type
        value = getattr(color, color_type, None)
        if isinstance(value, (str, int, float, bool)):
            result["value"] = value
    tint = getattr(color, "tint", None)
    if isinstance(tint, (int, float)) and tint:
        result["tint"] = tint
    return result


def _side(side: Any) -> dict[str, Any]:
    if side is None:
        return {}
    result: dict[str, Any] = {}
    if side.style:
        result["style"] = side.style
    color = _color(side.color)
    if color:
        result["color"] = color
    return result


def _style(cell: Cell) -> StyleIR | None:
    if not cell.has_style:
        return None
    font = {
        key: value
        for key, value in {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.b,
            "italic": cell.font.i,
            "underline": cell.font.u,
            "strike": cell.font.strike,
            "color": _color(cell.font.color),
        }.items()
        if value not in (None, {}, False)
    }
    fill = {
        key: value
        for key, value in {
            "type": cell.fill.fill_type,
            "foreground": _color(cell.fill.fgColor),
            "background": _color(cell.fill.bgColor),
        }.items()
        if value not in (None, {})
    }
    border = {
        key: value
        for key, value in {
            "left": _side(cell.border.left),
            "right": _side(cell.border.right),
            "top": _side(cell.border.top),
            "bottom": _side(cell.border.bottom),
            "diagonal": _side(cell.border.diagonal),
        }.items()
        if value
    }
    alignment = {
        key: value
        for key, value in {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "shrink_to_fit": cell.alignment.shrink_to_fit,
            "text_rotation": cell.alignment.text_rotation,
            "indent": cell.alignment.indent,
        }.items()
        if value not in (None, False, 0)
    }
    return StyleIR(
        number_format=cell.number_format,
        font=font,
        fill=fill,
        border=border,
        alignment=alignment,
    )


def _formula_text(value: Any, data_type: str) -> str | None:
    if data_type != "f":
        return None
    if isinstance(value, str):
        return value if value.startswith("=") else f"={value}"
    text = getattr(value, "text", None)
    if isinstance(text, str):
        return text if text.startswith("=") else f"={text}"
    return str(value)


def _display_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def _merge_map(
    worksheet: Worksheet,
) -> tuple[dict[tuple[int, int], tuple[int, int]], dict[tuple[int, int], tuple[int, int]]]:
    members: dict[tuple[int, int], tuple[int, int]] = {}
    spans: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        master = (merged_range.min_row, merged_range.min_col)
        spans[master] = (
            merged_range.max_row - merged_range.min_row + 1,
            merged_range.max_col - merged_range.min_col + 1,
        )
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                members[(row, column)] = master
    return members, spans


def _effective_bounds(worksheet: Worksheet) -> tuple[int, int, int, int]:
    coordinates = [
        (cell.row, cell.column)
        for cell in worksheet._cells.values()
        if cell.value is not None or cell.has_style
    ]
    for merged_range in worksheet.merged_cells.ranges:
        coordinates.extend(
            [
                (merged_range.min_row, merged_range.min_col),
                (merged_range.max_row, merged_range.max_col),
            ]
        )
    if not coordinates:
        return (1, 1, 1, 1)
    rows, columns = zip(*coordinates)
    return min(rows), min(columns), max(rows), max(columns)


def _source(workbook: Path, sheet: str, *, cell: str | None = None, range: str | None = None) -> SourceRef:
    return SourceRef(
        sheet=sheet,
        cell=cell,
        range=range,
        workbook_path=str(workbook.resolve()),
    )


def _sheet_grid(
    formula_sheet: Worksheet,
    value_sheet: Worksheet,
    workbook_path: Path,
) -> tuple[RegionIR, list[DiagnosticIR], tuple[int, int, int, int]]:
    min_row, min_column, max_row, max_column = _effective_bounds(formula_sheet)
    members, spans = _merge_map(formula_sheet)
    cells: list[CellIR] = []
    diagnostics: list[DiagnosticIR] = []
    for row in range(min_row, max_row + 1):
        for column in range(min_column, max_column + 1):
            formula_cell = formula_sheet.cell(row, column)
            cached_cell = value_sheet.cell(row, column)
            coordinate = f"{get_column_letter(column)}{row}"
            formula = _formula_text(formula_cell.value, formula_cell.data_type)
            cached_value = cached_cell.value if formula is not None else formula_cell.value
            master = members.get((row, column))
            row_span, col_span = spans.get((row, column), (1, 1))
            if formula is not None and cached_value is None:
                diagnostics.append(
                    DiagnosticIR(
                        code="FORMULA_CACHE_MISSING",
                        severity=DiagnosticSeverity.WARNING,
                        message=f"公式 {coordinate} 没有可用的缓存计算值",
                        source=_source(workbook_path, formula_sheet.title, cell=coordinate),
                        details={"formula": formula},
                    )
                )
            cells.append(
                CellIR(
                    coordinate=coordinate,
                    row=row,
                    column=column,
                    raw_value=_json_value(formula_cell.value),
                    display_value=_display_value(cached_value),
                    data_type=formula_cell.data_type,
                    formula=formula,
                    style=_style(formula_cell),
                    merged_master=(
                        f"{get_column_letter(master[1])}{master[0]}" if master else None
                    ),
                    row_span=row_span,
                    col_span=col_span,
                    source=_source(workbook_path, formula_sheet.title, cell=coordinate),
                )
            )
    range_ref = (
        f"{get_column_letter(min_column)}{min_row}:"
        f"{get_column_letter(max_column)}{max_row}"
    )
    table = TableIR(
        table_id="raw-grid",
        cells=cells,
        source=_source(workbook_path, formula_sheet.title, range=range_ref),
        metadata={
            "min_row": min_row,
            "min_column": min_column,
            "max_row": max_row,
            "max_column": max_column,
            "merge_count": len(formula_sheet.merged_cells.ranges),
        },
    )
    region = RegionIR(
        region_id="raw-grid",
        region_type=RegionType.FREEFORM,
        source=_source(workbook_path, formula_sheet.title, range=range_ref),
        tables=[table],
        confidence=1.0,
        metadata={"ingest_stage": "xlsx", "semantic_extraction": False},
    )
    return region, diagnostics, (min_row, min_column, max_row, max_column)


def _bind_manifest(
    manifest: ScreenshotManifest,
    sheets: list[SheetIR],
    workbook_path: Path,
) -> list[DiagnosticIR]:
    document_diagnostics: list[DiagnosticIR] = []
    by_name = {sheet.name: sheet for sheet in sheets}
    valid_types = {member.value: member for member in AssetType}
    for item in manifest.assets:
        sheet = by_name.get(item.sheet)
        if sheet is None:
            document_diagnostics.append(
                DiagnosticIR(
                    code="SCREENSHOT_SHEET_NOT_FOUND",
                    severity=DiagnosticSeverity.ERROR,
                    message=f"截图资产 {item.asset_id} 指定的工作表不存在: {item.sheet}",
                    details={"asset_id": item.asset_id, "manifest": str(manifest.source_path)},
                )
            )
            continue
        asset_type = valid_types.get(item.asset_type)
        if asset_type is None:
            sheet.diagnostics.append(
                DiagnosticIR(
                    code="SCREENSHOT_ASSET_TYPE_INVALID",
                    severity=DiagnosticSeverity.ERROR,
                    message=f"截图资产类型不受支持: {item.asset_type}",
                    source=_source(workbook_path, sheet.name),
                    details={"asset_id": item.asset_id},
                )
            )
            continue
        exists = item.path.is_file()
        asset = AssetIR(
            asset_id=item.asset_id,
            asset_type=asset_type,
            uri=str(item.path),
            media_type=mimetypes.guess_type(item.path.name)[0],
            description=item.description,
            source=_source(workbook_path, sheet.name),
            anchor=item.anchor,
            extraction_status="bound" if exists else "missing",
            metadata={
                **item.metadata,
                "source_kind": "user_manifest",
                "manifest_path": str(manifest.source_path),
                "ocr": item.ocr,
                "vlm": item.vlm,
                "region_id": item.region_id,
            },
        )
        sheet.assets.append(asset)
        if not exists:
            sheet.diagnostics.append(
                DiagnosticIR(
                    code="SCREENSHOT_FILE_MISSING",
                    severity=DiagnosticSeverity.ERROR,
                    message=f"截图文件不存在: {item.path}",
                    source=_source(workbook_path, sheet.name),
                    region_id=item.region_id,
                    details={"asset_id": item.asset_id, "path": str(item.path)},
                )
            )
        if item.region_id:
            region = next(
                (candidate for candidate in sheet.regions if candidate.region_id == item.region_id),
                None,
            )
            if region is None:
                sheet.diagnostics.append(
                    DiagnosticIR(
                        code="SCREENSHOT_REGION_NOT_FOUND",
                        severity=DiagnosticSeverity.WARNING,
                        message=f"截图绑定区域不存在: {item.region_id}",
                        source=_source(workbook_path, sheet.name),
                        region_id=item.region_id,
                        details={"asset_id": item.asset_id},
                    )
                )
            else:
                region.asset_ids.append(item.asset_id)
    return document_diagnostics


def attach_drawings(
    sheets: list[SheetIR],
    workbook_path: Path,
    asset_dir: Path,
    *,
    include_images: bool,
    include_shapes: bool,
) -> list[DiagnosticIR]:
    """Extract images/shapes for every sheet from the OOXML package.

    Shared by the legacy and sparse ingestors so asset extraction stays
    byte-identical across both paths. Returns document-level diagnostics; a
    single malformed drawing yields a per-sheet warning, while a broken package
    yields one ``OOXML_PACKAGE_INVALID`` document diagnostic.
    """

    document_diagnostics: list[DiagnosticIR] = []
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            sheet_parts = workbook_sheet_parts(archive)
            for sheet in sheets:
                part = sheet_parts.get(sheet.name)
                if not part:
                    sheet.diagnostics.append(
                        DiagnosticIR(
                            code="OOXML_WORKSHEET_PART_MISSING",
                            severity=DiagnosticSeverity.ERROR,
                            message="无法定位工作表对应的 OOXML 部件",
                            source=_source(workbook_path, sheet.name),
                        )
                    )
                    continue
                drawing_assets, drawing_diagnostics = extract_sheet_drawings(
                    archive,
                    sheet_part=part,
                    sheet_name=sheet.name,
                    output_dir=asset_dir / f"sheet-{sheet.index + 1}",
                    include_images=include_images,
                    include_shapes=include_shapes,
                )
                asset_numbers: dict[str, int] = {}
                for drawing in drawing_assets:
                    asset_numbers[drawing.kind] = asset_numbers.get(drawing.kind, 0) + 1
                    sheet.assets.append(
                        AssetIR(
                            asset_id=(
                                f"{sheet.sheet_id}-{drawing.kind}-"
                                f"{asset_numbers[drawing.kind]}"
                            ),
                            asset_type=AssetType(drawing.kind),
                            uri=drawing.uri,
                            media_type=drawing.media_type,
                            description=drawing.description,
                            source=_source(workbook_path, sheet.name),
                            anchor=drawing.anchor,
                            extraction_status="extracted",
                            metadata=drawing.metadata,
                        )
                    )
                for diagnostic in drawing_diagnostics:
                    sheet.diagnostics.append(
                        DiagnosticIR(
                            code=diagnostic.code,
                            severity=DiagnosticSeverity.WARNING,
                            message=diagnostic.message,
                            source=_source(workbook_path, sheet.name),
                            details=diagnostic.details,
                        )
                    )
    except (zipfile.BadZipFile, ET.ParseError, KeyError, ValueError) as error:
        document_diagnostics.append(
            DiagnosticIR(
                code="OOXML_PACKAGE_INVALID",
                severity=DiagnosticSeverity.ERROR,
                message="无法解析 XLSX 的 OOXML 资产关系",
                details={"error": str(error)},
            )
        )
    return document_diagnostics


def bind_manifest_assets(
    manifest_path: Path,
    sheets: list[SheetIR],
    workbook_path: Path,
) -> list[DiagnosticIR]:
    """Load and bind a screenshot manifest (shared by both ingestors)."""

    document_diagnostics: list[DiagnosticIR] = []
    try:
        manifest = load_screenshot_manifest(manifest_path)
        document_diagnostics.extend(_bind_manifest(manifest, sheets, workbook_path))
    except (OSError, ValueError) as error:
        document_diagnostics.append(
            DiagnosticIR(
                code="SCREENSHOT_MANIFEST_INVALID",
                severity=DiagnosticSeverity.ERROR,
                message="截图清单无法读取",
                details={"manifest": str(manifest_path), "error": str(error)},
            )
        )
    return document_diagnostics


class XlsxIngestor:
    def __init__(self, options: XlsxIngestOptions | None = None) -> None:
        self.options = options or XlsxIngestOptions()

    def ingest(self, workbook: Path) -> DocumentIR:
        workbook_path = Path(workbook).resolve()
        if not workbook_path.is_file():
            raise FileNotFoundError(workbook_path)
        formula_book = load_workbook(workbook_path, data_only=False, read_only=False)
        value_book = load_workbook(workbook_path, data_only=True, read_only=False)
        asset_dir = (
            self.options.asset_dir.resolve()
            if self.options.asset_dir
            else workbook_path.parent / f"{workbook_path.stem}_assets"
        )
        sheets: list[SheetIR] = []
        document_diagnostics: list[DiagnosticIR] = []
        try:
            for index, formula_sheet in enumerate(formula_book.worksheets):
                value_sheet = value_book[formula_sheet.title]
                region, diagnostics, bounds = _sheet_grid(
                    formula_sheet, value_sheet, workbook_path
                )
                sheets.append(
                    SheetIR(
                        sheet_id=f"sheet-{index + 1}",
                        name=formula_sheet.title,
                        index=index,
                        regions=[region],
                        diagnostics=diagnostics,
                        metadata={
                            "state": formula_sheet.sheet_state,
                            "effective_bounds": list(bounds),
                            "merged_ranges": [
                                str(item) for item in formula_sheet.merged_cells.ranges
                            ],
                        },
                    )
                )

            if self.options.include_images or self.options.include_shapes:
                document_diagnostics.extend(
                    attach_drawings(
                        sheets,
                        workbook_path,
                        asset_dir,
                        include_images=self.options.include_images,
                        include_shapes=self.options.include_shapes,
                    )
                )

            if self.options.screenshot_manifest:
                document_diagnostics.extend(
                    bind_manifest_assets(
                        self.options.screenshot_manifest, sheets, workbook_path
                    )
                )

            properties = formula_book.properties
            return DocumentIR(
                document_id=workbook_path.stem,
                title=properties.title or workbook_path.stem,
                source_path=str(workbook_path),
                sheets=sheets,
                diagnostics=document_diagnostics,
                metadata={
                    "ingestor": "openpyxl+ooxml",
                    "creator": properties.creator,
                    "last_modified_by": properties.lastModifiedBy,
                    "asset_directory": str(asset_dir),
                },
            )
        finally:
            formula_book.close()
            value_book.close()


def ingest_xlsx(
    workbook: str | Path,
    *,
    asset_dir: str | Path | None = None,
    screenshot_manifest: str | Path | None = None,
    include_images: bool = True,
    include_shapes: bool = True,
) -> DocumentIR:
    options = XlsxIngestOptions(
        asset_dir=Path(asset_dir) if asset_dir else None,
        screenshot_manifest=Path(screenshot_manifest) if screenshot_manifest else None,
        include_images=include_images,
        include_shapes=include_shapes,
    )
    return XlsxIngestor(options).ingest(Path(workbook))


# The legacy openpyxl double-load ingestor, kept as the guaranteed fallback.
LegacyOpenpyxlIngestor = XlsxIngestor


__all__ = [
    "LegacyOpenpyxlIngestor",
    "XlsxIngestOptions",
    "XlsxIngestor",
    "attach_drawings",
    "bind_manifest_assets",
    "ingest_xlsx",
]
