"""Sparse OOXML ingestor.

Loads the workbook **once** (openpyxl, ``data_only=False``) to obtain style
tables, merges, properties, and the sparse set of real cells — openpyxl's
``_cells`` already holds only cells present in the XML, so nothing is densified.
Cached formula results are read by streaming each sheet's XML for the ``<v>``
under an ``<f>`` in the *same* ``<c>`` node, eliminating the legacy second
``load_workbook(data_only=True)``.

Content bounds are computed from value cells + merge extents only, so
distant style-only cells and an inflated ``<dimension>`` never inflate the area
later materialised by :func:`excelspec.ingest.adapter.sparse_to_document`.
"""

from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

from ..models.document_ir import DocumentIR
from .base import UnsupportedWorkbookError
from .ooxml import workbook_sheet_parts
from .sparse_model import SparseCell, SparseSheet, SparseWorkbookIR
from .workbook import (
    XlsxIngestOptions,
    _display_value,
    _formula_text,
    _json_value,
    _merge_map,
    _style,
    attach_drawings,
    bind_manifest_assets,
)

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _convert_cached(t: str | None, v_text: str | None) -> Any:
    """Convert a streamed cached ``<v>`` to a Python value like data_only load."""

    if v_text is None:
        return None
    if t == "b":
        return v_text not in ("0", "false", "False", "")
    if t in ("e", "str"):
        return v_text
    # numeric (t absent or "n")
    try:
        if "." in v_text or "e" in v_text.lower():
            return float(v_text)
        return int(v_text)
    except ValueError:
        return v_text


def _stream_cached_values(
    archive: zipfile.ZipFile, part: str | None
) -> dict[str, Any]:
    """Return coordinate -> cached Python value for every formula cell.

    Streams the worksheet XML with ``iterparse`` so a large sheet is not held in
    memory. Only ``<c>`` nodes containing an ``<f>`` contribute.
    """

    if not part:
        return {}
    cached: dict[str, Any] = {}
    with archive.open(part) as handle:
        for _event, element in ET.iterparse(handle, events=("end",)):
            if _local(element.tag) != "c":
                continue
            has_formula = any(_local(child.tag) == "f" for child in element)
            if not has_formula:
                element.clear()
                continue
            coord = element.get("r")
            t = element.get("t")
            v_text = None
            for child in element:
                if _local(child.tag) == "v":
                    v_text = child.text
                    break
            if coord is not None:
                cached[coord] = _convert_cached(t, v_text)
            element.clear()
    return cached


class SparseOoxmlIngestor:
    def __init__(self, options: XlsxIngestOptions | None = None) -> None:
        self.options = options or XlsxIngestOptions()

    def ingest(self, workbook: Path) -> DocumentIR:
        from .adapter import sparse_to_document

        workbook_path = Path(workbook).resolve()
        if not workbook_path.is_file():
            raise FileNotFoundError(workbook_path)
        try:
            book = load_workbook(workbook_path, data_only=False, read_only=False)
        except zipfile.BadZipFile as error:  # not an OOXML zip / corrupt
            raise UnsupportedWorkbookError(f"无法作为 OOXML 打开: {error}") from error
        asset_dir = (
            self.options.asset_dir.resolve()
            if self.options.asset_dir
            else workbook_path.parent / f"{workbook_path.stem}_assets"
        )
        try:
            sparse = self._build_sparse(book, workbook_path)
            document, sheets = sparse_to_document(sparse, asset_dir=asset_dir)
            if self.options.include_images or self.options.include_shapes:
                document.diagnostics.extend(
                    attach_drawings(
                        sheets,
                        workbook_path,
                        asset_dir,
                        include_images=self.options.include_images,
                        include_shapes=self.options.include_shapes,
                    )
                )
            if self.options.screenshot_manifest:
                document.diagnostics.extend(
                    bind_manifest_assets(
                        self.options.screenshot_manifest, sheets, workbook_path
                    )
                )
            return document
        finally:
            book.close()

    def build_sparse_workbook(self, workbook: Path) -> SparseWorkbookIR:
        """Return the sparse IR with drawings/manifest attached to each sheet.

        Used by the zero-config detection path: the detector consumes sparse
        cells and drawing anchors without any grid materialisation.
        """

        workbook_path = Path(workbook).resolve()
        if not workbook_path.is_file():
            raise FileNotFoundError(workbook_path)
        try:
            book = load_workbook(workbook_path, data_only=False, read_only=False)
        except zipfile.BadZipFile as error:
            raise UnsupportedWorkbookError(f"无法作为 OOXML 打开: {error}") from error
        asset_dir = (
            self.options.asset_dir.resolve()
            if self.options.asset_dir
            else workbook_path.parent / f"{workbook_path.stem}_assets"
        )
        try:
            sparse = self._build_sparse(book, workbook_path)
            if self.options.include_images or self.options.include_shapes:
                sparse.document_diagnostics.extend(
                    attach_drawings(
                        sparse.sheets,
                        workbook_path,
                        asset_dir,
                        include_images=self.options.include_images,
                        include_shapes=self.options.include_shapes,
                    )
                )
            if self.options.screenshot_manifest:
                sparse.document_diagnostics.extend(
                    bind_manifest_assets(
                        self.options.screenshot_manifest, sparse.sheets, workbook_path
                    )
                )
            sparse.metadata["asset_directory"] = str(asset_dir)
            return sparse
        finally:
            book.close()

    # -- sparse construction ---------------------------------------------------

    def _build_sparse(self, book, workbook_path: Path) -> SparseWorkbookIR:
        styles: dict[int, Any] = {}
        style_index: dict[str, int] = {}

        def style_id_for(cell) -> int | None:
            style = _style(cell)
            if style is None:
                return None
            key = json.dumps(style.to_dict(), sort_keys=True, ensure_ascii=False)
            existing = style_index.get(key)
            if existing is not None:
                return existing
            new_id = len(styles)
            style_index[key] = new_id
            styles[new_id] = style
            return new_id

        sparse_sheets: list[SparseSheet] = []
        total_xml = total_value = total_style_only = 0
        with zipfile.ZipFile(workbook_path) as archive:
            parts = workbook_sheet_parts(archive)
            for index, worksheet in enumerate(book.worksheets):
                members, spans = _merge_map(worksheet)
                formula_coords = [
                    cell.coordinate
                    for cell in worksheet._cells.values()
                    if isinstance(cell, Cell) and cell.data_type == "f"
                ]
                cached_map = (
                    _stream_cached_values(archive, parts.get(worksheet.title))
                    if formula_coords
                    else {}
                )

                cells: dict[tuple[int, int], SparseCell] = {}
                style_only: dict[tuple[int, int], int] = {}
                value_coords: list[tuple[int, int]] = []
                for cell in worksheet._cells.values():
                    # A merged member with no style is materialised as a plain
                    # empty cell (data_type 'n', no style) — matching legacy — so
                    # skip it here. A styled member (e.g. a merge border) must be
                    # recorded so its style survives. Real Cell nodes (including
                    # typed-empty inlineStr) are always recorded.
                    is_merged_placeholder = isinstance(cell, MergedCell)
                    if is_merged_placeholder and not cell.has_style:
                        continue
                    row, column = cell.row, cell.column
                    has_value = cell.value is not None
                    formula = (
                        _formula_text(cell.value, cell.data_type) if has_value else None
                    )
                    if formula is not None:
                        cached = cached_map.get(cell.coordinate)
                    else:
                        cached = cell.value
                    style_id = style_id_for(cell)
                    cells[(row, column)] = SparseCell(
                        row=row,
                        column=column,
                        coordinate=cell.coordinate,
                        raw_value=_json_value(cell.value),
                        display_value=_display_value(cached),
                        data_type=cell.data_type,
                        formula=formula,
                        cached_value=cached,
                        style_id=style_id,
                    )
                    if is_merged_placeholder:
                        # Styled merge border: not a real value/content cell and
                        # not counted in sparse stats or content bounds.
                        continue
                    if has_value:
                        value_coords.append((row, column))
                    elif style_id is not None:
                        style_only[(row, column)] = style_id

                bounds = _content_bounds(value_coords, worksheet)
                sparse_sheets.append(
                    SparseSheet(
                        name=worksheet.title,
                        sheet_id=f"sheet-{index + 1}",
                        index=index,
                        state=worksheet.sheet_state,
                        cells=cells,
                        style_only=style_only,
                        merges=[str(item) for item in worksheet.merged_cells.ranges],
                        merge_spans=spans,
                        merge_members=members,
                        content_bounds=bounds,
                    )
                )
                # Stats count real content cells only (value + style-only),
                # never merged-border placeholders.
                total_value += len(value_coords)
                total_style_only += len(style_only)
                total_xml += len(value_coords) + len(style_only)

        properties = book.properties
        return SparseWorkbookIR(
            path=str(workbook_path),
            sheets=sparse_sheets,
            styles=styles,
            properties={
                "title": properties.title,
                "creator": properties.creator,
                "last_modified_by": properties.lastModifiedBy,
            },
            metadata={
                "sparse_stats": {
                    "xml_cell_count": total_xml,
                    "value_cell_count": total_value,
                    "style_only_cell_count": total_style_only,
                    "style_count": len(styles),
                }
            },
        )


def _content_bounds(
    value_coords: list[tuple[int, int]], worksheet
) -> tuple[int, int, int, int] | None:
    """Bounds over value cells + merge extents only (never distant styles)."""

    coords = list(value_coords)
    for merged_range in worksheet.merged_cells.ranges:
        coords.append((merged_range.min_row, merged_range.min_col))
        coords.append((merged_range.max_row, merged_range.max_col))
    if not coords:
        return (1, 1, 1, 1)
    rows = [row for row, _ in coords]
    columns = [column for _, column in coords]
    return min(rows), min(columns), max(rows), max(columns)


__all__ = ["SparseOoxmlIngestor"]
