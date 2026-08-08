"""Low-level OOXML drawing extraction independent of openpyxl internals."""

from __future__ import annotations

import mimetypes
import posixpath
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl.utils import get_column_letter

REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


@dataclass(slots=True)
class DrawingAsset:
    kind: str
    uri: str
    anchor: str | None = None
    media_type: str | None = None
    description: str | None = None
    metadata: dict[str, object] = field(default_factory=dict)


@dataclass(slots=True)
class DrawingDiagnostic:
    code: str
    message: str
    details: dict[str, object] = field(default_factory=dict)


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _resolve_part(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _rels_part(part: str) -> str:
    directory, name = posixpath.split(part)
    return posixpath.join(directory, "_rels", f"{name}.rels")


def _relationships(zf: zipfile.ZipFile, part: str) -> dict[str, tuple[str, str | None]]:
    rels_name = _rels_part(part)
    if rels_name not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read(rels_name))
    result: dict[str, tuple[str, str | None]] = {}
    for rel in root:
        if _local(rel.tag) == "Relationship" and rel.get("Id") and rel.get("Target"):
            result[rel.get("Id", "")] = (rel.get("Target", ""), rel.get("TargetMode"))
    return result


def workbook_sheet_parts(zf: zipfile.ZipFile) -> dict[str, str]:
    """Return display sheet name -> actual worksheet OOXML part."""

    workbook_part = "xl/workbook.xml"
    root = ET.fromstring(zf.read(workbook_part))
    rels = _relationships(zf, workbook_part)
    mapping: dict[str, str] = {}
    relationship_id = f"{{{DOC_REL_NS}}}id"
    for element in root.iter():
        if _local(element.tag) != "sheet":
            continue
        name, rid = element.get("name"), element.get(relationship_id)
        if name and rid and rid in rels:
            target, mode = rels[rid]
            if mode != "External":
                mapping[name] = _resolve_part(workbook_part, target)
    return mapping


def _anchor_position(anchor: ET.Element) -> str | None:
    points: dict[str, tuple[int, int]] = {}
    for child in anchor:
        local = _local(child.tag)
        if local not in {"from", "to"}:
            continue
        row = col = None
        for coordinate in child:
            if _local(coordinate.tag) == "row" and coordinate.text is not None:
                row = int(coordinate.text) + 1
            elif _local(coordinate.tag) == "col" and coordinate.text is not None:
                col = int(coordinate.text) + 1
        if row is not None and col is not None:
            points[local] = (row, col)
    if "from" not in points:
        return None
    start = f"{get_column_letter(points['from'][1])}{points['from'][0]}"
    if "to" not in points:
        return start
    end = f"{get_column_letter(points['to'][1])}{points['to'][0]}"
    return start if start == end else f"{start}:{end}"


def _anchor_row(anchor_ref: str | None) -> int | None:
    if not anchor_ref:
        return None
    match = re.search(r"(\d+)", anchor_ref)
    return int(match.group(1)) if match else None


def _drawing_ids(zf: zipfile.ZipFile, sheet_part: str) -> list[str]:
    root = ET.fromstring(zf.read(sheet_part))
    rid_name = f"{{{DOC_REL_NS}}}id"
    return [
        element.get(rid_name, "")
        for element in root.iter()
        if _local(element.tag) == "drawing" and element.get(rid_name)
    ]


def _object_description(element: ET.Element) -> str | None:
    for child in element.iter():
        if _local(child.tag) == "cNvPr":
            description = (child.get("descr") or "").strip()
            name = (child.get("name") or "").strip()
            if description:
                return description
            if name and not re.match(r"^(Picture|Shape)\s*\d+$", name, re.I):
                return name
            return None
    return None


def _shape_text(element: ET.Element) -> str:
    lines: list[str] = []
    for paragraph in (item for item in element.iter() if _local(item.tag) == "p"):
        text = "".join(
            node.text or "" for node in paragraph.iter() if _local(node.tag) == "t"
        ).strip()
        if text:
            lines.append(text)
    if lines:
        return "\n".join(lines)
    return "".join(
        node.text or "" for node in element.iter() if _local(node.tag) == "t"
    ).strip()


def _picture_embed_id(element: ET.Element) -> str | None:
    embed_name = f"{{{DOC_REL_NS}}}embed"
    for child in element.iter():
        if _local(child.tag) == "blip" and child.get(embed_name):
            return child.get(embed_name)
    return None


def _safe_filename(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .")
    return value[:100] or "asset"


def extract_sheet_drawings(
    zf: zipfile.ZipFile,
    *,
    sheet_part: str,
    sheet_name: str,
    output_dir: Path,
    include_images: bool = True,
    include_shapes: bool = True,
) -> tuple[list[DrawingAsset], list[DrawingDiagnostic]]:
    assets: list[DrawingAsset] = []
    diagnostics: list[DrawingDiagnostic] = []
    try:
        sheet_rels = _relationships(zf, sheet_part)
        drawing_ids = _drawing_ids(zf, sheet_part)
    except (ET.ParseError, KeyError, ValueError) as error:
        return [], [DrawingDiagnostic("OOXML_SHEET_DRAWING_INVALID", str(error))]

    image_number = shape_number = 0
    for drawing_id in drawing_ids:
        relationship = sheet_rels.get(drawing_id)
        if relationship is None:
            diagnostics.append(
                DrawingDiagnostic(
                    "OOXML_DRAWING_RELATIONSHIP_MISSING",
                    f"工作表 drawing 关系 {drawing_id} 不存在",
                    {"relationship_id": drawing_id},
                )
            )
            continue
        target, mode = relationship
        drawing_part = _resolve_part(sheet_part, target)
        if mode == "External" or drawing_part not in zf.namelist():
            diagnostics.append(
                DrawingDiagnostic(
                    "OOXML_DRAWING_PART_MISSING",
                    f"无法读取 drawing 部件 {drawing_part}",
                    {"relationship_id": drawing_id, "target": target},
                )
            )
            continue
        try:
            drawing_root = ET.fromstring(zf.read(drawing_part))
            drawing_rels = _relationships(zf, drawing_part)
        except (ET.ParseError, KeyError) as error:
            diagnostics.append(
                DrawingDiagnostic(
                    "OOXML_DRAWING_INVALID",
                    f"drawing 部件解析失败: {drawing_part}",
                    {"error": str(error)},
                )
            )
            continue

        for anchor in drawing_root:
            if _local(anchor.tag) not in {"oneCellAnchor", "twoCellAnchor", "absoluteAnchor"}:
                continue
            anchor_ref = _anchor_position(anchor)
            for drawing_object in anchor:
                object_kind = _local(drawing_object.tag)
                if object_kind == "pic":
                    if not include_images:
                        continue
                    image_number += 1
                    embed_id = _picture_embed_id(drawing_object)
                    relationship = drawing_rels.get(embed_id or "")
                    if relationship is None:
                        diagnostics.append(
                            DrawingDiagnostic(
                                "OOXML_IMAGE_RELATIONSHIP_MISSING",
                                f"图片关系 {embed_id or '<empty>'} 不存在",
                                {"drawing_part": drawing_part, "anchor": anchor_ref},
                            )
                        )
                        continue
                    media_target, media_mode = relationship
                    media_part = _resolve_part(drawing_part, media_target)
                    if media_mode == "External" or media_part not in zf.namelist():
                        diagnostics.append(
                            DrawingDiagnostic(
                                "OOXML_IMAGE_PART_MISSING",
                                f"无法读取图片部件 {media_part}",
                                {"drawing_part": drawing_part, "anchor": anchor_ref},
                            )
                        )
                        continue
                    extension = Path(media_part).suffix.lower() or ".bin"
                    stem = _safe_filename(f"{sheet_name}-image-{image_number}")
                    output_path = output_dir / f"{stem}{extension}"
                    output_dir.mkdir(parents=True, exist_ok=True)
                    output_path.write_bytes(zf.read(media_part))
                    assets.append(
                        DrawingAsset(
                            kind="image",
                            uri=str(output_path.resolve()),
                            anchor=anchor_ref,
                            media_type=mimetypes.guess_type(output_path.name)[0],
                            description=_object_description(drawing_object),
                            metadata={
                                "ooxml_part": media_part,
                                "drawing_part": drawing_part,
                                "relationship_id": embed_id or "",
                                "drawing_order": image_number,
                                "anchor_row": _anchor_row(anchor_ref),
                            },
                        )
                    )
                elif object_kind in {"sp", "grpSp", "cxnSp"}:
                    if not include_shapes:
                        continue
                    text = _shape_text(drawing_object)
                    if not text:
                        continue
                    shape_number += 1
                    assets.append(
                        DrawingAsset(
                            kind="shape",
                            uri=f"ooxml://{drawing_part}#shape-{shape_number}",
                            anchor=anchor_ref,
                            description=text,
                            metadata={
                                "text": text,
                                "drawing_part": drawing_part,
                                "shape_kind": object_kind,
                            },
                        )
                    )
    return assets, diagnostics


__all__ = [
    "DrawingAsset",
    "DrawingDiagnostic",
    "extract_sheet_drawings",
    "workbook_sheet_parts",
]
