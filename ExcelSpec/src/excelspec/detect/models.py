"""Candidate region model shared by the detector and router."""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import StrEnum
from typing import Any

from openpyxl.utils import get_column_letter

from ..models.document_ir import DiagnosticIR, RegionType


class CandidateRegionType(StrEnum):
    TABLE = "table"
    KEY_VALUE = "key_value"
    TEXT = "text"
    IMAGE = "image"
    SHAPE = "shape"
    LAYOUT = "layout"
    FREEFORM = "freeform"

    def to_region_type(self) -> RegionType:
        """Map to the canonical DocumentIR ``RegionType``."""

        mapping = {
            CandidateRegionType.TABLE: RegionType.TABLE,
            CandidateRegionType.KEY_VALUE: RegionType.KEY_VALUE,
            CandidateRegionType.TEXT: RegionType.FREEFORM,
            CandidateRegionType.IMAGE: RegionType.IMAGE,
            CandidateRegionType.SHAPE: RegionType.FREEFORM,
            CandidateRegionType.LAYOUT: RegionType.LAYOUT,
            CandidateRegionType.FREEFORM: RegionType.FREEFORM,
        }
        return mapping[self]


@dataclass(frozen=True, slots=True)
class CellBounds:
    """Inclusive 1-based rectangle ``(min_row, min_col, max_row, max_col)``."""

    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def row_count(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def col_count(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def area(self) -> int:
        return self.row_count * self.col_count

    def as_tuple(self) -> tuple[int, int, int, int]:
        return (self.min_row, self.min_col, self.max_row, self.max_col)

    def a1(self) -> str:
        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )

    def contains(self, row: int, col: int) -> bool:
        return (
            self.min_row <= row <= self.max_row
            and self.min_col <= col <= self.max_col
        )


@dataclass(slots=True)
class CandidateRegion:
    """An explainable region proposal produced from sparse cells only."""

    region_id: str
    sheet_name: str
    bounds: CellBounds
    region_type: CandidateRegionType
    confidence: float
    detection_method: str
    features: dict[str, Any] = field(default_factory=dict)
    title: str | None = None
    title_cell: str | None = None   # coordinate of the peeled heading, if any
    source_cells: list[str] = field(default_factory=list)
    asset_refs: list[str] = field(default_factory=list)
    diagnostics: list[DiagnosticIR] = field(default_factory=list)


__all__ = [
    "CandidateRegion",
    "CandidateRegionType",
    "CellBounds",
]
