"""Zero-config assembly: SparseWorkbookIR -> DocumentIR via detect + route."""

from __future__ import annotations

from pathlib import Path

from openpyxl.utils import get_column_letter, range_boundaries

from ..ingest.sparse_model import SparseWorkbookIR
from ..models.document_ir import (
    AssetIR,
    AssetType,
    DiagnosticIR,
    DiagnosticSeverity,
    DocumentIR,
    SheetIR,
)
from ..profile.enrich import enrich_regions, profile_required_diagnostics
from ..profile.model import SemanticProfile
from ..providers import (
    NullOcrProvider,
    NullVlmProvider,
    OcrProvider,
    ProviderContext,
    VlmProvider,
)
from .detector import RegionDetector
from .models import CandidateRegion, CandidateRegionType
from .router import RegionRouter


def _override_for(profile: SemanticProfile | None, sheet_name: str, role: str | None):
    if profile is None:
        return None
    for override in profile.overrides:
        if override.sheet and override.sheet == sheet_name:
            return override
        if override.sheet_alias and override.sheet_alias == role:
            return override
    return None


def _ignored(bounds, ignore_ranges: list[str]) -> bool:
    for ignore in ignore_ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ignore)
        except ValueError:
            continue
        if (
            bounds.min_row >= min_row
            and bounds.max_row <= max_row
            and bounds.min_col >= min_col
            and bounds.max_col <= max_col
        ):
            return True
    return False


def _parse_bounds(range_ref: str | None) -> tuple[int, int, int, int] | None:
    """Return ``(min_row, min_col, max_row, max_col)`` for an A1 reference."""

    if not range_ref:
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    except ValueError:
        return None
    return min_row, min_col, max_row, max_col


def _bounds_a1(bounds: tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def _union_bounds(
    values: list[tuple[int, int, int, int]],
) -> tuple[int, int, int, int] | None:
    if not values:
        return None
    return (
        min(value[0] for value in values),
        min(value[1] for value in values),
        max(value[2] for value in values),
        max(value[3] for value in values),
    )


def _bounds_area(bounds: tuple[int, int, int, int]) -> int:
    return (bounds[2] - bounds[0] + 1) * (bounds[3] - bounds[1] + 1)


def _bounds_distance(
    left: tuple[int, int, int, int], right: tuple[int, int, int, int]
) -> int:
    row_gap = max(0, right[0] - left[2] - 1, left[0] - right[2] - 1)
    col_gap = max(0, right[1] - left[3] - 1, left[1] - right[3] - 1)
    return row_gap + col_gap


def _is_embedded_raster(asset: AssetIR) -> bool:
    return asset.asset_type in {AssetType.IMAGE, AssetType.LAYOUT, AssetType.CHART}


def _bordered_positions(sparse_sheet, styles) -> set[tuple[int, int]]:
    positions: set[tuple[int, int]] = set()
    for position, style_id in sparse_sheet.style_only.items():
        style = styles.get(style_id)
        if style and style.border:
            positions.add(position)
    for position, cell in sparse_sheet.cells.items():
        style = styles.get(cell.style_id) if cell.style_id is not None else None
        if style and style.border:
            positions.add(position)
    return positions


def _expand_visual_bounds(
    sparse_sheet,
    styles,
    base: tuple[int, int, int, int],
) -> tuple[int, int, int, int]:
    """Expand a capture over its connected bordered/merged visual component.

    Only sparse, explicitly styled positions are inspected. A distant style at
    XFD1048576 can therefore never inflate an unrelated screenshot range.
    """

    bordered = _bordered_positions(sparse_sheet, styles)
    min_row, min_col, max_row, max_col = base
    seeds = {
        position
        for position in bordered
        if min_row - 1 <= position[0] <= max_row + 1
        and min_col - 1 <= position[1] <= max_col + 1
    }
    component: set[tuple[int, int]] = set()
    pending = list(seeds)
    while pending and len(component) < 50_000:
        position = pending.pop()
        if position in component:
            continue
        component.add(position)
        row, col = position
        for neighbour in ((row - 1, col), (row + 1, col), (row, col - 1), (row, col + 1)):
            if neighbour in bordered and neighbour not in component:
                pending.append(neighbour)
    if component:
        base = _union_bounds(
            [
                base,
                (
                    min(row for row, _ in component),
                    min(col for _, col in component),
                    max(row for row, _ in component),
                    max(col for _, col in component),
                ),
            ]
        ) or base

    # Merges are cheap and commonly define the outside edge of a cell-drawn UI.
    changed = True
    while changed:
        changed = False
        for merge in sparse_sheet.merges:
            merge_bounds = _parse_bounds(merge)
            if merge_bounds is None or _bounds_distance(base, merge_bounds) > 1:
                continue
            expanded = _union_bounds([base, merge_bounds]) or base
            if expanded != base:
                base = expanded
                changed = True
    return base


def _visual_diagnostic(region, *, code: str, message: str, details: dict | None = None) -> None:
    region.metadata.setdefault("diagnostics", []).append(
        DiagnosticIR(
            code=code,
            severity=DiagnosticSeverity.INFO,
            message=message,
            source=region.source,
            region_id=region.region_id,
            details=details or {},
        ).to_dict()
    )


def assemble_document(
    sparse: SparseWorkbookIR,
    *,
    mode: str = "fast",
    profile: SemanticProfile | None = None,
    ocr: OcrProvider | None = None,
    vlm: VlmProvider | None = None,
) -> tuple[DocumentIR, dict[str, list[str]]]:
    """Build a DocumentIR from the sparse IR using detection + routing."""

    path = Path(sparse.path)
    detector = RegionDetector()
    router = RegionRouter(sparse)
    sheets: list[SheetIR] = []
    unrecognized: dict[str, list[str]] = {}

    for sparse_sheet in sparse.sheets:
        role = profile.sheet_role(sparse_sheet.name) if profile else None
        override = _override_for(profile, sparse_sheet.name, role)
        if override and override.exclude_sheet:
            continue

        candidates = detector.detect_sheet_regions(sparse_sheet, sparse.styles)
        if override and override.ignore:
            candidates = [
                c for c in candidates if not _ignored(c.bounds, override.ignore)
            ]

        regions = [router.route(sparse_sheet, candidate) for candidate in candidates]
        if override and override.visual_range:
            for region in regions:
                if region.metadata.get("visual"):
                    region.metadata["visual_capture_range_override"] = override.visual_range

        sheet_diagnostics: list[DiagnosticIR] = list(sparse_sheet.diagnostics)
        for candidate in candidates:
            sheet_diagnostics.extend(candidate.diagnostics)
        if profile:
            sheet_diagnostics.extend(enrich_regions(profile, sparse_sheet.name, regions))

        metadata = {"extraction_mode": mode, "detected_region_count": len(regions)}
        if role:
            metadata["sheet_role"] = role
        sheets.append(
            SheetIR(
                sheet_id=sparse_sheet.sheet_id,
                name=sparse_sheet.name,
                index=sparse_sheet.index,
                regions=regions,
                assets=list(sparse_sheet.assets),
                diagnostics=sheet_diagnostics,
                metadata=metadata,
            )
        )
        unrecognized[sparse_sheet.name] = [
            candidate.bounds.a1()
            for candidate in candidates
            if candidate.region_type == CandidateRegionType.FREEFORM
        ]

    if mode in ("auto", "visual"):
        _capture_layout_regions(sparse, sheets)
        _apply_visual_providers(
            sheets, ocr=ocr or NullOcrProvider(), vlm=vlm or NullVlmProvider()
        )

    document = DocumentIR(
        document_id=path.stem,
        title=sparse.properties.get("title") or path.stem,
        source_path=str(path),
        sheets=sheets,
        diagnostics=list(sparse.document_diagnostics),
        metadata={
            "ingestor": "sparse-ooxml",
            "extraction_mode": mode,
            "profile_id": profile.profile_id if profile else None,
            "document_type": profile.document_type if profile else None,
            "asset_directory": sparse.metadata.get("asset_directory"),
            "sparse_stats": sparse.metadata.get("sparse_stats", {}),
        },
    )
    if profile:
        document.diagnostics.extend(profile_required_diagnostics(profile, document))
    return document, unrecognized


def _apply_visual_providers(
    sheets: list[SheetIR], *, ocr: OcrProvider, vlm: VlmProvider
) -> None:
    """Run available OCR/VLM providers over visual regions (auto/visual only).

    Only called when a provider reports ``available``. A provider failure records
    a diagnostic and never discards the region's assets. Results are tagged with
    provider / source / confidence.
    """

    if not getattr(ocr, "available", False) and not getattr(vlm, "available", False):
        return
    for sheet in sheets:
        for region in sheet.regions:
            if not region.metadata.get("visual"):
                continue
            asset_id = region.asset_ids[0] if region.asset_ids else None
            context = ProviderContext(
                sheet=sheet.name,
                region_id=region.region_id,
                title=region.title,
                source_range=region.source.range if region.source else None,
            )
            for provider, kind, method in (
                (vlm, "vlm", "describe"),
                (ocr, "ocr", "extract"),
            ):
                if not getattr(provider, "available", False):
                    continue
                try:
                    result = getattr(provider, method)(asset_id, context)
                except Exception as error:  # noqa: BLE001 - provider is best-effort
                    region.metadata.setdefault("diagnostics", []).append(
                        DiagnosticIR(
                            code=f"provider.{kind}_failed",
                            severity=DiagnosticSeverity.WARNING,
                            message=f"{kind} provider 失败，保留视觉资源: {region.region_id} ({error})",
                            source=region.source,
                            region_id=region.region_id,
                        ).to_dict()
                    )
                    continue
                region.metadata[f"{kind}_result"] = {
                    "text": result.text,
                    "provider": result.provider,
                    "source": result.source,
                    "confidence": result.confidence,
                }


def _capture_layout_regions(sparse: SparseWorkbookIR, sheets: list[SheetIR]) -> None:
    """auto/visual: screenshot layout regions, reusing one Excel session.

    A capture failure never drops structured content — it only records a
    warning on the region.
    """

    sparse_by_name = {sheet.name: sheet for sheet in sparse.sheets}
    targets: list[tuple[SheetIR, object, str, str]] = []
    for sheet in sheets:
        sparse_sheet = sparse_by_name.get(sheet.name)
        if sparse_sheet is None:
            continue
        visual_regions = [region for region in sheet.regions if region.metadata.get("visual")]
        embedded = [asset for asset in sheet.assets if _is_embedded_raster(asset)]
        assets_by_id = {asset.asset_id: asset for asset in sheet.assets}
        for region in visual_regions:
            base = _parse_bounds(region.source.range if region.source else None)
            if base is None:
                continue
            linked_assets = [
                assets_by_id[asset_id]
                for asset_id in region.asset_ids
                if asset_id in assets_by_id
            ]
            linked_rasters = [asset for asset in linked_assets if _is_embedded_raster(asset)]
            if linked_rasters:
                region.metadata["screenshot_strategy"] = "reuse_embedded_asset"
                region.metadata["visual_source_asset_ids"] = [
                    asset.asset_id for asset in linked_rasters
                ]
                _visual_diagnostic(
                    region,
                    code="route.screenshot_reused_embedded_asset",
                    message=f"区域 {region.region_id} 已有嵌入原图，跳过 Excel 截图",
                )
                continue

            # A tiny title-like layout followed by exactly one embedded image is
            # a common Japanese spec pattern. The image is the real visual; A2:A3
            # is merely its heading and must not be captured as a useless strip.
            nearby_embedded = [
                asset
                for asset in embedded
                if (anchor := _parse_bounds(asset.anchor)) is not None
                and _bounds_distance(base, anchor) <= 20
            ]
            if _bounds_area(base) < 12 and len(nearby_embedded) == 1:
                region.metadata["screenshot_strategy"] = "reuse_nearby_embedded_asset"
                region.metadata["visual_source_asset_ids"] = [nearby_embedded[0].asset_id]
                _visual_diagnostic(
                    region,
                    code="route.screenshot_reused_nearby_asset",
                    message=(
                        f"区域 {region.region_id} 是小型标题区域；复用同 Sheet 的嵌入图片，"
                        "跳过标题截图"
                    ),
                    details={"asset_id": nearby_embedded[0].asset_id},
                )
                continue

            override = region.metadata.get("visual_capture_range_override")
            capture_bounds = _parse_bounds(override) if isinstance(override, str) else None
            strategy = "profile_override" if capture_bounds else "detected_visual_bounds"
            if capture_bounds is None:
                anchor_bounds = [
                    parsed
                    for asset in linked_assets
                    if asset.asset_type == AssetType.SHAPE
                    and (parsed := _parse_bounds(asset.anchor)) is not None
                ]
                capture_bounds = _union_bounds([base, *anchor_bounds]) or base
                capture_bounds = _expand_visual_bounds(
                    sparse_sheet, sparse.styles, capture_bounds
                )
                if anchor_bounds:
                    strategy = "shape_anchor_union"

            if _bounds_area(capture_bounds) < 12:
                region.metadata["screenshot_strategy"] = "skip_tiny_visual_range"
                _visual_diagnostic(
                    region,
                    code="route.screenshot_skipped_tiny_range",
                    message=(
                        f"区域 {region.region_id} 的视觉范围过小 "
                        f"({_bounds_a1(capture_bounds)})，为避免无意义截图而跳过"
                    ),
                )
                continue
            capture_range = _bounds_a1(capture_bounds)
            region.metadata["screenshot_strategy"] = strategy
            region.metadata["visual_capture_range"] = capture_range
            targets.append((sheet, region, capture_range, strategy))
    if not targets:
        return

    from ..render import ExcelCaptureSession

    asset_root = Path(sparse.metadata.get("asset_directory") or ".")
    with ExcelCaptureSession(sparse.path) as session:
        for sheet, region, capture_range, strategy in targets:
            destination = asset_root / "screenshots" / f"{sheet.sheet_id}-{region.region_id}.png"
            try:
                image = session.capture(sheet.name, capture_range, destination)
            except Exception as error:  # noqa: BLE001 - screenshot is best-effort
                region.metadata.setdefault("diagnostics", []).append(
                    DiagnosticIR(
                        code="route.screenshot_failed",
                        severity=DiagnosticSeverity.WARNING,
                        message=f"截图失败，保留结构化内容: {region.region_id} ({error})",
                        source=region.source,
                        region_id=region.region_id,
                    ).to_dict()
                )
                continue
            asset_id = f"{sheet.sheet_id}-{region.region_id}-screenshot"
            sheet.assets.append(
                AssetIR(
                    asset_id=asset_id,
                    asset_type=AssetType.SCREENSHOT,
                    uri=str(image),
                    media_type="image/png",
                    description=region.title or region.region_id,
                    source=region.source,
                    anchor=capture_range,
                    extraction_status="rendered",
                    metadata={
                        "capture_method": "excel_com",
                        "capture_strategy": strategy,
                        "region_id": region.region_id,
                        "capture_range": capture_range,
                    },
                )
            )
            region.asset_ids.append(asset_id)


__all__ = ["assemble_document"]
