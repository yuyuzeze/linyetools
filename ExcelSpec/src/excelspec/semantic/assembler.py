"""Assemble a SemanticDocumentIR from a routed DocumentIR (forward transform)."""

from __future__ import annotations

from openpyxl.utils import column_index_from_string, get_column_letter

from ..models.document_ir import CellIR, DocumentIR, RegionIR, SheetIR, TableIR
from ..models.semantic import (
    KeyValueEntry,
    SemanticAsset,
    SemanticColumn,
    SemanticDocumentIR,
    SemanticRegion,
    SemanticRegionType,
    SemanticRow,
    SemanticSheet,
    SemanticTable,
)
from .references import extract_references


def _cell_text(cell: CellIR | None) -> str:
    if cell is None:
        return ""
    if cell.display_value is not None:
        return str(cell.display_value).strip()
    if cell.raw_value is None:
        return ""
    return str(cell.raw_value).strip()


def _cell_value(cell: CellIR | None):
    if cell is None:
        return None
    if cell.formula is not None:
        return cell.display_value
    return cell.raw_value if cell.raw_value is not None else cell.display_value


def _region_semantic_type(region: RegionIR) -> SemanticRegionType:
    candidate = region.metadata.get("candidate_type")
    if isinstance(candidate, str):
        try:
            return SemanticRegionType(candidate)
        except ValueError:
            pass
    return SemanticRegionType(region.region_type.value)


def _reading_text(cells: list[CellIR]) -> str:
    by_row: dict[int, list[CellIR]] = {}
    for cell in cells:
        if _cell_text(cell):
            by_row.setdefault(cell.row, []).append(cell)
    lines = []
    for row in sorted(by_row):
        ordered = sorted(by_row[row], key=lambda c: c.column)
        lines.append(" ".join(_cell_text(c) for c in ordered))
    return "\n".join(lines)


def _build_table(table: TableIR, region: RegionIR) -> SemanticTable:
    cells = table.cells
    if not cells:
        return SemanticTable(header_rows=table.header_rows)
    by_position = {(c.row, c.column): c for c in cells}
    columns_present = sorted({c.column for c in cells})
    min_row = min(c.row for c in cells)
    max_row = max(c.row for c in cells)
    header_rows = max(1, table.header_rows)

    header_labels = table.metadata.get("header_labels", {}) or {}
    field_mapping = {
        item["source_column"]: item
        for item in (table.metadata.get("field_mapping", []) or [])
        if isinstance(item, dict) and item.get("source_column")
    }

    columns: list[SemanticColumn] = []
    key_for_column: dict[int, str] = {}
    for column in columns_present:
        letter = get_column_letter(column)
        source_header = header_labels.get(letter) or None
        mapping = field_mapping.get(letter, {})
        semantic_name = table.column_semantics.get(letter) or mapping.get("semantic_name")
        confidence = float(mapping.get("confidence", 1.0 if semantic_name else 0.0))
        columns.append(
            SemanticColumn(
                column_id=letter,
                source_header=source_header,
                semantic_name=semantic_name,
                display_name=source_header or letter,
                confidence=confidence,
            )
        )
        # stable key: semantic name if mapped, else the column id
        key_for_column[column] = semantic_name or letter

    rows: list[SemanticRow] = []
    for row in range(min_row + header_rows, max_row + 1):
        row_cells = [by_position.get((row, col)) for col in columns_present]
        if all(_cell_text(c) == "" for c in row_cells):
            continue  # do not emit knowledge for a fully blank row
        values: dict = {}
        source_values: dict = {}
        formulas: dict = {}
        for column in columns_present:
            cell = by_position.get((row, column))
            key = key_for_column[column]
            header = header_labels.get(get_column_letter(column)) or get_column_letter(column)
            # missing cell -> explicit None in place (never a leftward shift)
            values[key] = _cell_value(cell)
            source_values[header] = _cell_value(cell)
            if cell is not None and cell.formula is not None:
                formulas[key] = cell.formula
        row_range = (
            f"{get_column_letter(columns_present[0])}{row}:"
            f"{get_column_letter(columns_present[-1])}{row}"
        )
        rows.append(
            SemanticRow(
                row_id=f"row-{row}",
                source_range=row_range,
                values=values,
                source_values=source_values,
                formulas=formulas,
                confidence=region.confidence or 1.0,
            )
        )
    return SemanticTable(columns=columns, rows=rows, header_rows=header_rows)


def _build_key_values(region: RegionIR) -> list[KeyValueEntry]:
    entries: list[KeyValueEntry] = []
    for key, value in region.values.items():
        entries.append(
            KeyValueEntry(
                key=key,
                semantic_name=None,
                value=value,
                confidence=region.confidence or 1.0,
            )
        )
    return entries


def _region_cells(region: RegionIR) -> list[CellIR]:
    return [cell for table in region.tables for cell in table.cells]


def assemble_semantic(
    document: DocumentIR,
    *,
    profile_id: str | None = None,
    processing_mode: str | None = None,
    source_hash: str | None = None,
    document_type: str | None = None,
) -> SemanticDocumentIR:
    """Build a SemanticDocumentIR. Deterministic: stable region/row/column order."""

    references = extract_references(document)
    refs_by_cell: dict[tuple[str, str], list[str]] = {}
    for reference in references:
        refs_by_cell.setdefault(
            (reference.source_sheet, reference.source_cell), []
        ).append(reference.reference_id)

    semantic_sheets: list[SemanticSheet] = []
    semantic_regions: list[SemanticRegion] = []
    semantic_assets: list[SemanticAsset] = []
    referenced_assets: set[str] = set()

    doc_title = document.title
    section_root = [doc_title] if doc_title else []

    for sheet in document.sheets:
        role = sheet.metadata.get("sheet_role")
        region_ids: list[str] = []
        for region in sheet.regions:
            sem_type = _region_semantic_type(region)
            cells = _region_cells(region)
            section_path = [*section_root, sheet.name]
            if region.title:
                section_path = [*section_path, region.title]

            sem_region = SemanticRegion(
                region_id=f"{sheet.sheet_id}:{region.region_id}",
                region_type=sem_type,
                sheet=sheet.name,
                sheet_role=role if isinstance(role, str) else None,
                title=region.title,
                section_path=section_path,
                source_range=region.source.range if region.source else None,
                confidence=region.confidence if region.confidence is not None else 0.0,
                detection_method=region.metadata.get("detection_method"),
                asset_refs=list(region.asset_ids),
                metadata={
                    key: value
                    for key, value in region.metadata.items()
                    if key in (
                        "features",
                        "candidate_type",
                        "materialized_cell_count",
                        "visual",
                        "title_range",
                        "header_decision",
                    )
                },
                diagnostics=list(region.metadata.get("diagnostics", []) or []),
            )
            referenced_assets.update(region.asset_ids)

            # formula references anchored inside this region
            formula_refs: list[str] = []
            for cell in cells:
                formula_refs.extend(refs_by_cell.get((sheet.name, cell.coordinate), []))
            sem_region.formula_refs = formula_refs

            if sem_type == SemanticRegionType.TABLE and region.tables:
                sem_region.table = _build_table(region.tables[0], region)
            elif sem_type == SemanticRegionType.KEY_VALUE:
                sem_region.key_values = _build_key_values(region)
                sem_region.text = _reading_text(cells)
            elif sem_type in (
                SemanticRegionType.IMAGE,
                SemanticRegionType.SHAPE,
                SemanticRegionType.LAYOUT,
            ):
                parts = [region.title] if region.title else []
                body = _reading_text(cells)
                if body:
                    parts.append(body)
                sem_region.text = "\n".join(parts) or None
            else:  # text / freeform
                sem_region.text = _reading_text(cells)

            semantic_regions.append(sem_region)
            region_ids.append(sem_region.region_id)

        for asset in sheet.assets:
            semantic_assets.append(
                SemanticAsset(
                    asset_id=asset.asset_id,
                    asset_type=asset.asset_type.value,
                    uri=asset.uri,
                    sheet=sheet.name,
                    description=asset.description,
                    anchor=asset.anchor,
                    referenced=asset.asset_id in referenced_assets,
                    metadata=dict(asset.metadata),
                )
            )

        semantic_sheets.append(
            SemanticSheet(
                sheet_id=sheet.sheet_id,
                name=sheet.name,
                index=sheet.index,
                sheet_role=role if isinstance(role, str) else None,
                region_ids=region_ids,
            )
        )

    sections = [
        {"path": [*section_root, sheet.name], "sheet": sheet.name}
        for sheet in document.sheets
    ]

    return SemanticDocumentIR(
        document_id=document.document_id,
        title=document.title,
        document_type=document_type or document.metadata.get("document_type"),
        source_path=document.source_path,
        source_hash=source_hash,
        profile_id=profile_id or document.metadata.get("profile_id"),
        processing_mode=processing_mode or document.metadata.get("extraction_mode"),
        sheets=semantic_sheets,
        sections=sections,
        regions=semantic_regions,
        assets=semantic_assets,
        references=references,
        diagnostics=[d.to_dict() for d in document.diagnostics],
        metadata={"ingestor": document.metadata.get("ingestor")},
    )


__all__ = ["assemble_semantic"]
