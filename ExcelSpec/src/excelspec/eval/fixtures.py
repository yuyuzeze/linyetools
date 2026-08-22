"""Programmatic annotated evaluation cases mimicking JP spec layouts.

Every case is a small openpyxl workbook built at run time (never committed) with
a hand annotation of the regions/headers a correct detector should produce.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import EvalCase, ExpectedRegion

_BOLD = Font(bold=True)
_FILL = PatternFill("solid", fgColor="DDDDDD")
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE = Font(bold=True, size=14)


def _header(cell) -> None:
    cell.font = _BOLD
    cell.fill = _FILL
    cell.border = _BORDER


def _write(ws, rows, start_row=1, start_col=1) -> None:
    for r, row in enumerate(rows, start=start_row):
        for c, value in enumerate(row, start=start_col):
            if value is not None:
                ws.cell(r, c, value)


# --------------------------------------------------------------------------- #
# builders
# --------------------------------------------------------------------------- #

def _single_row_header(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "T"
    _write(ws, [["項目ID", "項目名", "型"], ["F001", "契約番号", "string"], ["F002", "氏名", "string"], ["F003", "年齢", "number"]])
    for c in range(1, 4):
        _header(ws.cell(1, c))
    wb.save(path)


def _two_row_merged_header(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws.merge_cells("A1:B1"); ws["A1"] = "識別情報"; ws["C1"] = "説明"
    _write(ws, [["項目ID", "項目名", None]], start_row=2)
    ws["C2"] = "備考"
    _write(ws, [["F001", "契約番号", "主キー"], ["F002", "氏名", "-"]], start_row=3)
    for coord in ("A1", "C1", "A2", "B2", "C2"):
        _header(ws[coord])
    wb.save(path)


def _three_row_header(path: Path) -> None:
    # A genuine 3-row header: row1 has two merged groups (not a lone banner),
    # row2 sub-groups, row3 leaf labels, data from row4.
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws.merge_cells("A1:A3"); ws["A1"] = "No"
    ws.merge_cells("B1:E1"); ws["B1"] = "画面項目"
    ws.merge_cells("B2:C2"); ws["B2"] = "識別"
    ws.merge_cells("D2:E2"); ws["D2"] = "属性"
    _write(ws, [["ID", "名称", "型", "桁"]], start_row=3, start_col=2)
    _write(ws, [["1", "F001", "契約番号", "文字列", "20"]], start_row=4)
    for coord in ("A1", "B1", "B2", "D2", "B3", "C3", "D3", "E3"):
        _header(ws[coord])
    wb.save(path)


def _title_above_table(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "T"
    ws["A1"] = "画面項目一覧"; ws["A1"].font = _TITLE
    _write(ws, [["項目ID", "項目名", "型"]], start_row=2)
    _write(ws, [["F001", "契約番号", "string"], ["F002", "氏名", "string"]], start_row=3)
    for c in range(1, 4):
        _header(ws.cell(2, c))
    wb.save(path)


def _blank_row_in_table(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "T"
    _write(ws, [["ID", "名称"], ["1", "a"], ["2", "b"]])
    _write(ws, [["3", "c"], ["4", "d"]], start_row=5)  # row 4 blank
    for c in range(1, 3):
        _header(ws.cell(1, c))
    wb.save(path)


def _multi_table(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "T"
    _write(ws, [["ID", "名称"], ["1", "a"], ["2", "b"]])
    _write(ws, [["No", "内容"], ["1", "x"], ["2", "y"]], start_row=6)  # blank 4,5
    for c in range(1, 3):
        _header(ws.cell(1, c)); _header(ws.cell(6, c))
    wb.save(path)


def _horizontal_kv(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "表紙"
    _write(ws, [["画面ID", "SCR-001"], ["画面名", "利用者検索"], ["作成者", "山田"]])
    for r in range(1, 4):
        _header(ws.cell(r, 1))
    wb.save(path)


def _multi_kv_per_row(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "表紙"
    _write(ws, [["画面ID", "SCR-001", "版数", "1.0"], ["作成者", "山田", "承認者", "田中"]])
    for coord in ("A1", "C1", "A2", "C2"):
        _header(ws[coord])
    wb.save(path)


def _revision_history(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "改訂履歴"
    _write(ws, [["No", "日付", "版数", "内容", "担当"],
                ["1", "2026-01-01", "1.0", "新規作成", "山田"],
                ["2", "2026-02-01", "1.1", "項目追加", "田中"]])
    for c in range(1, 6):
        _header(ws.cell(1, c))
    wb.save(path)


def _screen_items(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "画面入出力項目一覧"
    ws.merge_cells("A1:A2"); ws["A1"] = "項目ID"
    ws.merge_cells("B1:B2"); ws["B1"] = "画面項目名"
    ws.merge_cells("C1:D1"); ws["C1"] = "属性"
    ws["C2"] = "型"; ws["D2"] = "桁数"
    ws.merge_cells("E1:E2"); ws["E1"] = "必須"
    _write(ws, [["F001", "契約番号", "文字列", "20", "○"],
                ["F002", "氏名", "文字列", "40", "○"]], start_row=3)
    for coord in ("A1", "B1", "C1", "C2", "D2", "E1"):
        _header(ws[coord])
    wb.save(path)


def _action_list(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "画面アクション一覧"
    _write(ws, [["No", "アクション", "イベント", "処理内容"],
                ["1", "検索", "click", "一覧を更新"],
                ["2", "クリア", "click", "条件をリセット"]])
    for c in range(1, 5):
        _header(ws.cell(1, c))
    wb.save(path)


def _graph_paper_layout(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "画面レイアウト"
    ws["A1"] = "画面レイアウト"; ws["A1"].font = _TITLE
    # a sparse merged box with borders, little text = visual layout
    ws.merge_cells("B3:H3"); ws["B3"] = "検索条件"
    ws.merge_cells("B10:H10"); ws["B10"] = "結果一覧"
    for row in range(3, 15):
        for col in range(2, 9):
            ws.cell(row, col).border = _BORDER
    wb.save(path)


def _cross_sheet_formula(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "表紙"
    base = wb.create_sheet("基本情報")
    base["B3"] = "SCR-DEMO-001"
    _write(ws, [["画面ID", "='基本情報'!B3"], ["版数", "=基本情報!B3"]])
    ws["A1"].font = _BOLD
    wb.save(path)


def _distant_style_pollution(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "T"
    _write(ws, [["ID", "名称"], ["1", "a"], ["2", "b"]])
    for c in range(1, 3):
        _header(ws.cell(1, c))
    ws.cell(row=5000, column=200).fill = _FILL  # distant style-only cell
    wb.save(path)


def _hidden_rows_and_sheet(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "T"
    _write(ws, [["ID", "名称"], ["1", "a"], ["2", "b"]])
    for c in range(1, 3):
        _header(ws.cell(1, c))
    ws.row_dimensions[2].hidden = True
    hidden = wb.create_sheet("隠しシート")
    hidden["A1"] = "内部メモ"
    hidden.sheet_state = "hidden"
    wb.save(path)


def build_cases() -> list[EvalCase]:
    return [
        EvalCase("single_row_header", "单行表头表格", _single_row_header,
                 [ExpectedRegion("T", "A1:C4", "table", header_rows=1, row_count=3)], tags=["header"]),
        EvalCase("two_row_merged_header", "两行合并表头", _two_row_merged_header,
                 [ExpectedRegion("T", "A1:C4", "table", header_rows=2, row_count=2)], tags=["header", "merge"]),
        EvalCase("three_row_header", "三行复杂表头", _three_row_header,
                 [ExpectedRegion("T", "A1:E4", "table", header_rows=3, row_count=1)], tags=["header", "merge"]),
        EvalCase("title_above_table", "表格前节区标题", _title_above_table,
                 [ExpectedRegion("T", "A2:C4", "table", header_rows=1, row_count=2, title="画面項目一覧")], tags=["title"]),
        EvalCase("blank_row_in_table", "表格中空行", _blank_row_in_table,
                 [ExpectedRegion("T", "A1:B6", "table", header_rows=1)], tags=["blank"]),
        EvalCase("multi_table", "同 Sheet 多表", _multi_table,
                 [ExpectedRegion("T", "A1:B3", "table"), ExpectedRegion("T", "A6:B8", "table")], tags=["multi"]),
        EvalCase("horizontal_kv", "横向 Key/Value", _horizontal_kv,
                 [ExpectedRegion("表紙", "A1:B3", "key_value")], tags=["kv"]),
        EvalCase("multi_kv_per_row", "一行多 Key/Value", _multi_kv_per_row,
                 [ExpectedRegion("表紙", "A1:D2", "key_value")], tags=["kv"]),
        EvalCase("revision_history", "修正履历", _revision_history,
                 [ExpectedRegion("改訂履歴", "A1:E3", "table", header_rows=1, row_count=2)], tags=["table"]),
        EvalCase("screen_items", "画面项目定义", _screen_items,
                 [ExpectedRegion("画面入出力項目一覧", "A1:E4", "table", header_rows=2, row_count=2)],
                 expected_fields={"画面入出力項目一覧": {"B": "field_name", "E": "required"}}, tags=["header", "merge", "field"]),
        EvalCase("action_list", "Action 列表", _action_list,
                 [ExpectedRegion("画面アクション一覧", "A1:D3", "table", header_rows=1, row_count=2)], tags=["table"]),
        EvalCase("graph_paper_layout", "方眼纸布局", _graph_paper_layout,
                 [ExpectedRegion("画面レイアウト", "A1:A1", "text"),
                  ExpectedRegion("画面レイアウト", "B3:H14", "layout")], tags=["layout"]),
        EvalCase("cross_sheet_formula", "跨 Sheet 引用", _cross_sheet_formula,
                 [ExpectedRegion("表紙", "A1:B2", "key_value"),
                  ExpectedRegion("基本情報", "B3:B3", "text")], expected_references=2, tags=["formula"]),
        EvalCase("distant_style_pollution", "远距离样式污染", _distant_style_pollution,
                 [ExpectedRegion("T", "A1:B3", "table")], tags=["stress"]),
        EvalCase("hidden_rows_and_sheet", "隐藏行/隐藏 Sheet", _hidden_rows_and_sheet,
                 [ExpectedRegion("T", "A1:B3", "table"),
                  ExpectedRegion("隠しシート", "A1:A1", "text")], tags=["hidden"]),
    ]


__all__ = ["build_cases"]
