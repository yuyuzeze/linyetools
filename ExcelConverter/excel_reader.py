from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def read_sheet_as_grid(worksheet: Worksheet) -> list[list[str]]:
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    grid: list[list[str]] = [
        [_cell_to_str(worksheet.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]

    for merged in worksheet.merged_cells.ranges:
        value = worksheet.cell(merged.min_row, merged.min_col).value
        text = _cell_to_str(value)
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                grid[r - 1][c - 1] = text

    return grid


def load_workbook_from_path(path: Path):
    return load_workbook(path, data_only=True)
