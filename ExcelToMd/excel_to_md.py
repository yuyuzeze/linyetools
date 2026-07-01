"""
Excel (.xlsx) → Markdown。

表格一律为 **GFM 管道表**（`| ... |`）。合并单元格仅在 **左上角一格** 写内容，同合并区内其它格导出为空，
避免整段文字在管道表里重复出现（不再使用 <table>/<tr>/<td>）。

跳过整表无文字行；剔除全空列；输出折叠连续空行。

默认读取文本框/形状内文字（顺序与箭头拓扑不保真）；可用 --no-shapes 关闭。
默认从嵌入图导出图片文件并在 Markdown 中直接引用（文件名形如「图1-1 xxx」）；可用 --no-export-images 关闭。

依赖：openpyxl；形状/图片解析用标准库 zipfile + xml.etree。
"""

from __future__ import annotations

import argparse
import html
import re
import sys
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import quote
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _sheet_bounds(ws: Worksheet) -> tuple[int, int, int, int]:
    min_row = ws.min_row or 1
    min_col = ws.min_column or 1
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    return min_row, min_col, max_row, max_col


def _merge_master_map(ws: Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    """合并区内每个坐标 → 该区左上角坐标。"""
    m: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        top = (mr.min_row, mr.min_col)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                m[(r, c)] = top
    return m


def _pipe_cell_value(
    ws: Worksheet, r: int, c: int, master_map: dict[tuple[int, int], tuple[int, int]]
) -> object:
    """
    管道表导出用：合并区只在左上角一格保留值，其余合并格导出为空（避免整段文字重复）。
    """
    top = master_map.get((r, c), (r, c))
    if (r, c) != top:
        return None
    return ws.cell(top[0], top[1]).value


def _value_is_blank(v: object) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    return False


def _row_has_visible_text(
    ws: Worksheet,
    r: int,
    min_col: int,
    max_col: int,
    master_map: dict[tuple[int, int], tuple[int, int]],
) -> bool:
    for c in range(min_col, max_col + 1):
        if not _value_is_blank(_pipe_cell_value(ws, r, c, master_map)):
            return True
    return False


def _format_pipe_cell(v: object) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    s = " ".join(line.strip() for line in s.split("\n") if line.strip())
    return s.replace("\\", "\\\\").replace("|", "\\|")


def _trim_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    """去掉在全部行中均为空的列（合并单元格产生的冗余物理列）。"""
    if not rows:
        return rows
    ncols = len(rows[0])
    keep = [c for c in range(ncols) if any(row[c].strip() for row in rows)]
    if not keep:
        return [[]]
    return [[row[c] for c in keep] for row in rows]


def _row_nonempty_cells(
    ws: Worksheet,
    r: int,
    min_col: int,
    max_col: int,
    master_map: dict[tuple[int, int], tuple[int, int]],
) -> list[str]:
    """一行内从左到右的非空单元格文本（合并格仅左上角）。"""
    out: list[str] = []
    for c in range(min_col, max_col + 1):
        v = _pipe_cell_value(ws, r, c, master_map)
        if not _value_is_blank(v):
            out.append(_format_pipe_cell(v))
    return out


_TABLE_HEADER_HINTS = (
    "画面項目",
    "項目名",
    "種別",
    "データ元",
    "初期値",
    "イベント",
    "必須",
    "活性",
    "補足説明",
    "チェック",
    "画面レイアウト",
)


def _is_likely_table_header_row(cells: list[str]) -> bool:
    """判断一行是否像主表表头（日语设计书常见列名）。"""
    if len(cells) < 3:
        return False
    if any(re.fullmatch(r"No\.?", t.strip(), re.I) for t in cells):
        return True
    if any("画面項目" in t or "項目名" in t for t in cells):
        return True
    hit = sum(1 for h in _TABLE_HEADER_HINTS if any(h in t for t in cells))
    return len(cells) >= 5 and hit >= 2


def _detect_main_table_start_row(
    ws: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    master_map: dict[tuple[int, int], tuple[int, int]],
) -> int | None:
    """返回主表表头行号（1-based）；无法识别时返回 None。"""
    for r in range(min_row, max_row + 1):
        if not _row_has_visible_text(ws, r, min_col, max_col, master_map):
            continue
        cells = _row_nonempty_cells(ws, r, min_col, max_col, master_map)
        if _is_likely_table_header_row(cells):
            return r

    counts: list[tuple[int, int]] = []
    for r in range(min_row, min(min_row + 30, max_row + 1)):
        if not _row_has_visible_text(ws, r, min_col, max_col, master_map):
            continue
        counts.append((r, len(_row_nonempty_cells(ws, r, min_col, max_col, master_map))))
    if not counts:
        return None

    median = sorted(n for _, n in counts)[len(counts) // 2]
    threshold = max(6, median + 3)
    for r, n in counts:
        if n >= threshold:
            follow = 0
            for r2 in range(r + 1, min(r + 4, max_row + 1)):
                if len(_row_nonempty_cells(ws, r2, min_col, max_col, master_map)) >= 2:
                    follow += 1
            if follow >= 1:
                return r
    return None


_KNOWN_METADATA_LABELS = frozenset({
    "プロダクト",
    "プロジェクト",
    "サブシステム",
    "画面名",
    "画面ID",
    "作成日",
    "作成者",
    "更新日",
    "更新者",
    "版数",
    "バージョン",
    "文書番号",
    "I/O",
    "表示",
    "活性",
    "必須",
})

_COVER_ATTR_LABELS = frozenset({
    "文書番号",
    "バージョン",
    "作成者",
    "作成日",
    "更新者",
    "更新日",
    "承認者",
    "承認日",
    "レビュー者",
    "レビュー日",
})

_DOC_TYPE_MARKERS = (
    "画面設計書",
    "詳細設計書",
    "基本設計書",
    "要件定義書",
)

_STANDALONE_TITLE_CELLS = frozenset({
    "基本設計",
    "詳細設計",
    "画面設計",
    "要件定義",
})


def _normalize_label(text: str) -> str:
    return text.strip().rstrip("：:")


def _cell_raw_value(
    ws: Worksheet,
    r: int,
    c: int,
    master_map: dict[tuple[int, int], tuple[int, int]],
) -> object:
    """合并区仅在左上角返回原值，其余格视为空（避免表紙情報重复）。"""
    return _pipe_cell_value(ws, r, c, master_map)


def _next_col_after_cell(ws: Worksheet, r: int, c: int) -> int:
    """返回单元格 (r,c) 所在合并区之后的第一列（无合并则 c+1）。"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
            return mr.max_col + 1
    return c + 1


def _format_metadata_value(v: object) -> str:
    """表紙メタデータ用の表示（日付は Excel 風 YYYY/M/D）。"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return f"{v.year}/{v.month}/{v.day}"
    if isinstance(v, date):
        return f"{v.year}/{v.month}/{v.day}"
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    s = " ".join(line.strip() for line in s.split("\n") if line.strip())
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{int(m.group(1))}/{int(m.group(2))}/{int(m.group(3))}"
    return s.replace("\\", "\\\\").replace("|", "\\|")


def _is_cover_attr_label(text: str) -> bool:
    return _normalize_label(text) in _COVER_ATTR_LABELS


def _looks_like_screen_id(text: str) -> bool:
    t = text.strip()
    return bool(re.fullmatch(r"SCR-A\d+", t, re.I))


def _is_doc_type_text(text: str) -> bool:
    t = text.strip()
    return any(t.endswith(m) or m == t for m in _DOC_TYPE_MARKERS)


def _is_cover_sheet(
    ws: Worksheet,
    master_map: dict[tuple[int, int], tuple[int, int]],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> bool:
    """表紙シート（縦積みタイトル + 文書属性表）を判定。"""
    name = ws.title.strip()
    if name in ("表紙", "封面", "Cover"):
        return True

    has_doc_type = False
    has_attr_table = False
    for r in range(min_row, max_row + 1):
        cells = _row_nonempty_cells(ws, r, min_col, max_col, master_map)
        if any(_is_doc_type_text(c) for c in cells):
            has_doc_type = True
        if len(cells) == 2 and _is_cover_attr_label(cells[0]):
            has_attr_table = True

    if has_doc_type and has_attr_table:
        return (
            _detect_main_table_start_row(
                ws, min_row, max_row, min_col, max_col, master_map
            )
            is None
        )
    return False


def _classify_cover_line(text: str) -> str:
    """表紙タイトルブロック内の行種別。"""
    t = text.strip()
    if t in _STANDALONE_TITLE_CELLS:
        return "phase"
    if _looks_like_screen_id(t):
        return "screen_id"
    if _is_doc_type_text(t):
        return "doc_type"
    if len(t) > 30:
        return "project"
    if re.search(r"[_＿]", t) or "平時" in t:
        return "subsystem"
    return "screen_name"


def cover_sheet_to_markdown(ws: Worksheet) -> str:
    """
    表紙シート専用レイアウト:
      プロジェクト帯 → サブシステム → 設計フェーズ → 画面ID → 画面名 → 書類種別
      下部: 文書番号 / バージョン / 作成者 / 作成日 などの 2 列属性表
    """
    min_row, min_col, max_row, max_col = _sheet_bounds(ws)
    master_map = _merge_master_map(ws)

    preamble: list[str] = []
    attrs: list[tuple[str, str]] = []
    seen_attr: set[str] = set()

    for r in range(min_row, max_row + 1):
        cells = _row_nonempty_cells(ws, r, min_col, max_col, master_map)
        if not cells:
            continue
        if len(cells) == 2 and _is_cover_attr_label(cells[0]):
            label = _normalize_label(cells[0])
            if label not in seen_attr:
                seen_attr.add(label)
                attrs.append((label, cells[1]))
        elif len(cells) == 1:
            text = cells[0].strip()
            if text and (not preamble or preamble[-1] != text):
                preamble.append(text)

    fields: dict[str, str] = {}
    screen_name: str | None = None
    doc_type: str | None = None

    for text in preamble:
        role = _classify_cover_line(text)
        if role == "doc_type":
            doc_type = text
        elif role == "screen_id":
            fields["画面ID"] = text
        elif role == "phase":
            fields["設計フェーズ"] = text
        elif role == "project":
            fields["プロジェクト"] = text
        elif role == "subsystem":
            fields["サブシステム"] = text
        elif role == "screen_name" and screen_name is None:
            screen_name = text

    lines: list[str] = ["### 表紙\n\n"]
    if screen_name and doc_type:
        lines.append(f"**{screen_name} — {doc_type}**\n\n")
    elif doc_type:
        lines.append(f"**{doc_type}**\n\n")
    elif screen_name:
        lines.append(f"**{screen_name}**\n\n")

    info_order = ("プロジェクト", "サブシステム", "設計フェーズ", "画面ID")
    info_rows = [(k, fields[k]) for k in info_order if k in fields]
    if info_rows:
        lines.append("#### プロジェクト情報\n\n")
        lines.append("| 区分 | 内容 |\n| --- | --- |\n")
        for k, v in info_rows:
            lines.append(f"| {k} | {v} |\n")
        lines.append("\n")

    if attrs:
        lines.append("#### 文書属性\n\n")
        lines.append("| 項目 | 内容 |\n| --- | --- |\n")
        for k, v in attrs:
            lines.append(f"| {k} | {v} |\n")
        lines.append("\n")

    return "".join(lines) if len(lines) > 1 else "_（空）_\n"


def _is_known_metadata_label(text: str) -> bool:
    return _normalize_label(text) in _KNOWN_METADATA_LABELS


def _is_standalone_metadata_cell(text: str) -> bool:
    t = text.strip()
    if t.startswith("■") or t.startswith("※"):
        return True
    if t in _STANDALONE_TITLE_CELLS:
        return True
    if "一覧" in t and len(t) >= 6:
        return True
    if t == "凡例":
        return True
    return False


def _extract_row_metadata_entries(
    ws: Worksheet,
    r: int,
    min_col: int,
    max_col: int,
    master_map: dict[tuple[int, int], tuple[int, int]],
) -> list[tuple[str, ...]]:
    """
    行を左から走査し、既知ラベルは右隣の非空セルを値としてペア化。
    戻り値: ('standalone', text) または ('kv', label, value)
    """
    entries: list[tuple[str, ...]] = []
    c = min_col
    while c <= max_col:
        raw = _cell_raw_value(ws, r, c, master_map)
        if _value_is_blank(raw):
            c += 1
            continue

        text = _format_metadata_value(raw)
        if _is_standalone_metadata_cell(text):
            entries.append(("standalone", text))
            c = _next_col_after_cell(ws, r, c)
            continue

        if _is_known_metadata_label(text):
            label = _normalize_label(text)
            value_text: str | None = None
            value_col = c
            search = _next_col_after_cell(ws, r, c)
            for c2 in range(search, max_col + 1):
                raw2 = _cell_raw_value(ws, r, c2, master_map)
                if _value_is_blank(raw2):
                    continue
                candidate = _format_metadata_value(raw2)
                if _is_known_metadata_label(candidate):
                    break
                value_text = candidate
                value_col = c2
                break
            if value_text is not None:
                entries.append(("kv", label, value_text))
                c = _next_col_after_cell(ws, r, value_col)
            else:
                entries.append(("standalone", text))
                c = _next_col_after_cell(ws, r, c)
            continue

        entries.append(("standalone", text))
        c = _next_col_after_cell(ws, r, c)
    return entries


def _metadata_to_markdown(
    ws: Worksheet,
    min_row: int,
    table_start_row: int,
    min_col: int,
    max_col: int,
    master_map: dict[tuple[int, int], tuple[int, int]],
) -> str:
    """表紙メタデータ区 → key-value リスト（セル位置に基づくペア）。"""
    lines: list[str] = []
    seen_kv: set[str] = set()
    seen_standalone: set[str] = set()
    seen_legend: set[str] = set()

    for r in range(min_row, table_start_row):
        if not _row_has_visible_text(ws, r, min_col, max_col, master_map):
            continue
        entries = _extract_row_metadata_entries(ws, r, min_col, max_col, master_map)
        if not entries:
            continue

        legend_bits: list[str] = []
        seen_bit: set[str] = set()
        for entry in entries:
            if entry[0] == "kv":
                label = entry[1]
                if label in seen_kv:
                    continue
                seen_kv.add(label)
                lines.append(f"- **{label}**: {entry[2]}\n")
            else:
                text = entry[1]
                if text.startswith("■") or text in _STANDALONE_TITLE_CELLS or (
                    "一覧" in text and len(text) >= 6
                ):
                    if text in seen_standalone:
                        continue
                    seen_standalone.add(text)
                    lines.append(f"**{text}**\n")
                elif "：" in text or ":" in text:
                    if text not in seen_bit:
                        seen_bit.add(text)
                        legend_bits.append(text)
                elif text.startswith("※"):
                    if text in seen_standalone:
                        continue
                    seen_standalone.add(text)
                    lines.append(f"- {text}\n")
                else:
                    if text not in seen_bit:
                        seen_bit.add(text)
                        legend_bits.append(text)

        if legend_bits:
            legend_line = f"- {' | '.join(legend_bits)}\n"
            if legend_line not in seen_legend:
                seen_legend.add(legend_line)
                lines.append(legend_line)

    if not lines:
        return ""
    return "### 表紙情報\n\n" + "".join(lines) + "\n"


def _collect_pipe_rows(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    min_col: int,
    max_col: int,
    master_map: dict[tuple[int, int], tuple[int, int]],
) -> list[list[str]]:
    raw_rows: list[list[str]] = []
    for r in range(start_row, end_row + 1):
        if not _row_has_visible_text(ws, r, min_col, max_col, master_map):
            continue
        raw_rows.append(
            [
                _format_pipe_cell(_pipe_cell_value(ws, r, c, master_map))
                for c in range(min_col, max_col + 1)
            ]
        )
    return raw_rows


def _rows_to_pipe_markdown(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    trimmed = _trim_empty_columns(rows)
    if not trimmed or not trimmed[0]:
        return ""
    body_lines = ["| " + " | ".join(cells) + " |" for cells in trimmed]
    ncol_out = len(trimmed[0])
    sep = "| " + " | ".join(["---"] * ncol_out) + " |"
    return "\n".join([body_lines[0], sep] + body_lines[1:])


def sheet_to_markdown_content(ws: Worksheet) -> str:
    """表紙 / 表紙情報 + 一覧；シート種別を自動判定。"""
    min_row, min_col, max_row, max_col = _sheet_bounds(ws)
    if max_col < min_col:
        return "_（空）_"

    master_map = _merge_master_map(ws)
    if _is_cover_sheet(ws, master_map, min_row, max_row, min_col, max_col):
        return cover_sheet_to_markdown(ws)
    table_start = _detect_main_table_start_row(
        ws, min_row, max_row, min_col, max_col, master_map
    )

    parts: list[str] = []
    if table_start is not None and table_start > min_row:
        meta = _metadata_to_markdown(
            ws, min_row, table_start, min_col, max_col, master_map
        )
        if meta:
            parts.append(meta)
        data_start = table_start
    else:
        data_start = min_row

    table_md = _rows_to_pipe_markdown(
        _collect_pipe_rows(ws, data_start, max_row, min_col, max_col, master_map)
    )
    if table_md:
        if parts:
            parts.append("### 一覧\n\n")
        parts.append(table_md)

    if not parts:
        return "_（空）_"
    return "".join(parts)


def sheet_to_pipe_markdown_table(ws: Worksheet) -> str:
    """兼容旧名；等同 sheet_to_markdown_content。"""
    return sheet_to_markdown_content(ws)


def collapse_extra_blank_lines(text: str) -> str:
    """去掉连续空行，只保留至多一个换行。"""
    lines = text.splitlines()
    out: list[str] = []
    prev_blank = False
    for line in lines:
        blank = line.strip() == ""
        if blank and prev_blank:
            continue
        out.append(line)
        prev_blank = blank
    return "\n".join(out).strip() + "\n"


def _local_tag(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _drawing_zip_path(target: str) -> str:
    """sheet rels 中 Target 多为 ../drawings/drawing1.xml。"""
    target = target.replace("\\", "/")
    name = target.split("/")[-1]
    return f"xl/drawings/{name}"


def _drawing_rels_zip_path(drawing_part: str) -> str:
    """xl/drawings/drawing1.xml -> xl/drawings/_rels/drawing1.xml.rels"""
    name = drawing_part.rsplit("/", 1)[-1]
    return f"xl/drawings/_rels/{name}.rels"


def _resolve_media_zip_path(drawing_part: str, target: str) -> str:
    """把 drawing 的 rel Target（如 ../media/image1.png）解析为 zip 内路径。"""
    target = target.replace("\\", "/").lstrip("/")
    if target.startswith("xl/"):
        return target
    base = drawing_part.rsplit("/", 1)[0]
    parts = base.split("/")
    for piece in target.split("/"):
        if piece == "..":
            if parts:
                parts.pop()
        elif piece and piece != ".":
            parts.append(piece)
    return "/".join(parts)


def _md_uri_path(relative_posix: str) -> str:
    """Markdown 链接里用的相对 URI（含空格等则分段编码）。"""
    parts = [p for p in relative_posix.replace("\\", "/").split("/") if p != ""]
    return "/".join(quote(p, safe="") for p in parts)


def _load_rels_id_to_target(zf: zipfile.ZipFile, rels_part: str) -> dict[str, str]:
    if rels_part not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read(rels_part))
    m: dict[str, str] = {}
    for rel in root:
        if _local_tag(rel.tag) != "Relationship":
            continue
        rid, tgt = rel.attrib.get("Id"), rel.attrib.get("Target")
        if rid and tgt:
            m[rid] = tgt
    return m


def _anchor_sort_tuple(anchor_el: ET.Element) -> tuple[int, int]:
    """用锚点左上角行列排序（无则极大值）。"""
    for child in anchor_el:
        if _local_tag(child.tag) != "from":
            continue
        row, col = 10**9, 10**9
        for sub in child:
            t = _local_tag(sub.tag)
            if t == "row" and sub.text is not None and sub.text.strip().isdigit():
                row = int(sub.text)
            elif t == "col" and sub.text is not None and sub.text.strip().isdigit():
                col = int(sub.text)
        return (row, col)
    return (10**9, 10**9)


def _pic_embed_rid(pic_el: ET.Element) -> str | None:
    r_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
    for el in pic_el.iter():
        if _local_tag(el.tag) == "blip":
            rid = el.attrib.get(r_ns)
            if rid:
                return rid
    return None


def _pic_caption_from_excel(pic_el: ET.Element) -> str:
    """优先用图片「说明」再「名称」；排除默认 Picture N。"""
    name, descr = "", ""
    for el in pic_el.iter():
        if _local_tag(el.tag) == "cNvPr":
            name = (el.attrib.get("name") or "").strip()
            descr = (el.attrib.get("descr") or "").strip()
            break
    for candidate in (descr, name):
        if not candidate:
            continue
        if re.match(r"^Picture\s*\d+$", candidate, re.I):
            continue
        return candidate
    return ""


def _sanitize_figure_label(s: str, max_len: int = 120) -> str:
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:max_len] if s else ""


def _unique_stem(figures_dir: Path, base: str, ext: str) -> str:
    """base 不含扩展名；返回可用文件名（不含路径） base.ext 或 base_2.ext。"""
    stem = base + ext
    if not (figures_dir / stem).exists():
        return stem
    for n in range(2, 10_000):
        alt = f"{base}_{n}{ext}"
        if not (figures_dir / alt).exists():
            return alt
    return base + ext


def export_sheet_embedded_images(
    xlsx_path: Path,
    figures_dir: Path,
    sheet_index: int,
) -> tuple[list[Path], str]:
    """
    导出当前 sheet 的嵌入位图文件。
    文件名：图{sheet_index}-{序号} [Excel中图片说明/名称].ext
    返回 (生成的图片路径列表, 可追加到 md 的 Markdown 片段)。
    """
    figures_dir.mkdir(parents=True, exist_ok=True)
    rel_fig = figures_dir.name
    img_paths: list[Path] = []
    md_lines: list[str] = []

    r_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    sheet_part = f"xl/worksheets/sheet{sheet_index}.xml"
    rels_part = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        if sheet_part not in zf.namelist() or rels_part not in zf.namelist():
            return [], ""

        sheet_root = ET.fromstring(zf.read(sheet_part))
        r_ids: list[str] = []
        for el in sheet_root.iter():
            if _local_tag(el.tag) == "drawing":
                rid = el.attrib.get(r_ns)
                if rid:
                    r_ids.append(rid)

        sheet_rid_to_target = _load_rels_id_to_target(zf, rels_part)
        drawing_parts = [
            _drawing_zip_path(sheet_rid_to_target[r])
            for r in r_ids
            if r in sheet_rid_to_target
        ]

        entries: list[tuple[tuple[int, int], int, str, str, str]] = []
        # (sort_row, sort_col), order, drawing_part, embed_rid, caption
        seq = 0
        for dp in drawing_parts:
            if dp not in zf.namelist():
                continue
            droot = ET.fromstring(zf.read(dp))
            drawing_rels = _drawing_rels_zip_path(dp)
            rid_to_media = _load_rels_id_to_target(zf, drawing_rels)

            for anchor in droot:
                if _local_tag(anchor.tag) not in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
                    continue
                sort_key = _anchor_sort_tuple(anchor)
                for child in anchor:
                    if _local_tag(child.tag) != "pic":
                        continue
                    rid = _pic_embed_rid(child)
                    if not rid or rid not in rid_to_media:
                        continue
                    target = rid_to_media[rid]
                    media_zip = _resolve_media_zip_path(dp, target)
                    if media_zip not in zf.namelist():
                        continue
                    cap = _pic_caption_from_excel(child)
                    entries.append((sort_key, seq, media_zip, rid, cap))
                    seq += 1

        entries.sort(key=lambda t: (t[0][0], t[0][1], t[1]))

        for idx, (_sk, _seq, media_zip, _rid, caption) in enumerate(entries, start=1):
            raw = zf.read(media_zip)
            ext = Path(media_zip).suffix.lower() or ".bin"
            if ext not in (".png", ".jpeg", ".jpg", ".gif", ".bmp", ".tif", ".tiff", ".wmf", ".emf"):
                ext = ".bin"

            label = _sanitize_figure_label(caption)
            if label:
                base = f"図{sheet_index}-{idx} {label}"
            else:
                base = f"図{sheet_index}-{idx}"
            base = base[:120].rstrip(" .")

            img_name = _unique_stem(figures_dir, base, ext)
            img_path = figures_dir / img_name
            img_path.write_bytes(raw)
            img_paths.append(img_path)

            alt = label or f"図{sheet_index}-{idx}"
            href = _md_uri_path(f"{rel_fig}/{img_name}")
            md_lines.append(f"![{alt}]({href})\n")

    if not md_lines:
        return [], ""

    appendix = (
        f"\n### 埋め込み画像（{rel_fig}/）\n\n"
        + "".join(md_lines)
        + "\n"
    )
    return img_paths, appendix


def extract_shape_texts_from_xlsx(xlsx_path: Path, sheet_index: int) -> list[str]:
    """
    从 DrawingML 收集 a:t 文本；去重、顺序不保证。
    sheet_index 为 1-based，与 xl/worksheets/sheet{N}.xml 一致。
    """
    texts: list[str] = []
    seen: set[str] = set()
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        sheet_part = f"xl/worksheets/sheet{sheet_index}.xml"
        rels_part = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if sheet_part not in zf.namelist() or rels_part not in zf.namelist():
            return []

        root = ET.fromstring(zf.read(sheet_part))
        r_ids: list[str] = []
        for el in root.iter():
            if _local_tag(el.tag) == "drawing":
                rid = el.attrib.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                if rid:
                    r_ids.append(rid)

        rels_root = ET.fromstring(zf.read(rels_part))
        rid_to_target: dict[str, str] = {}
        for rel in rels_root:
            if _local_tag(rel.tag) != "Relationship":
                continue
            rid, tgt = rel.attrib.get("Id"), rel.attrib.get("Target")
            if rid and tgt:
                rid_to_target[rid] = tgt

        drawing_paths = [_drawing_zip_path(rid_to_target[r]) for r in r_ids if r in rid_to_target]

        for dp in drawing_paths:
            if dp not in zf.namelist():
                continue
            droot = ET.fromstring(zf.read(dp))
            for el in droot.iter():
                if el.tag == f"{{{ns_a}}}t" and el.text:
                    t = el.text.strip()
                    if t and t not in seen:
                        seen.add(t)
                        texts.append(t)

    return texts


def build_markdown_for_sheet(
    ws: Worksheet,
    *,
    workbook_path: Path,
    sheet_index: int,
    include_shapes: bool,
    figures_appendix: str = "",
) -> str:
    parts: list[str] = [f"## {ws.title}\n"]
    parts.append(
        "<!-- "
        f"source: {workbook_path.name} | sheet: {sheet_index} | "
        f"exported: {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}"
        " -->\n"
    )

    parts.append(sheet_to_markdown_content(ws))
    parts.append("\n")

    if include_shapes:
        shape_lines = extract_shape_texts_from_xlsx(workbook_path, sheet_index)
        if shape_lines:
            parts.append("\n### シェイプ内テキスト（DrawingML、順序は保証されません）\n")
            for line in shape_lines:
                parts.append(f"- {html.escape(line)}\n")

    if figures_appendix:
        parts.append(figures_appendix)

    return collapse_extra_blank_lines("".join(parts))


def convert_workbook(
    xlsx_path: Path,
    out_dir: Path,
    *,
    one_file: bool,
    include_shapes: bool,
    export_images: bool,
) -> list[Path]:
    xlsx_path = xlsx_path.resolve()
    out_dir = out_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    written: list[Path] = []
    figures_dir = out_dir / f"{xlsx_path.stem}_images" if export_images else None

    if one_file:
        chunks = [f"# {xlsx_path.stem}\n\n"]
        for i, ws in enumerate(wb.worksheets, start=1):
            fig_append = ""
            fig_imgs: list[Path] = []
            if figures_dir is not None:
                fig_imgs, fig_append = export_sheet_embedded_images(xlsx_path, figures_dir, i)
                written.extend(fig_imgs)
            chunks.append(
                build_markdown_for_sheet(
                    ws,
                    workbook_path=xlsx_path,
                    sheet_index=i,
                    include_shapes=include_shapes,
                    figures_appendix=fig_append,
                )
            )
            chunks.append("\n")
        out_path = out_dir / f"{xlsx_path.stem}.md"
        out_path.write_text(collapse_extra_blank_lines("".join(chunks)), encoding="utf-8")
        written.append(out_path)
    else:
        for i, ws in enumerate(wb.worksheets, start=1):
            fig_append = ""
            fig_imgs: list[Path] = []
            if figures_dir is not None:
                fig_imgs, fig_append = export_sheet_embedded_images(xlsx_path, figures_dir, i)
                written.extend(fig_imgs)
            safe = re.sub(r'[<>:"/\\|?*]', "_", ws.title) or f"sheet{i}"
            out_path = out_dir / f"{xlsx_path.stem}__{safe}.md"
            body = build_markdown_for_sheet(
                ws,
                workbook_path=xlsx_path,
                sheet_index=i,
                include_shapes=include_shapes,
                figures_appendix=fig_append,
            )
            out_path.write_text(
                collapse_extra_blank_lines(f"# {ws.title}\n\n{body}"),
                encoding="utf-8",
            )
            written.append(out_path)

    wb.close()
    return written


def _collect_xlsx_files(path: Path, *, recursive: bool = False) -> list[Path]:
    """收集待转换的 .xlsx（跳过 Excel 临时文件 ~$*.xlsx）。"""
    path = path.resolve()
    if path.is_file():
        return [path]
    if not path.is_dir():
        return []
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    files = [
        f
        for f in sorted(path.glob(pattern))
        if f.is_file() and not f.name.startswith("~$")
    ]
    return files


def _batch_output_dir(out_dir: Path, input_root: Path, xlsx_path: Path) -> Path:
    """批量模式下为每个 xlsx 生成独立输出子目录。"""
    try:
        rel_parent = xlsx_path.parent.relative_to(input_root)
    except ValueError:
        rel_parent = Path(".")
    if rel_parent == Path("."):
        return out_dir / xlsx_path.stem
    safe_parts = [re.sub(r'[<>:"/\\|?*]', "_", p) for p in rel_parent.parts]
    return out_dir / Path(*safe_parts) / xlsx_path.stem


def convert_path(
    input_path: Path,
    out_dir: Path,
    *,
    one_file: bool,
    include_shapes: bool,
    export_images: bool,
    recursive: bool = False,
) -> list[Path]:
    """
    单文件：直接写入 out_dir。
    文件夹：out_dir 下为每个 xlsx 创建独立子文件夹并分别转换。
    """
    input_path = input_path.resolve()
    out_dir = out_dir.resolve()
    xlsx_files = _collect_xlsx_files(input_path, recursive=recursive)
    if not xlsx_files:
        return []

    written: list[Path] = []
    batch_mode = input_path.is_dir()
    for xlsx_path in xlsx_files:
        target_dir = (
            _batch_output_dir(out_dir, input_path, xlsx_path)
            if batch_mode
            else out_dir
        )
        written.extend(
            convert_workbook(
                xlsx_path,
                target_dir,
                one_file=one_file,
                include_shapes=include_shapes,
                export_images=export_images,
            )
        )
    return written


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="excel_to_md",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description="将 .xlsx 转为 Markdown：GFM 管道表（合并格仅左上角有值，自动剔除全空行列）；默认导出嵌入图并在 md 中直接引用。UTF-8。",
        epilog="""
默认行为：
  导出单元格网格为管道表，自动去掉合并产生的全空行/列。
  从 DrawingML 抽取形状/文本框内可见文字（列表形式；顺序可能与画面不一致，箭头拓扑不保真）。
  导出嵌入图片到「工作簿名_images/」，在 .md 末尾以 Markdown 图片语法直接引用。

可选关闭：
  --no-shapes         不解析形状/文本框内文字。
  --no-export-images  不导出嵌入图片。

依赖：仅 pip 安装 openpyxl；不调用 LibreOffice 或 Excel。
""".strip(),
    )
    parser.add_argument(
        "input",
        type=Path,
        help="入力 .xlsx ファイル、または .xlsx を含むフォルダ",
    )
    parser.add_argument("-o", "--output-dir", type=Path, required=True, help="出力ディレクトリ（自動作成）")
    parser.add_argument("--one-file", action="store_true", help="全シートを 1 つの .md にまとめる")
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="フォルダ入力時、サブフォルダ内の .xlsx も再帰的に変換",
    )
    parser.add_argument(
        "--no-shapes",
        action="store_true",
        help="不解析形状/文本框内文字",
    )
    parser.add_argument(
        "--no-export-images",
        action="store_true",
        help="不导出嵌入图片",
    )
    args = parser.parse_args(list(argv) if argv is not None else None)

    input_path = args.input.resolve()
    if not input_path.exists():
        print(f"エラー：パスが存在しません {input_path}", file=sys.stderr)
        return 2

    if input_path.is_file():
        if input_path.suffix.lower() != ".xlsx":
            print("エラー：.xlsx のみ対応しています", file=sys.stderr)
            return 2
    elif not input_path.is_dir():
        print(f"エラー：ファイルまたはフォルダを指定してください {input_path}", file=sys.stderr)
        return 2

    written = convert_path(
        input_path,
        args.output_dir,
        one_file=args.one_file,
        include_shapes=not args.no_shapes,
        export_images=not args.no_export_images,
        recursive=args.recursive,
    )
    if not written:
        print(f"エラー：変換対象の .xlsx が見つかりません {input_path}", file=sys.stderr)
        return 2

    for path in written:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
